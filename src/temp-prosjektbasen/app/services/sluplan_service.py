from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import joinedload
from zoneinfo import ZoneInfo

from app.models.db import db
from app.models.project import Project
from app.models.user import User
from app.models.fag import Fag
from app.models.sluplan import (
    SluplanComment,
    SluplanDependency,
    SluplanDiscipline,
    SluplanFile,
    SluplanPerson,
    SluplanProject,
    SluplanTask,
)

__all__ = [
    "list_projects",
    "list_base_projects",
    "create_project",
    "get_project_snapshot",
    "list_resource_catalog",
    "get_tasks",
    "get_plan_snapshot",
    "export_plan",
    "import_plan",
    "create_task",
    "update_task_fields",
    "reset_plan",
    "add_comment_to_task",
    "add_file_attachment",
    "generate_task_ics",
    "import_systems_from_excel",
    "get_alerts",
    "summarize_by_discipline",
    "summarize_by_user",
    "summarize_by_time_window",
]

MENTION_PATTERN = re.compile(r"@([\w\-æøåÆØÅ]+)")

STANDARD_MILESTONES = [
    "Mechanical Complete",
    "Functional Test",
    "Integrated Test",
    "Full-Scale Test",
    "Complete Handover Documentation (FDV)",
]

STANDARD_SYSTEM_SUBTASKS = [
    "MC Complete",
    "SD/Software Ready",
    "Protocols Complete",
    "Self-Test",
    "Commissioning",
    "Functional Test (BH and TE)",
]


# ---------------------------------------------------------------------------
# Serializers & helpers
# ---------------------------------------------------------------------------


def _iso_date(value: date) -> str:
    return value.isoformat()


def _parse_iso_date(raw: str) -> date:
    year, month, day = map(int, raw.split("-"))
    return date(year, month, day)


def _parse_iso_datetime(raw: Optional[str]) -> Optional[datetime]:
    if not raw:
        return None
    try:
        normalized = raw.replace("Z", "+00:00")
        dt = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if dt.tzinfo:
        return dt.astimezone(timezone.utc)
    return dt.replace(tzinfo=timezone.utc)


def _to_utc_iso(value: Optional[datetime]) -> Optional[str]:
    if not value:
        return None
    if value.tzinfo:
        value = value.astimezone(timezone.utc)
    else:
        value = value.replace(tzinfo=timezone.utc)
    return value.isoformat()


def _serialize_project(project: SluplanProject) -> Dict[str, Any]:
    base_project = project.project
    resolved_name = project.name
    resolved_order = project.order_number
    resolved_location = None

    if base_project:
        base_name = (base_project.project_name or "").strip()
        base_order = (base_project.order_number or "").strip()
        resolved_location = base_project.location

        if base_name:
            resolved_name = base_name
        if base_order:
            resolved_order = base_order

    return {
        "id": project.id,
        "name": resolved_name,
        "order_number": resolved_order,
        "start_date": _iso_date(project.start_date),
        "end_date": _iso_date(project.end_date),
        "project_id": base_project.id if base_project else None,
        "project_name": base_project.project_name if base_project else None,
        "location": resolved_location,
        "created_at": _to_utc_iso(project.created_at),
        "updated_at": _to_utc_iso(project.updated_at),
    }


def _serialize_comment(comment: SluplanComment) -> Dict[str, Any]:
    mentions = comment.mentions.split(",") if comment.mentions else []
    return {
        "id": str(comment.id),
        "text": comment.text,
        "author": comment.author,
        "created_at": _to_utc_iso(comment.created_at),
        "mentions": mentions,
    }


def _serialize_file(attachment: SluplanFile) -> Dict[str, Any]:
    return {
        "id": str(attachment.id),
        "filename": attachment.filename,
        "size": attachment.size,
        "content_type": attachment.content_type,
        "uploaded_at": _to_utc_iso(attachment.uploaded_at),
    }


def _serialize_task(task: SluplanTask, children: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    return {
        "id": str(task.id),
        "title": task.title,
        "name": task.title,
        "start": _iso_date(task.start_date),
        "end": _iso_date(task.end_date),
        "status": task.status,
        "kind": task.kind,
        "assignee": task.assignee.name if task.assignee else None,
        "assignee_email": task.assignee.email if task.assignee else None,
        "discipline": task.discipline.name if task.discipline else None,
        "children": children if children is not None else [],
        "comments": [_serialize_comment(comment) for comment in task.comments or []],
        "files": [_serialize_file(attachment) for attachment in task.files or []],
    }


def _serialize_dependency(dep: SluplanDependency) -> Dict[str, Any]:
    return {
        "id": str(dep.id),
        "fromId": str(dep.from_task_id),
        "toId": str(dep.to_task_id),
        "type": dep.type or "FS",
    }


# ---------------------------------------------------------------------------
# Project helpers
# ---------------------------------------------------------------------------


def _upcoming_monday(reference: Optional[date] = None) -> date:
    base = reference or date.today()
    offset = (7 - base.weekday()) % 7
    return base + timedelta(days=offset)


def _ensure_default_project() -> SluplanProject:
    project = db.session.execute(select(SluplanProject).order_by(SluplanProject.created_at)).scalar_one_or_none()
    if project:
        return project

    base_project = db.session.execute(select(Project).order_by(Project.created_at)).scalar_one_or_none()
    start = _upcoming_monday()
    end = start + timedelta(days=30)
    project = SluplanProject(
        name=base_project.project_name if base_project else "Standardplan",
        order_number=base_project.order_number if base_project else None,
        project=base_project,
        start_date=start,
        end_date=end,
    )
    db.session.add(project)
    db.session.flush()
    _seed_default_plan(project, systems=None)
    db.session.commit()
    return project


def _project_by_id(project_id: Optional[int]) -> SluplanProject:
    if project_id is None:
        return _ensure_default_project()
    project = db.session.get(SluplanProject, int(project_id))
    if not project:
        raise ValueError("project_finnes_ikke")
    return project


# ---------------------------------------------------------------------------
# Data building
# ---------------------------------------------------------------------------


def _load_tasks_for_project(project: SluplanProject) -> List[SluplanTask]:
    statement = (
        select(SluplanTask)
        .options(
            joinedload(SluplanTask.assignee),
            joinedload(SluplanTask.discipline),
            joinedload(SluplanTask.comments),
            joinedload(SluplanTask.files),
        )
        .where(SluplanTask.project_id == project.id)
        .order_by(SluplanTask.parent_id, SluplanTask.start_date, SluplanTask.id)
    )
    return list(db.session.execute(statement).scalars())


def _build_task_tree(tasks: Iterable[SluplanTask]) -> List[Dict[str, Any]]:
    grouped: Dict[Optional[int], List[SluplanTask]] = {}
    for task in tasks:
        grouped.setdefault(task.parent_id, []).append(task)

    for children in grouped.values():
        children.sort(key=lambda item: (item.start_date, item.id))

    def _build(node: SluplanTask) -> Dict[str, Any]:
        child_nodes = grouped.get(node.id, [])
        children = [_build(child) for child in child_nodes]
        return _serialize_task(node, children)

    roots = grouped.get(None, [])
    return [_build(task) for task in roots]


def _status_bucket(status: Optional[str]) -> str:
    if not status:
        return "planned"
    normalized = status.strip().lower()
    if normalized in {"ferdig", "done", "completed"}:
        return "completed"
    if normalized in {"pågår", "in progress", "in_progress", "progress"}:
        return "in_progress"
    return "planned"


def _task_duration_days(task: SluplanTask) -> int:
    delta = (task.end_date - task.start_date).days
    return max(0, delta) + 1


def _duration_from_dates(start: date, end: date) -> int:
    delta = (end - start).days
    return max(0, delta) + 1


def _ensure_person(name: Optional[str], email: Optional[str] = None) -> Optional[SluplanPerson]:
    if not name and not email:
        return None

    name = (name or "").strip()
    email = (email or "").strip()

    query = select(SluplanPerson)
    if email:
        query = query.where(func.lower(SluplanPerson.email) == email.lower())
    elif name:
        query = query.where(func.lower(SluplanPerson.name) == name.lower())

    user: Optional[User] = None
    if email:
        user = db.session.execute(select(User).where(func.lower(User.email) == email.lower())).scalar_one_or_none()
    elif name:
        parts = name.split()
        if len(parts) >= 2:
            first, last = parts[0], parts[-1]
            user = db.session.execute(
                select(User).where(
                    func.lower(User.first_name) == first.lower(),
                    func.lower(User.last_name) == last.lower(),
                )
            ).scalar_one_or_none()

    if not name and user:
        name = (user.full_name() or "").strip()

    person = db.session.execute(query).scalar_one_or_none()
    if person:
        updated = False
        if user and person.user_id != user.id:
            person.user_id = user.id
            updated = True
        if email and person.email != email:
            person.email = email
            updated = True
        target_name = (user.full_name() if user else name) or ""
        target_name = target_name.strip()
        if target_name and person.name != target_name:
            person.name = target_name
            updated = True
        if updated:
            db.session.add(person)
            db.session.flush()
        return person

    person = SluplanPerson(
        name=name or email or "Ukjent",
        email=email or (user.email if user else None),
        user_id=user.id if user else None,
    )
    db.session.add(person)
    db.session.flush()
    return person


def _ensure_discipline(name: Optional[str]) -> Optional[SluplanDiscipline]:
    if not name:
        return None
    cleaned = name.strip()
    if not cleaned:
        return None
    discipline = db.session.execute(
        select(SluplanDiscipline).where(func.lower(SluplanDiscipline.name) == cleaned.lower())
    ).scalar_one_or_none()
    if discipline:
        return discipline
    discipline = SluplanDiscipline(name=cleaned)
    db.session.add(discipline)
    db.session.flush()
    return discipline


def _create_task(
    project: SluplanProject,
    title: str,
    start_day: date,
    end_day: date,
    *,
    kind: str = "task",
    status: str = "planlagt",
    parent: Optional[SluplanTask] = None,
) -> SluplanTask:
    task = SluplanTask(
        project=project,
        title=title,
        start_date=start_day,
        end_date=end_day,
        status=status,
        kind=kind,
        parent=parent,
    )
    db.session.add(task)
    db.session.flush()
    return task


def _seed_default_plan(project: SluplanProject, systems: Optional[Iterable[str]]) -> None:
    start_day = _upcoming_monday()

    milestones: List[SluplanTask] = []
    for index, milestone in enumerate(STANDARD_MILESTONES):
        milestone_day = start_day + timedelta(days=index * 7)
        task = _create_task(project, milestone, milestone_day, milestone_day, kind="milestone")
        milestones.append(task)

    for current, nxt in zip(milestones, milestones[1:]):
        dependency = SluplanDependency(from_task=current, to_task=nxt, type="FS")
        db.session.add(dependency)

    if not systems:
        return

    spacing = max(len(STANDARD_SYSTEM_SUBTASKS), 5)
    for offset, system_name in enumerate(systems):
        system_start = start_day + timedelta(days=len(milestones) * spacing + offset * spacing)
        parent = _create_task(project, system_name, system_start, system_start + timedelta(days=len(STANDARD_SYSTEM_SUBTASKS) - 1), kind="system")
        previous_child: Optional[SluplanTask] = None
        for day_offset, sub_title in enumerate(STANDARD_SYSTEM_SUBTASKS):
            day = system_start + timedelta(days=day_offset)
            child = _create_task(project, f"{system_name} · {sub_title}", day, day, kind="subtask", parent=parent)
            if previous_child:
                db.session.add(SluplanDependency(from_task=previous_child, to_task=child))
            previous_child = child


def _delete_project_tasks(project: SluplanProject) -> None:
    tasks = db.session.execute(
        select(SluplanTask).options(joinedload(SluplanTask.children)).where(SluplanTask.project_id == project.id)
    ).scalars()
    for task in tasks:
        db.session.delete(task)
    db.session.flush()


def _collect_dependencies_for_project(project: SluplanProject) -> List[SluplanDependency]:
    statement = (
        select(SluplanDependency)
        .join(SluplanTask, SluplanDependency.from_task_id == SluplanTask.id)
        .where(SluplanTask.project_id == project.id)
    )
    return list(db.session.execute(statement).scalars())


# ---------------------------------------------------------------------------
# Public API – Projects & metadata
# ---------------------------------------------------------------------------


def list_projects(search: Optional[str] = None) -> List[Dict[str, Any]]:
    query = (
        select(SluplanProject)
        .options(joinedload(SluplanProject.project))
        .order_by(SluplanProject.created_at.desc())
    )
    if search:
        pattern = f"%{search.lower()}%"
        query = (
            query.join(Project, SluplanProject.project_id == Project.id, isouter=True)
            .where(
            or_(
                func.lower(SluplanProject.name).like(pattern),
                func.lower(SluplanProject.order_number).like(pattern),
                func.lower(Project.project_name).like(pattern),
                func.lower(Project.order_number).like(pattern),
                func.lower(Project.location).like(pattern),
            )
        )
        )
    projects = db.session.execute(query).scalars()
    return [_serialize_project(project) for project in projects]


def list_base_projects(search: Optional[str] = None, limit: int = 25) -> List[Dict[str, Any]]:
    query = select(Project).order_by(Project.created_at.desc())
    if search:
        pattern = f"%{search.lower()}%"
        query = query.where(
            or_(
                func.lower(Project.project_name).like(pattern),
                func.lower(Project.order_number).like(pattern),
            )
        )
    projects = db.session.execute(query.limit(limit)).scalars()
    return [
        {
            "id": project.id,
            "project_name": project.project_name,
            "order_number": project.order_number,
            "location": project.location,
        }
        for project in projects
    ]


def create_project(payload: Dict[str, Any]) -> Dict[str, Any]:
    name = (payload.get("name") or "").strip()
    if not name and payload.get("project_id"):
        base = db.session.get(Project, int(payload["project_id"]))
        if base:
            name = base.project_name
    if not name:
        raise ValueError("project_name_mangler")

    start_raw = (payload.get("start_date") or "").strip()
    end_raw = (payload.get("end_date") or "").strip()
    if not start_raw or not end_raw:
        raise ValueError("project_dato_mangler")

    start_day = _parse_iso_date(start_raw)
    end_day = _parse_iso_date(end_raw)
    if end_day < start_day:
        raise ValueError("project_slutt_for_start")

    project_ref: Optional[Project] = None
    if payload.get("project_id"):
        project_ref = db.session.get(Project, int(payload["project_id"]))
        if not project_ref:
            raise ValueError("project_ref_finnes_ikke")

    project = SluplanProject(
        name=name,
        order_number=(payload.get("order_number") or (project_ref.order_number if project_ref else None)),
        start_date=start_day,
        end_date=end_day,
        project=project_ref,
    )
    db.session.add(project)
    db.session.flush()

    systems = payload.get("systems")
    if systems and isinstance(systems, list):
        systems_iter = [str(item) for item in systems if str(item).strip()]
    else:
        systems_iter = None

    _seed_default_plan(project, systems_iter)
    db.session.commit()
    return _serialize_project(project)


def get_project_snapshot(project_id: Optional[int]) -> Dict[str, Any]:
    project = _project_by_id(project_id)
    return _serialize_project(project)


def list_resource_catalog() -> Dict[str, Any]:
    users = db.session.execute(select(User).order_by(User.first_name, User.last_name)).scalars()
    disciplines = db.session.execute(select(Fag).order_by(Fag.name)).scalars()
    return {
        "users": [
            {
                "id": user.id,
                "name": f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip() or user.email,
                "first_name": (user.first_name or "").strip(),
                "last_name": (user.last_name or "").strip(),
                "email": user.email,
                "fag": user.fag,
                "role": user.role,
                "location": user.location,
            }
            for user in users
        ],
        "disciplines": [
            {
                "id": discipline.id,
                "name": discipline.name,
            }
            for discipline in disciplines
        ],
    }


# ---------------------------------------------------------------------------
# Plan operations
# ---------------------------------------------------------------------------


def get_tasks(project_id: Optional[int] = None) -> List[Dict[str, Any]]:
    project = _project_by_id(project_id)
    tasks = _load_tasks_for_project(project)
    return _build_task_tree(tasks)


def get_plan_snapshot(project_id: Optional[int] = None) -> Dict[str, Any]:
    project = _project_by_id(project_id)
    tasks = _load_tasks_for_project(project)
    dependencies = _collect_dependencies_for_project(project)
    return {
        "project": _serialize_project(project),
        "tasks": _build_task_tree(tasks),
        "dependencies": [_serialize_dependency(dep) for dep in dependencies],
    }


def export_plan(project_id: Optional[int] = None) -> Dict[str, Any]:
    snapshot = get_plan_snapshot(project_id)
    snapshot["generated_at"] = datetime.now(timezone.utc).isoformat()
    return snapshot


def reset_plan(project_id: Optional[int] = None, *, systems: Optional[Iterable[str]] = None) -> Dict[str, Any]:
    project = _project_by_id(project_id)
    _delete_project_tasks(project)
    _seed_default_plan(project, systems)
    db.session.commit()
    return get_plan_snapshot(project.id)


def create_task(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("ugyldig_payload")

    project = _project_by_id(payload.get("project_id"))

    title = (payload.get("title") or "").strip()
    if not title:
        raise ValueError("title_mangler")

    start_raw = (payload.get("start") or "").strip()
    if not start_raw:
        raise ValueError("start_mangler")
    start_day = _parse_iso_date(start_raw)

    end_raw = (payload.get("end") or "").strip() or start_raw
    end_day = _parse_iso_date(end_raw)
    if end_day < start_day:
        end_day = start_day

    assignee = _ensure_person(payload.get("assignee"), payload.get("assignee_email"))
    discipline = _ensure_discipline(payload.get("discipline"))

    parent: Optional[SluplanTask] = None
    if payload.get("parent_id"):
        parent = db.session.get(SluplanTask, int(payload["parent_id"]))
        if parent and parent.project_id != project.id:
            raise ValueError("parent_tilhorer_annet_prosjekt")

    task = SluplanTask(
        project=project,
        title=title,
        start_date=start_day,
        end_date=end_day,
        status=(payload.get("status") or "planlagt").strip() or "planlagt",
        kind=(payload.get("kind") or "task").strip() or "task",
        parent=parent,
        assignee=assignee,
        discipline=discipline,
    )
    db.session.add(task)
    db.session.commit()
    return _serialize_task(task)


def _task_with_project(task_id: str) -> Tuple[SluplanTask, SluplanProject]:
    task = db.session.get(SluplanTask, int(task_id))
    if not task:
        raise ValueError("task_finnes_ikke")
    project = task.project
    if not project:
        raise ValueError("project_finnes_ikke")
    return task, project


def update_task_fields(task_id: str, fields: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(fields, dict):
        raise ValueError("ugyldig_payload")
    task, _project = _task_with_project(task_id)

    if "assignee" in fields or "assignee_email" in fields:
        task.assignee = _ensure_person(fields.get("assignee"), fields.get("assignee_email"))

    if "status" in fields and isinstance(fields["status"], str) and fields["status"].strip():
        task.status = fields["status"].strip()

    if "title" in fields and isinstance(fields["title"], str) and fields["title"].strip():
        task.title = fields["title"].strip()

    if "kind" in fields and isinstance(fields["kind"], str) and fields["kind"].strip():
        task.kind = fields["kind"].strip()

    if "discipline" in fields:
        task.discipline = _ensure_discipline(fields.get("discipline"))

    if "start" in fields or "end" in fields:
        start = str(fields["start"]) if "start" in fields else None
        end = str(fields["end"]) if "end" in fields else None
        if start:
            start_day = _parse_iso_date(start)
            task.start_date = start_day
            if task.end_date < start_day:
                task.end_date = start_day
        if end:
            end_day = _parse_iso_date(end)
            if end_day < task.start_date:
                end_day = task.start_date
            task.end_date = end_day

    db.session.commit()
    return _serialize_task(task)


def add_comment_to_task(task_id: str, text: str, author: Optional[str] = None) -> Dict[str, Any]:
    message = (text or "").strip()
    if not message:
        raise ValueError("kommentar_tom")
    task, _ = _task_with_project(task_id)
    mentions = MENTION_PATTERN.findall(message)
    comment = SluplanComment(
        task=task,
        text=message,
        author=(author or "").strip() or None,
        mentions=",".join(mentions) or None,
    )
    db.session.add(comment)
    db.session.commit()
    db.session.refresh(task)
    return _serialize_task(task)


def add_file_attachment(task_id: str, filename: str, size: int, content_type: Optional[str] = None) -> Dict[str, Any]:
    task, _ = _task_with_project(task_id)
    cleaned = (filename or "").strip()
    if not cleaned:
        raise ValueError("filnavn_mangler")
    try:
        numeric_size = int(size)
    except (TypeError, ValueError):
        numeric_size = 0
    attachment = SluplanFile(
        task=task,
        filename=cleaned,
        size=max(0, numeric_size),
        content_type=content_type,
    )
    db.session.add(attachment)
    db.session.commit()
    db.session.refresh(task)
    return _serialize_task(task)


def generate_task_ics(task_id: str) -> str:
    task, project = _task_with_project(task_id)

    summary = task.title.strip()
    description_parts = []
    if task.assignee:
        description_parts.append(f"Ressurs: {task.assignee.name}")
    if task.status:
        description_parts.append(f"Status: {task.status}")
    if task.kind:
        description_parts.append(f"Type: {task.kind}")
    if project:
        description_parts.append(f"Prosjekt: {project.name}")
    if task.comments:
        description_parts.append(f"Siste kommentar: {task.comments[-1].text}")
    description = "\n".join(description_parts) if description_parts else "SluPlan-oppgave"

    tz = ZoneInfo("Europe/Oslo")
    start_dt = datetime.combine(task.start_date, time(hour=7, minute=0, tzinfo=tz))
    end_dt = datetime.combine(task.end_date, time(hour=15, minute=30, tzinfo=tz))
    if end_dt <= start_dt:
        end_dt = start_dt + timedelta(hours=8, minutes=30)

    def fmt(dt: datetime) -> str:
        return dt.strftime("%Y%m%dT%H%M%S")

    uid = f"sluplan-{task.id}@prosjektbasen"
    dtstamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Prosjektbasen//SluPlan//NO",
        "CALSCALE:GREGORIAN",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{dtstamp}",
        f"SUMMARY:{summary}",
        f"DTSTART;TZID=Europe/Oslo:{fmt(start_dt)}",
        f"DTEND;TZID=Europe/Oslo:{fmt(end_dt)}",
        f"DESCRIPTION:{description}",
        "STATUS:CONFIRMED",
        "END:VEVENT",
        "END:VCALENDAR",
    ]
    return "\r\n".join(lines) + "\r\n"


def import_systems_from_excel(file_storage, *, project_id: Optional[int] = None) -> Dict[str, Any]:
    project = _project_by_id(project_id)

    raw_bytes = file_storage.read()
    if not raw_bytes:
        raise ValueError("tom_fil")

    try:
        workbook = load_workbook(BytesIO(raw_bytes), data_only=True)
    except Exception as exc:  # pragma: no cover
        raise ValueError("ugyldig_excel") from exc

    sheet = workbook.active
    if sheet.max_row < 2:
        raise ValueError("ingen_data")

    header = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    try:
        system_col_index = next(idx for idx, name in enumerate(header) if name.lower() == "system")
    except StopIteration as exc:
        raise ValueError("mangler_system_kolonne") from exc

    unique_systems: List[str] = []
    seen = set()
    for row in sheet.iter_rows(min_row=2):
        value = row[system_col_index].value
        if value is None:
            continue
        name = str(value).strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_systems.append(name)

    if not unique_systems:
        raise ValueError("ingen_systemer")

    _seed_default_plan(project, systems=unique_systems)
    db.session.commit()
    return get_plan_snapshot(project.id)


def get_alerts(project_id: Optional[int] = None, reference_date: Optional[date] = None) -> Dict[str, Any]:
    project = _project_by_id(project_id)
    today = reference_date or date.today()
    statement = (
        select(SluplanTask)
        .options(joinedload(SluplanTask.assignee))
        .where(SluplanTask.project_id == project.id)
    )

    upcoming: List[Dict[str, Any]] = []
    due_today: List[Dict[str, Any]] = []
    overdue: List[Dict[str, Any]] = []

    for task in db.session.execute(statement).scalars():
        if (task.status or "").strip().lower() == "ferdig":
            continue
        delta_days = (task.end_date - today).days
        payload = {
            "id": str(task.id),
            "title": task.title,
            "start": _iso_date(task.start_date),
            "end": _iso_date(task.end_date),
            "assignee": task.assignee.name if task.assignee else None,
            "assignee_email": task.assignee.email if task.assignee else None,
            "status": task.status,
            "kind": task.kind,
            "days_until_due": delta_days,
        }
        if delta_days < 0:
            overdue.append(payload)
        elif delta_days == 0:
            due_today.append(payload)
        elif delta_days <= 7:
            upcoming.append(payload)

    def sort_key(item: Dict[str, Any]) -> Tuple[int, str]:
        return (item.get("days_until_due") or 0, item.get("title") or "")

    upcoming.sort(key=sort_key)
    due_today.sort(key=sort_key)
    overdue.sort(key=sort_key)

    return {
        "project": _serialize_project(project),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "reference_date": today.isoformat(),
        "upcoming": upcoming,
        "today": due_today,
        "overdue": overdue,
    }


def summarize_by_discipline(project_id: Optional[int] = None) -> List[Dict[str, Any]]:
    project = _project_by_id(project_id)
    statement = (
        select(SluplanTask)
        .options(joinedload(SluplanTask.discipline))
        .where(SluplanTask.project_id == project.id)
    )
    summary: Dict[str, Dict[str, Any]] = {}
    for task in db.session.execute(statement).scalars():
        key = task.discipline.name if task.discipline else (task.kind.capitalize() if task.kind else "Annet")
        entry = summary.setdefault(
            key,
            {
                "discipline": key,
                "task_count": 0,
                "planned": 0,
                "in_progress": 0,
                "completed": 0,
                "total_duration_days": 0,
            },
        )
        entry["task_count"] += 1
        entry[_status_bucket(task.status)] += 1
        entry["total_duration_days"] += _task_duration_days(task)

    result: List[Dict[str, Any]] = []
    for item in summary.values():
        count = item["task_count"] or 1
        result.append(
            {
                **item,
                "average_duration_days": round(item["total_duration_days"] / count, 2),
            }
        )
    result.sort(key=lambda entry: entry["discipline"].lower())
    return result


def summarize_by_user(project_id: Optional[int] = None) -> List[Dict[str, Any]]:
    project = _project_by_id(project_id)
    statement = (
        select(SluplanTask)
        .options(joinedload(SluplanTask.assignee))
        .where(SluplanTask.project_id == project.id)
    )
    summary: Dict[str, Dict[str, Any]] = {}
    for task in db.session.execute(statement).scalars():
        key = task.assignee.name if task.assignee else "Ikke tildelt"
        entry = summary.setdefault(
            key,
            {
                "user": key,
                "task_count": 0,
                "planned": 0,
                "in_progress": 0,
                "completed": 0,
                "total_duration_days": 0,
            },
        )
        entry["task_count"] += 1
        entry[_status_bucket(task.status)] += 1
        entry["total_duration_days"] += _task_duration_days(task)

    result: List[Dict[str, Any]] = []
    for item in summary.values():
        count = item["task_count"] or 1
        result.append(
            {
                **item,
                "average_duration_days": round(item["total_duration_days"] / count, 2),
            }
        )
    result.sort(key=lambda entry: entry["user"].lower())
    return result


def summarize_by_time_window(
    project_id: Optional[int] = None,
    from_value: Optional[str] = None,
    to_value: Optional[str] = None,
) -> Dict[str, Any]:
    project = _project_by_id(project_id)
    today = date.today()

    try:
        start = _parse_iso_date(from_value) if from_value else today
    except Exception as exc:
        raise ValueError("from_ugyldig") from exc

    try:
        end = _parse_iso_date(to_value) if to_value else start + timedelta(days=30)
    except Exception as exc:
        raise ValueError("to_ugyldig") from exc

    if end < start:
        start, end = end, start

    statement = (
        select(SluplanTask)
        .options(joinedload(SluplanTask.assignee), joinedload(SluplanTask.discipline))
        .where(SluplanTask.project_id == project.id)
    )

    planned = in_progress = completed = 0
    total_duration = 0
    tasks_payload: List[Dict[str, Any]] = []

    for task in db.session.execute(statement).scalars():
        if task.end_date < start or task.start_date > end:
            continue
        bucket = _status_bucket(task.status)
        if bucket == "completed":
            completed += 1
        elif bucket == "in_progress":
            in_progress += 1
        else:
            planned += 1

        duration = _task_duration_days(task)
        total_duration += duration
        tasks_payload.append(
            {
                "id": str(task.id),
                "title": task.title,
                "start": _iso_date(task.start_date),
                "end": _iso_date(task.end_date),
                "assignee": task.assignee.name if task.assignee else None,
                "assignee_email": task.assignee.email if task.assignee else None,
                "status": task.status,
                "discipline": task.discipline.name if task.discipline else None,
                "duration_days": duration,
            }
        )

    total_tasks = planned + in_progress + completed
    tasks_payload.sort(key=lambda item: (item["start"], item["title"]))

    return {
        "project": _serialize_project(project),
        "from": start.isoformat(),
        "to": end.isoformat(),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_tasks": total_tasks,
        "planned": planned,
        "in_progress": in_progress,
        "completed": completed,
        "total_duration_days": total_duration,
        "tasks": tasks_payload,
    }


def import_plan(payload: Dict[str, Any], *, project_id: Optional[int] = None) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("ugyldig_planformat")
    project = _project_by_id(project_id)

    tasks_data = payload.get("tasks")
    if not isinstance(tasks_data, list):
        raise ValueError("tasks_må_være_liste")
    dependencies_data = payload.get("dependencies") or []
    if not isinstance(dependencies_data, list):
        raise ValueError("dependencies_må_være_liste")

    _delete_project_tasks(project)

    id_mapping: Dict[str, SluplanTask] = {}

    def _import_task(data: Dict[str, Any], parent: Optional[SluplanTask] = None) -> SluplanTask:
        title = (data.get("title") or data.get("name") or "").strip()
        if not title:
            raise ValueError("task_title_mangler")
        start_raw = (data.get("start") or data.get("start_date") or "").strip()
        end_raw = (data.get("end") or data.get("end_date") or start_raw).strip()
        start_day = _parse_iso_date(start_raw)
        end_day = _parse_iso_date(end_raw) if end_raw else start_day
        if end_day < start_day:
            end_day = start_day

        assignee = _ensure_person(data.get("assignee"), data.get("assignee_email"))
        discipline = _ensure_discipline(data.get("discipline"))

        task = SluplanTask(
            project=project,
            title=title,
            start_date=start_day,
            end_date=end_day,
            status=(data.get("status") or "planlagt").strip() or "planlagt",
            kind=(data.get("kind") or "task").strip() or "task",
            parent=parent,
            assignee=assignee,
            discipline=discipline,
        )
        db.session.add(task)
        db.session.flush()

        original_id = data.get("id")
        if original_id is not None:
            id_mapping[str(original_id)] = task

        for comment_data in data.get("comments") or []:
            text = (comment_data.get("text") or "").strip()
            if not text:
                continue
            mentions = comment_data.get("mentions")
            mentions_value = ",".join(str(item) for item in mentions) if isinstance(mentions, list) else None
            created_at = _parse_iso_datetime(comment_data.get("created_at"))
            comment = SluplanComment(
                task=task,
                text=text,
                author=(comment_data.get("author") or "").strip() or None,
                mentions=mentions_value,
                created_at=created_at or datetime.utcnow(),
            )
            db.session.add(comment)

        for file_data in data.get("files") or []:
            filename = (file_data.get("filename") or "").strip()
            if not filename:
                continue
            size = file_data.get("size")
            try:
                numeric_size = int(size)
            except (TypeError, ValueError):
                numeric_size = 0
            uploaded_at = _parse_iso_datetime(file_data.get("uploaded_at"))
            attachment = SluplanFile(
                task=task,
                filename=filename,
                size=max(0, numeric_size),
                content_type=file_data.get("content_type"),
                uploaded_at=uploaded_at or datetime.utcnow(),
            )
            db.session.add(attachment)

        for child in data.get("children") or []:
            _import_task(child, parent=task)

        return task

    for root in tasks_data:
        _import_task(root, parent=None)

    for dependency in dependencies_data:
        from_id = dependency.get("fromId") or dependency.get("from_id")
        to_id = dependency.get("toId") or dependency.get("to_id")
        if from_id is None or to_id is None:
            continue
        from_task = id_mapping.get(str(from_id))
        to_task = id_mapping.get(str(to_id))
        if not from_task or not to_task:
            continue
        db.session.add(SluplanDependency(from_task=from_task, to_task=to_task, type=dependency.get("type") or "FS"))

    db.session.commit()
    return get_plan_snapshot(project.id)
