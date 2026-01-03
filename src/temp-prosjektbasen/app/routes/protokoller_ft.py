import io
import os
import json
import re
from datetime import datetime
from flask import request, jsonify, send_file, current_app
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule, DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

# Hent alt vi trenger fra core
from .protokoller_core import (
    bp,
    extract_text_from_file,
    parse_rows_from_text,
    TFM_DICT,
    DATA_DIR,
    get_unique_system_id,   # ← viktig for kanonisk system-ID
)

# ──────────────────────────────────────────────────────────────────────────────
# Funksjonstest • GENERER UNDERLAG (REN JSON)
# ──────────────────────────────────────────────────────────────────────────────
@bp.route("/generate_funksjonstest", methods=["POST"])
@login_required
def generate_funksjonstest_underlag():
    """
    Funksjonstest: leser opplastede filer, bruker felles parser og returnerer JSON-liste.
    Filtrerer på system_kriterier (prefiks), og fyller inn komponent/TFM-desc robust.
    """
    files = request.files.getlist("files")
    out: list[dict] = []

    # system_kriterier, f.eks. "36" eller "36, 37"
    raw_criteria = (request.form.get("system_kriterier") or "").strip()
    allowed_prefixes = [s for s in re.split(r"[,\s]+", raw_criteria) if s]

    def _allowed(uid: str) -> bool:
        if not allowed_prefixes:
            return True
        s = str(uid or "")
        return any(s.startswith(pref) for pref in allowed_prefixes)

    # Hjelpere (samme prinsipp som i MC)
    SYS_RE  = re.compile(r"\b\d{3,4}\.\d{3,4}\b")
    COMP_RE = re.compile(r"-?([A-Za-z]{2,4}\d{2,5})(?:%[^\s+]+)?")

    def parse_component_from_full_id(full_id: str) -> str | None:
        s = str(full_id or "")
        m = COMP_RE.search(s)
        return m.group(1) if m else None

    def tfm_prefix_from_component(komp: str) -> str:
        if not komp:
            return ""
        letters = re.match(r"^[A-Za-z]{2,}", komp)
        return (letters.group(0)[:2].upper()) if letters else ""

    def canonical_system(row: dict) -> str:
        # 1) allerede kanonisk?
        uid = row.get("unique_system")
        if uid:
            return uid
        # 2) fra system_number
        sysnum = row.get("system_number") or ""
        if sysnum:
            return get_unique_system_id(sysnum)
        # 3) fra full_id
        fid = row.get("full_id") or ""
        m = SYS_RE.search(fid)
        return get_unique_system_id(m.group(0)) if m else ""

    # Prosesser filer
    for f in files:
        if not f or not f.filename:
            continue
        try:
            try:
                f.seek(0)
            except Exception:
                pass

            text = extract_text_from_file(f)
            for row in parse_rows_from_text(text, f.filename):
                uid = canonical_system(row)
                if not _allowed(uid):
                    continue

                # komponent: prioritér feltet, ellers trekk ut fra full_id
                komp = row.get("komponent") or parse_component_from_full_id(row.get("full_id", "")) or ""
                tfm_key = tfm_prefix_from_component(komp)
                desc = TFM_DICT.get(tfm_key, "Ukjent beskrivelse")

                out.append({
                    "source": row.get("source", f.filename),
                    "unique_system": uid,
                    "system_full_name": row.get("system_full_name", ""),
                    "system_number": row.get("system_number", ""),
                    "full_id": row.get("full_id", ""),
                    "komponent": komp,
                    "funksjonsvalg": row.get("funksjonsvalg", "Øvrig"),
                    "desc": desc,
                })

        except Exception as e:
            out.append({
                "source": getattr(f, "filename", "ukjent"),
                "system_number": "",
                "komponent": "",
                "error": f"{type(e).__name__}: {e}",
            })

    return jsonify(out)


# ──────────────────────────────────────────────────────────────────────────────
# Funksjonstest • LAST NED PROTOKOLL (EXCEL)
# ──────────────────────────────────────────────────────────────────────────────

def _normalize_rows_from_request():
    """
    Godtar både {rows:[...]} og direkte [...].
    Tåler at enkelte rader er JSON-strenger og forsøker å json.loads dem.
    """
    payload = request.get_json(silent=True)

    if isinstance(payload, dict):
        rows = payload.get("rows", [])
    elif isinstance(payload, list):
        rows = payload
    else:
        rows = []

    norm = []
    for r in rows:
        if isinstance(r, dict):
            norm.append(r)
        elif isinstance(r, str):
            try:
                obj = json.loads(r)
                if isinstance(obj, dict):
                    norm.append(obj)
            except Exception:
                # ignorer rå strenger
                continue
    return norm

@bp.route("/download_funksjonstest_protokoll", methods=["POST"])
@login_required
def download_funksjonstest_protokoll():
    from flask import current_app
    from openpyxl.styles import Alignment
    log = current_app.logger

    # ---------- robust payload -> rows (uendret prinsipp) ----------
    def _load_payload():
        payload = request.get_json(silent=True)
        if payload is None:
            raw = request.get_data(cache=False, as_text=True)
            try:
                payload = json.loads(raw)
            except Exception:
                payload = raw
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:
                pass
        return payload

    def _norm_rows(payload):
        if isinstance(payload, dict):
            rows = payload.get("rows", [])
        elif isinstance(payload, list):
            rows = payload
        else:
            rows = []
        norm = []
        for r in rows:
            if isinstance(r, dict):
                norm.append(r)
            elif isinstance(r, str):
                try:
                    obj = json.loads(r)
                    if isinstance(obj, dict):
                        norm.append(obj)
                except Exception:
                    pass
        return norm

    payload = _load_payload()
    rows = _norm_rows(payload)
    if not rows:
        return jsonify({"error": "Ingen gyldige rader mottatt."}), 400

    # ---------- last mal ----------
    template_path = os.path.join(DATA_DIR, "Funksjonstest.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Malfil ikke funnet",
                        "hint": f"Letet etter: {template_path}"}), 500
    wb = load_workbook(template_path)

    # Finn malark (som før)
    malark = None
    for name in wb.sheetnames:
        if name.lower() in ("mal", "malverk", "template", "ft_mal"):
            malark = wb[name]; break
    if malark is None:
        malark = wb[wb.sheetnames[0]]

    # ---------- grupper per system (beholder denne logikken) ----------
    def _sys(r):
        # prioriterer felt som finnes i datasettet; faller tilbake til "Uspesifisert"
        for k in ("system_number", "system", "unique_system"):
            v = (r.get(k) or "").strip()
            if v:
                return v
        return "Uspesifisert"

    groups = {}
    for r in rows:
        groups.setdefault(_sys(r), []).append(r)

    # ---------- helpers ----------
    START_ROW = 29
    COL = {"A": 1, "B": 2, "C": 3, "E": 5, "G": 7, "H": 8}

    FUNK_CHOICES = [
        "Start og Stopp funksjoner",
        "Reguleringsfunksjoner",
        "Sikkerhetsfunksjoner",
        "Øvrig",
    ]

    def normalize_funksjonsvalg(val: str | None) -> str:
        s = (val or "").strip().lower()
        if not s:
            return "Øvrig"
        if "start" in s:
            return "Start og Stopp funksjoner"
        if "reguler" in s:
            return "Reguleringsfunksjoner"
        if "sikker" in s:
            return "Sikkerhetsfunksjoner"
        return "Øvrig"

    def replace_placeholders(ws, system_text: str):
        patt = re.compile(r"\{SYSTEM(NAVN)?\}", re.I)
        for r in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=40):
            for cell in r:
                v = cell.value
                if isinstance(v, str) and patt.search(v):
                    cell.value = patt.sub(system_text, v)

    # Validering + CF for datarader (Status i A, Funksjonsvalg i H)
    def apply_validations_and_cf(ws, start_row: int, end_row: int):
        if start_row > end_row:
            return

        # Status (A) – nedtrekk + CF
        dv_status = DataValidation(
            type="list",
            formula1='"Ikke startet,Under arbeid,Avvik,Utført"',
            allow_blank=True
        )
        ws.add_data_validation(dv_status)
        dv_status.add(f"A{start_row}:A{end_row}")

        # CF-farger for Status (samme logikk som før)
        FILL_GRAY  = PatternFill("solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
        FILL_BLUE  = PatternFill("solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
        FILL_RED   = PatternFill("solid", start_color="FFF8D7DA", end_color="FFF8D7DA")
        FILL_GREEN = PatternFill("solid", start_color="FFD4EDDA", end_color="FFD4EDDA")
        FONT_BLACK = Font(color="FF000000")
        FONT_WHITE = Font(color="FFFFFFFF")

        def rule(expr, fill, font):
            dxf = DifferentialStyle(fill=fill, font=font)
            return Rule(type="expression", dxf=dxf, stopIfTrue=False, formula=[expr])

        rngA = f"A{start_row}:A{end_row}"
        ws.conditional_formatting.add(rngA, rule('=EXACT($A1,"Ikke startet")', FILL_GRAY,  FONT_BLACK))
        ws.conditional_formatting.add(rngA, rule('=EXACT($A1,"Under arbeid")', FILL_BLUE,  FONT_BLACK))
        ws.conditional_formatting.add(rngA, rule('=EXACT($A1,"Avvik")',        FILL_RED,   FONT_WHITE))
        ws.conditional_formatting.add(rngA, rule('=EXACT($A1,"Utført")',       FILL_GREEN, FONT_BLACK))

        # Funksjonsvalg (H) – nedtrekk
        lista = '"' + ",".join(FUNK_CHOICES) + '"'
        dv_funk = DataValidation(type="list", formula1=lista, allow_blank=True)
        ws.add_data_validation(dv_funk)
        dv_funk.add(f"H{start_row}:H{end_row}")

    # ---------- lag ark per system og fyll (uten seksjonsoverskrifter) ----------
    for system, liste in groups.items():
        title = f"{system} - FT"
        ws = wb[title] if title in wb.sheetnames else wb.copy_worksheet(malark)
        ws.title = title

        replace_placeholders(ws, system)

        # Tøm dataområde (A–N) fra rad 29 og ned 1000 rader
        for rr in range(START_ROW, START_ROW + 1000):
            for cc in range(1, 14 + 1):  # A..N
                ws.cell(rr, cc).value = None

        r = START_ROW
        first_data_row = r

        # Skriv alle rader fortløpende
        for obj in liste:
            system_txt = _sys(obj)
            komponent  = obj.get("komponent", "")
            test       = obj.get("testutfoerelse", "") or obj.get("test", "")
            aksept     = obj.get("aksept", "") or obj.get("forventet_resultat", "")
            status_val = obj.get("status") or "Ikke startet"
            funksjonsvalg = normalize_funksjonsvalg(obj.get("funksjonsvalg") or obj.get("integrert"))

            ws.cell(r, COL["A"], status_val)      # Status
            ws.cell(r, COL["B"], system_txt)      # System
            ws.cell(r, COL["C"], komponent)       # Komponent
            ws.cell(r, COL["E"], test)            # Testutførelse
            ws.cell(r, COL["G"], aksept)          # Akseptkriterie
            ws.cell(r, COL["H"], funksjonsvalg)   # Funksjonsvalg (NY)
            r += 1

        last_data_row = r - 1
        if last_data_row >= first_data_row:
            apply_validations_and_cf(ws, first_data_row, last_data_row)

    # --- Fjern malarket helt til slutt ---
    try:
        if malark.title in wb.sheetnames:
            wb.remove(wb[malark.title])
    except Exception as e:
        log.warning("Kunne ikke fjerne malarket %r: %s", malark.title, e)

    # Sorter ark alfabetisk (som før)
    try:
        wb._sheets = [wb[n] for n in sorted(wb.sheetnames)]
    except Exception:
        pass

    # ---------- send fil ----------
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"Funksjonstest_Protokoll_{datetime.utcnow():%Y%m%d_%H%M%S}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )