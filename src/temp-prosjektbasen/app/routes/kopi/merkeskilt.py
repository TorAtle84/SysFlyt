import os
import tempfile
import traceback
import re
import json

from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
import pandas as pd
from openpyxl import Workbook
from app.models.db import db
from app.models.merkeskilt import Merkskilt
from app.models import User
from app.models.project import Project

merkeskilt_bp = Blueprint('merkeskilt', __name__)

# ────────────────────────────────────────────────────────────────────────────────
# Hent kode→beskrivelse‐mapping (tfm.xlsx)
# ────────────────────────────────────────────────────────────────────────────────
basedir = os.path.abspath(os.path.dirname(__file__))
mapping_path = os.path.abspath(os.path.join(basedir, "..", "data", "tfm.xlsx"))

try:
    mapping_df = pd.read_excel(mapping_path, sheet_name="TFM", header=None, dtype=str)
    mapping_df = mapping_df.fillna("Ikke i bruk")  # <- ny linje
    kode_mapping = dict(zip(mapping_df[0].astype(str), mapping_df[1].astype(str)))
except Exception as e:
    kode_mapping = {}
    current_app.logger.error(f"Kunne ikke lese tfm.xlsx: {e}")

def hent_teknikere_og_admin():
    brukere = User.query.filter(User.role.in_(["tekniker", "admin"])).all()
    resultat_liste = []
    for b in brukere:
        navn = f"{getattr(b, 'first_name', '')} {getattr(b, 'last_name', '')}".strip() or getattr(b, "email", "")
        resultat_liste.append({"id": b.id, "navn": navn})
    return resultat_liste

@merkeskilt_bp.route("/merkeskilt")
def vis_merkeskilt():
    # 1) Teknikere/admin for dropdown (dersom du trenger det)
    teknikere = hent_teknikere_og_admin()

    # 2) TFM-innstillinger
    try:
        settings_path = os.path.join(basedir, "..", "data", "tfm-settings.json")
        with open(settings_path, "r", encoding="utf-8") as f:
            tfm_settings = json.load(f)
    except:
        tfm_settings = {}
    tfm_liste = []
    for _, row in mapping_df.iterrows():
        kode         = str(row[0]).strip()
        beskrivelse  = str(row[1]).strip()
        aktiv        = tfm_settings.get(kode, True)
        tfm_liste.append((kode, beskrivelse, aktiv))

    # 3) Hent alle prosjekter for «Send til prosjekt»
    projects = Project.query.order_by(Project.project_name).all()

    # 4) Render malen med alle nødvendige parametre
    return render_template(
        "merkeskilt.html",
        teknikere=teknikere,
        tfm_liste=tfm_liste,
        projects=projects
    )

@merkeskilt_bp.route("/api/generer", methods=["POST"])
def generer_komponenter():
    data = request.json or {}
    taglinjer = data.get("tags", [])
    formatvalg = data.get("format", "")
    resultat, feilmeldinger = [], []

    settings_path = os.path.abspath(os.path.join(basedir, "..", "data", "tfm-settings.json"))
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            tfm_settings = json.load(f)
    except:
        tfm_settings = {}

    if not formatvalg:
        return jsonify({"feil": ["Ingen format-streng valgt."]}), 400

    for tag in taglinjer:
        tag = tag.strip()
        eq_pos = tag.find("=")
        dash_pos = tag.find("-", eq_pos + 1) if eq_pos != -1 else -1
        percent_pos = tag.find("%", dash_pos + 1) if dash_pos != -1 else -1

        if eq_pos == -1 or dash_pos == -1:
            feilmeldinger.append(f"Feil på rad '{tag}' – mangler '=' eller '-'")
            continue

        byggnr = tag[:eq_pos]
        system = tag[eq_pos:dash_pos]
        komponent = tag[dash_pos:percent_pos] if percent_pos != -1 else tag[dash_pos:]
        typekode = tag[percent_pos:] if percent_pos != -1 else ""

        kode2 = komponent[1:3] if len(komponent) >= 3 else ""
        aktiv = tfm_settings.get(kode2, True)
        if not aktiv:
            continue

        try:
            komp_str = (formatvalg
                .replace("{byggnr}", byggnr)
                .replace("{system}", system)
                .replace("{komponent}", komponent)
                .replace("{typekode}", typekode))
        except:
            feilmeldinger.append(f"Feil i format '{formatvalg}' for '{tag}'")
            continue

        beskrivelse = kode_mapping.get(kode2, "")
        resultat.append({
            "komponent": komp_str,
            "beskrivelse": beskrivelse,
            "himlingsskilt": "Nei",
            "stripsskilt": "Nei"
        })

    if feilmeldinger:
        return jsonify({"feil": feilmeldinger}), 400
    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/scan-og-generer", methods=["POST"])
def scan_og_generer():
    formatvalg = request.form.get("format", "").strip()
    if not formatvalg:
        return jsonify({"feil": ["Ingen format-streng valgt."]}), 400

    kriteriestr = request.form.get("system_kriterier", "").strip()
    system_kriterier = [k for k in kriteriestr.split(",") if k.strip().isdigit()] if kriteriestr else []

    filer = request.files.getlist("files")
    if not filer:
        return jsonify({"feil": ["Ingen filer funnet."]}), 400

    settings_path = os.path.abspath(os.path.join(basedir, "..", "data", "tfm-settings.json"))
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            tfm_settings = json.load(f)
    except:
        tfm_settings = {}

    placeholders = {
        "{byggnr}": r"(?P<byggnr>\+[A-Za-z0-9]+)",
        "{system}": r"(?P<system>=[^-]+)",
        "{komponent}": r"(?P<komponent>-[^%]+)",
        "{typekode}": r"(?P<typekode>[%/].+)"
    }
    regex_str = formatvalg
    for ph, subpat in placeholders.items():
        regex_str = regex_str.replace(ph, subpat)
    try:
        mønster = re.compile(regex_str)
    except re.error as e:
        return jsonify({"feil": [f"Regex-feil: {e}"]}), 400

    funn_set = set()
    for f in filer:
        try:
            data_ark = pd.read_excel(f, sheet_name=None, dtype=str)
        except Exception as e:
            return jsonify({"feil": [f"Feil i fil '{f.filename}': {e}"]}), 400
        for df in data_ark.values():
            for verdi in df.astype(str).values.flatten():
                if not isinstance(verdi, str):
                    continue
                for match in mønster.finditer(verdi):
                    system_val = match.groupdict().get("system", "")
                    if "{system}" in formatvalg and system_kriterier:
                        if system_val[1:3] not in system_kriterier:
                            continue
                    kode2 = match.groupdict().get("komponent", "")[1:3]
                    if not tfm_settings.get(kode2, True):
                        continue
                    funn_set.add(match.group(0))

    resultat = []
    for tag in funn_set:
        match = mønster.match(tag)
        if not match:
            continue
        gr = match.groupdict()
        byggnr = gr.get("byggnr", "")
        system = gr.get("system", "")
        komponent = gr.get("komponent", "")
        typekode = gr.get("typekode", "")
        komponent_val = match.groupdict().get("komponent", "")
        kode2 = komponent_val[1:3] if komponent_val else ""
        beskrivelse = kode_mapping.get(kode2, "")
        komp_str = (
            formatvalg.replace("{byggnr}", byggnr)
                      .replace("{system}", system)
                      .replace("{komponent}", komponent)
                      .replace("{typekode}", typekode)
        )
        resultat.append({
            "komponent": komp_str,
            "beskrivelse": beskrivelse,
            "himlingsskilt": "Nei",
            "stripsskilt": "Nei"
        })

    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/lastned", methods=["POST"])
def last_ned_excel():
    data = request.json or {}
    rader = data.get("rader", [])
    ordrenummer = data.get("ordrenummer")
    if not ordrenummer:
        return jsonify({"error": "Ordrenummer er påkrevd."}), 400

    wb = Workbook()
    ws = wb.active

    # 1) Skriv overskrifter + autofilter
    headers = ["Komponent-ID", "Beskrivelse", "Himlingsskilt", "Stripsskilt"]
    ws.append(headers)
    ws.auto_filter.ref = "A1:D1"  # Autofilter

    # 2) Legg inn radene med riktig datatype
    for rad in rader:
        komponent = rad.get("komponent", "")
        beskrivelse = rad.get("beskrivelse", "")
        himling = rad.get("himlingsskilt", "Nei")
        strips = rad.get("stripsskilt", "Nei")

        radnr = ws.max_row + 1
        c_komp = ws.cell(row=radnr, column=1, value=komponent)
        c_komp.number_format = "@"  # Eksplisitt tekstformat
        c_komp.data_type = "s"

        ws.cell(row=radnr, column=2, value=beskrivelse)
        ws.cell(row=radnr, column=3, value=himling)
        ws.cell(row=radnr, column=4, value=strips)

    # 3) Autojuster kolonnebredder
    for kol in ws.columns:
        maks = max(len(str(c.value)) if c.value else 0 for c in kol)
        ws.column_dimensions[kol[0].column_letter].width = maks + 2

    # 4) Lagre og send
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name=f"{ordrenummer}-Merkeskilt.xlsx")


@merkeskilt_bp.route("/api/tfm-liste", methods=["GET"])
def hent_tfm_liste():
    try:
        settings_path = os.path.abspath(os.path.join(basedir, "..", "data", "tfm-settings.json"))
        with open(settings_path, "r", encoding="utf-8") as f:
            tfm_settings = json.load(f)
    except:
        tfm_settings = {}

    er_admin_flag = current_user.is_authenticated and getattr(current_user, "role", "") == "admin"

    resultat = []
    for idx, row in mapping_df.iterrows():
        kode = str(row[0]).strip()
        beskrivelse = str(row[1]).strip()
        aktiv = tfm_settings.get(kode, True)
        resultat.append({
            "kode": kode,
            "beskrivelse": beskrivelse,
            "aktiv": aktiv,
            "er_admin": er_admin_flag
        })

    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/tfm-liste/save", methods=["POST"])
def lagre_tfm_innstillinger():
    data = request.get_json() or {}
    innstillinger = data.get("innstillinger", [])
    tfm_dict = {opp["kode"]: bool(opp["aktiv"]) for opp in innstillinger if "kode" in opp}
    settings_path = os.path.abspath(os.path.join(basedir, "..", "data", "tfm-settings.json"))
    try:
        with open(settings_path, "w", encoding="utf-8") as f:
            json.dump(tfm_dict, f, ensure_ascii=False, indent=2)
    except Exception as e:
        current_app.logger.error(f"Kunne ikke lagre tfm-settings.json: {e}")
        return jsonify({"feil": ["Klarte ikke lagre."]}), 500
    return jsonify({"success": True}), 200

@merkeskilt_bp.route("/send_merkeskilt", methods=["POST"])
@login_required
def send_merkeskilt():
    data = request.get_json()
    project_id = data.get("project_id")
    rows = data.get("rows")

    if not project_id or not rows:
        return jsonify({"error": "Mangler data"}), 400

    # 1. Slett eksisterende rader for dette prosjektet
    Merkskilt.query.filter_by(project_id=project_id).delete()

    # 2. Legg inn nye rader
    for row in rows:
        ny_rad = Merkskilt(
            project_id=project_id,
            komponent_id=row.get("komponent_id"),
            beskrivelse=row.get("beskrivelse"),
            aktiv=row.get("aktiv") == "Ja",
            inaktiv=row.get("inaktiv") == "Ja"
        )
        db.session.add(ny_rad)

    db.session.commit()
    return jsonify({"success": True})
