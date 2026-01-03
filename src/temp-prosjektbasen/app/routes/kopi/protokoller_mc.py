import os, re, tempfile
from flask import jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import Rule, DifferentialStyle
from copy import copy

from .protokoller_core import bp, DATA_DIR, add_betingelser
# +++ NYTT: hent kanoniserer for system-ID + (valgfritt) TFM-lookup om du ønsker senere
from .protokoller_core import get_unique_system_id  # <- finnes i kjernen vår nå


# === HJELPERE (NYE) ==========================================================

_COMPONENT_RE = re.compile(r"-(?P<komp>[A-Za-z]{2,4}\d{2,5}[^\s%+]*)")

def _canonical_system_from_row(r: dict) -> str:
    """
    Henter kanonisk system-ID:
    - Primært r['unique_system'] hvis satt
    - sekundært r['system_number'] (kan være rå), eller trekk ut fra full_id
    """
    sys = r.get("unique_system")
    if sys:
        return sys
    # fallback: system_number om finnes
    sys = r.get("system_number") or ""
    if sys:
        return get_unique_system_id(sys)
    # siste fallback: prøv å hente ut fra full_id
    full_id = r.get("full_id") or ""
    # hent første 000.000 / 0000.0000 / 000.0000 / 0000.000
    m = re.search(r"\b\d{3,4}\.\d{3,4}\b", full_id)
    return get_unique_system_id(m.group(0) if m else "")

def _clean_display_id(full_id: str) -> str:
    if not full_id:
        return ""
    s = str(full_id).strip()

    # Fjern +BYGGNR=SYSTEM-innledning
    s = re.sub(r"^\+[^=\s+]+=\d{3,4}\.\d{3,4}", "", s).lstrip("+=").strip()

    m = _COMPONENT_RE.search(s)
    if m:
        return m.group("komp")  # f.eks. "JV010T/001"

    # Fallback: klipp ved whitespace eller '+'
    s = re.split(r"\s|\+", s, maxsplit=1)[0]
    return s

def _copy_cell_style(src_cell, dst_cell):
    """Kopier utseende fra src til dst (font/fill/border/alignment/format)."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def _sheet_sort_key(ws_title: str):
    """
    Sorter ark som '4330.201_MC' etter systemdelen numerisk:
    - Ekstraher 'A.B' foran '_MC', finn tall A og B
    - Sortér tuple (A, B, tittel) så ikke-koblede havner til slutt
    """
    base = ws_title.replace("_MC", "")
    m = re.match(r"^\s*(\d{1,4})\.(\d{1,4})\s*$", base)
    if m:
        return (int(m.group(1)), int(m.group(2)), ws_title)
    return (999999, 999999, ws_title)


# === ROUTE (REVIDERT) ========================================================
def _apply_status_cf(ws, start_row: int, end_row: int):
    """
    Betinget formatering for kolonner A–C (status).
    Farger: lys grå / blå / rød / grønn. Font svart, unntak 'Avvik' = hvit.
    """
    # Pastellfarger (ARGB, uten '#')
    FILL_GRAY  = PatternFill(fill_type="solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
    FILL_BLUE  = PatternFill(fill_type="solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
    FILL_RED   = PatternFill(fill_type="solid", start_color="FFF8D7DA", end_color="FFF8D7DA")
    FILL_GREEN = PatternFill(fill_type="solid", start_color="FFD4EDDA", end_color="FFD4EDDA")

    FONT_BLACK = Font(color="FF000000")
    FONT_WHITE = Font(color="FFFFFFFF")

    def _rule(expr: str, fill: PatternFill, font: Font):
        dxf = DifferentialStyle(fill=fill, font=font)
        return Rule(type="expression", dxf=dxf, stopIfTrue=False, formula=[expr])

    # Lag regler per kolonne (A, B, C)
    for col in ("A", "B", "C"):
        cell_range = f"{col}{start_row}:{col}{end_row}"

        # =EXACT($A3,"Ikke startet")
        ws.conditional_formatting.add(
            cell_range,
            _rule(f'=EXACT(${col}{start_row},"Ikke startet")', FILL_GRAY, FONT_BLACK)
        )
        ws.conditional_formatting.add(
            cell_range,
            _rule(f'=EXACT(${col}{start_row},"Under arbeid")', FILL_BLUE, FONT_BLACK)
        )
        ws.conditional_formatting.add(
            cell_range,
            _rule(f'=EXACT(${col}{start_row},"Avvik")', FILL_RED, FONT_WHITE)
        )
        ws.conditional_formatting.add(
            cell_range,
            _rule(f'=EXACT(${col}{start_row},"Utført")', FILL_GREEN, FONT_BLACK)
        )


@bp.route("/download_protokoll", methods=["POST"])
@login_required
def download_protokoll():
    from copy import copy
    data = __import__("flask").request.json
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"feil": "Ingen rader mottatt."}), 400

    mal_path = os.path.join(DATA_DIR, "Mekanisk_Komplett.xlsx")
    if not os.path.exists(mal_path):
        return jsonify({"feil": "Malfil ikke funnet."}), 500

    wb = load_workbook(mal_path)
    malark = wb["malverk"]

    print("\n--- LOGG MC: Data mottatt for MC-protokoll ---")
    print(f"Antall rader mottatt: {len(rows)}")
    for i, row in enumerate(rows[:3]):
        print(f"Rad {i+1}: {row}")
    print("-" * 50)

    # 1) Gruppér per KANONISK system (default 'Uspesifisert')
    grouped = {}
    for r in rows:
        sys = _canonical_system_from_row(r) or "Uspesifisert"
        grouped.setdefault(sys, []).append(r)

    # 2) Skriv ett ark per system
    for sys, liste in grouped.items():
        kopi = wb.copy_worksheet(malark)
        kopi.title = f"{sys}_MC"

        # --- Finn hvilke lokasjonsfelter som finnes i denne gruppen ---
        has_plass = any((r.get("locs") or {}).get("Plassering") for r in liste)
        has_rom   = any((r.get("locs") or {}).get("Rom") for r in liste)
        has_lok   = any((r.get("locs") or {}).get("Lokasjon") for r in liste)
        has_ktr   = any((r.get("locs") or {}).get("Kontrollområde") for r in liste)

        primary_label = None
        if   has_plass: primary_label = "Plassering"
        elif has_rom:   primary_label = "Rom"
        elif has_lok:   primary_label = "Lokasjon"

        # --- Sett inn kolonner etter E (F=6) om aktuelt ---
        insert_at = 6  # kolonne F
        if primary_label:
            kopi.insert_cols(insert_at, amount=1)
            # kopier header-stil (rad 1–2) fra E
            for r_i in (1, 2):
                _copy_cell_style(kopi.cell(r_i, 5), kopi.cell(r_i, insert_at))
            kopi.cell(1, insert_at, primary_label)
            kopi.cell(2, insert_at, None)
            insert_at += 1

        if has_ktr:
            kopi.insert_cols(insert_at, amount=1)
            for r_i in (1, 2):
                _copy_cell_style(kopi.cell(r_i, 5), kopi.cell(r_i, insert_at))
            kopi.cell(1, insert_at, "Kontrollområde")
            kopi.cell(2, insert_at, None)
            insert_at += 1

        # --- Nullstill rad 3–200 (verdier) uten å røre headerne ---
        for i in range(3, 201):
            for j in range(1, kopi.max_column + 1):
                kopi.cell(i, j).value = None

        # --- Datavalidering i A–C (uendret) ---
        dv = DataValidation(type="list", formula1='"Ikke startet,Under arbeid,Avvik,Utført"', allow_blank=True)
        kopi.add_data_validation(dv)
        for col in ["A", "B", "C"]:
            dv.add(f"{col}3:{col}200")
            for i in range(3, 201):
                kopi[f"{col}{i}"].value = "Ikke startet"

        # --- Skriv rader: D/E som før + evt. F/G for lokasjon ---
        for idx, r in enumerate(liste, start=3):
            full_id = r.get("full_id", "") or ""
            clean_id = _clean_display_id(full_id)

            kopi.cell(idx, 4, clean_id)         # D: Komponentnavn (nå ryddig -XX000 eller -XX000%TYPE)
            kopi.cell(idx, 5, r.get("desc", ""))# E: Beskrivelse (uendret)

            col_cursor = 6  # F starter her hvis innsatt
            locs = r.get("locs") or {}

            if primary_label:
                kopi.cell(idx, col_cursor, locs.get(primary_label, "") or "")
                col_cursor += 1

            if has_ktr:
                kopi.cell(idx, col_cursor, locs.get("Kontrollområde", "") or "")

        # --- Trim tomme rader som før ---
        last = 3 + len(liste) - 1
        if last < 200:
            kopi.delete_rows(last + 1, 200 - last)

        add_betingelser(kopi)
        
        end_row = last if last >= 3 else 3
        _apply_status_cf(kopi, start_row=3, end_row=end_row)

    # 3) Fjern malark, sorter faner numerisk på system
    wb.remove(malark)
    wb._sheets.sort(key=lambda s: _sheet_sort_key(s.title))

    # 4) Lever fil
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    username = "Ukjent"
    if hasattr(current_user, "is_authenticated") and current_user.is_authenticated:
        username = getattr(current_user, "first_name", None) or getattr(current_user, "username", None) or "Bruker"

    return send_file(tmp.name, as_attachment=True, download_name=f"{username}-Mekanisk_Komplett.xlsx")
