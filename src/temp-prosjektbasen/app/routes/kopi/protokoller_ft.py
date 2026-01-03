import io
import os
import json
import re
from datetime import datetime
from flask import request, jsonify, send_file, current_app
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule, DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

# Hent alt vi trenger fra core
from .protokoller_core import (
    bp,
    extract_text_from_file,
    parse_rows_from_text,
    TFM_DICT,
    DATA_DIR,
    get_unique_system_id,   # ← viktig for kanonisk system-ID
)

# ──────────────────────────────────────────────────────────────────────────────
# Funksjonstest • GENERER UNDERLAG (REN JSON)
# ──────────────────────────────────────────────────────────────────────────────
@bp.route("/generate_funksjonstest", methods=["POST"])
@login_required
def generate_funksjonstest_underlag():
    """
    Funksjonstest: leser opplastede filer, bruker felles parser og returnerer JSON-liste.
    Filtrerer på system_kriterier (prefiks), og fyller inn komponent/TFM-desc robust.
    """
    files = request.files.getlist("files")
    out: list[dict] = []

    # system_kriterier, f.eks. "36" eller "36, 37"
    raw_criteria = (request.form.get("system_kriterier") or "").strip()
    allowed_prefixes = [s for s in re.split(r"[,\s]+", raw_criteria) if s]

    def _allowed(uid: str) -> bool:
        if not allowed_prefixes:
            return True
        s = str(uid or "")
        return any(s.startswith(pref) for pref in allowed_prefixes)

    # Hjelpere (samme prinsipp som i MC)
    SYS_RE  = re.compile(r"\b\d{3,4}\.\d{3,4}\b")
    COMP_RE = re.compile(r"-?([A-Za-z]{2,4}\d{2,5})(?:%[^\s+]+)?")

    def parse_component_from_full_id(full_id: str) -> str | None:
        s = str(full_id or "")
        m = COMP_RE.search(s)
        return m.group(1) if m else None

    def tfm_prefix_from_component(komp: str) -> str:
        if not komp:
            return ""
        letters = re.match(r"^[A-Za-z]{2,}", komp)
        return (letters.group(0)[:2].upper()) if letters else ""

    def canonical_system(row: dict) -> str:
        # 1) allerede kanonisk?
        uid = row.get("unique_system")
        if uid:
            return uid
        # 2) fra system_number
        sysnum = row.get("system_number") or ""
        if sysnum:
            return get_unique_system_id(sysnum)
        # 3) fra full_id
        fid = row.get("full_id") or ""
        m = SYS_RE.search(fid)
        return get_unique_system_id(m.group(0)) if m else ""

    # Prosesser filer
    for f in files:
        if not f or not f.filename:
            continue
        try:
            try:
                f.seek(0)
            except Exception:
                pass

            text = extract_text_from_file(f)
            for row in parse_rows_from_text(text, f.filename):
                uid = canonical_system(row)
                if not _allowed(uid):
                    continue

                # komponent: prioritér feltet, ellers trekk ut fra full_id
                komp = row.get("komponent") or parse_component_from_full_id(row.get("full_id", "")) or ""
                tfm_key = tfm_prefix_from_component(komp)
                desc = TFM_DICT.get(tfm_key, "Ukjent beskrivelse")

                out.append({
                    "source": row.get("source", f.filename),
                    "unique_system": uid,
                    "system_full_name": row.get("system_full_name", ""),
                    "system_number": row.get("system_number", ""),
                    "full_id": row.get("full_id", ""),
                    "komponent": komp,
                    "funksjonsvalg": row.get("funksjonsvalg", "Øvrig"),
                    "desc": desc,
                })

        except Exception as e:
            out.append({
                "source": getattr(f, "filename", "ukjent"),
                "system_number": "",
                "komponent": "",
                "error": f"{type(e).__name__}: {e}",
            })

    return jsonify(out)


# ──────────────────────────────────────────────────────────────────────────────
# Funksjonstest • LAST NED PROTOKOLL (EXCEL)
# ──────────────────────────────────────────────────────────────────────────────

def _normalize_rows_from_request():
    """
    Godtar både {rows:[...]} og direkte [...].
    Tåler at enkelte rader er JSON-strenger og forsøker å json.loads dem.
    """
    payload = request.get_json(silent=True)

    if isinstance(payload, dict):
        rows = payload.get("rows", [])
    elif isinstance(payload, list):
        rows = payload
    else:
        rows = []

    norm = []
    for r in rows:
        if isinstance(r, dict):
            norm.append(r)
        elif isinstance(r, str):
            try:
                obj = json.loads(r)
                if isinstance(obj, dict):
                    norm.append(obj)
            except Exception:
                # ignorer rå strenger
                continue
    return norm

@bp.route("/download_funksjonstest_protokoll", methods=["POST"])
@login_required
def download_funksjonstest_protokoll():
    from flask import current_app
    from openpyxl.styles import Alignment
    log = current_app.logger

    # ---------- robust payload -> rows ----------
    def _load_payload():
        payload = request.get_json(silent=True)
        if payload is None:
            raw = request.get_data(cache=False, as_text=True)
            try:
                payload = json.loads(raw)
            except Exception:
                payload = raw
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:
                pass
        return payload

    def _norm_rows(payload):
        if isinstance(payload, dict):
            rows = payload.get("rows", [])
        elif isinstance(payload, list):
            rows = payload
        else:
            rows = []
        norm = []
        for r in rows:
            if isinstance(r, dict):
                norm.append(r)
            elif isinstance(r, str):
                try:
                    obj = json.loads(r)
                    if isinstance(obj, dict):
                        norm.append(obj)
                except Exception:
                    pass
        return norm

    payload = _load_payload()
    rows = _norm_rows(payload)
    if not rows:
        return jsonify({"error": "Ingen gyldige rader mottatt."}), 400

    # ---------- last mal ----------
    template_path = os.path.join(DATA_DIR, "Funksjonstest.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Malfil ikke funnet",
                        "hint": f"Letet etter: {template_path}"}), 500
    wb = load_workbook(template_path)

    # Finn malark
    malark = None
    for name in wb.sheetnames:
        if name.lower() in ("mal", "malverk", "template", "ft_mal"):
            malark = wb[name]; break
    if malark is None:
        malark = wb[wb.sheetnames[0]]

    # ---------- grupper per system ----------
    def _sys(r):
        s = r.get("system_number") or r.get("system") or ""
        return str(s).strip() or "Uspesifisert"

    groups = {}
    for r in rows:
        groups.setdefault(_sys(r), []).append(r)

    # ---------- helpers ----------
    START_ROW = 29
    COL = {"A":1, "B":2, "C":3, "E":5, "G":7}
    SECTION_ORDER = [
        "Start og Stopp funksjoner",
        "Reguleringsfunksjoner",
        "Sikkerhetsfunksjoner",
        "Øvrig",
    ]

    def replace_placeholders(ws, system_text: str):
        patt = re.compile(r"\{SYSTEM(NAVN)?\}", re.I)
        for r in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=40):
            for cell in r:
                v = cell.value
                if isinstance(v, str) and patt.search(v):
                    cell.value = patt.sub(system_text, v)

    # Validering + CF for et sett med (start, end)-områder i kolonne A
    def apply_status_validation_and_cf_ranges(ws, ranges):
        dv = DataValidation(
            type="list",
            formula1='"Ikke startet,Under arbeid,Avvik,Utført"',
            allow_blank=True
        )
        ws.add_data_validation(dv)
        for a, b in ranges:
            if a <= b:
                dv.add(f"A{a}:A{b}")

        # Farger
        FILL_GRAY  = PatternFill("solid", start_color="FFF2F2F2", end_color="FFF2F2F2")
        FILL_BLUE  = PatternFill("solid", start_color="FFDDEBF7", end_color="FFDDEBF7")
        FILL_RED   = PatternFill("solid", start_color="FFF8D7DA", end_color="FFF8D7DA")
        FILL_GREEN = PatternFill("solid", start_color="FFD4EDDA", end_color="FFD4EDDA")
        FONT_BLACK = Font(color="FF000000")
        FONT_WHITE = Font(color="FFFFFFFF")

        def rule(expr, fill, font):
            dxf = DifferentialStyle(fill=fill, font=font)
            return Rule(type="expression", dxf=dxf, stopIfTrue=False, formula=[expr])

        # Slå sammen alle intervaller til ett CF-område (går fint)
        if ranges:
            a_min = min(a for a,_ in ranges)
            b_max = max(b for _,b in ranges)
            rng = f"A{a_min}:A{b_max}"
            ws.conditional_formatting.add(rng, rule('=EXACT($A1,"Ikke startet")', FILL_GRAY,  FONT_BLACK))
            ws.conditional_formatting.add(rng, rule('=EXACT($A1,"Under arbeid")', FILL_BLUE,  FONT_BLACK))
            ws.conditional_formatting.add(rng, rule('=EXACT($A1,"Avvik")',        FILL_RED,   FONT_WHITE))
            ws.conditional_formatting.add(rng, rule('=EXACT($A1,"Utført")',       FILL_GREEN, FONT_BLACK))

    # Seksjonsrad-stil
    SECTION_FILL = PatternFill("solid", start_color="FFDCEBFF", end_color="FFDCEBFF")  # lys pastell blå
    SECTION_FONT = Font(bold=True)
    SECTION_ALIGN = Alignment(horizontal="center", vertical="center")

    def safe_merge_row(ws, row, start_col=1, end_col=14):
        """Unmerger alle ranges som overlapper row/start_col..end_col og merger deretter raden."""
        to_unmerge = []
        for cr in list(ws.merged_cells.ranges):
            if (cr.min_row <= row <= cr.max_row) and not (cr.max_col < start_col or cr.min_col > end_col):
                to_unmerge.append(str(cr))
        for cr in to_unmerge:
            ws.unmerge_cells(cr)
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        
    # ---------- lag ark per system og fyll ----------
    for system, liste in groups.items():
        title = f"{system} - FT"
        ws = wb[title] if title in wb.sheetnames else wb.copy_worksheet(malark)
        ws.title = title

        replace_placeholders(ws, system)

        # Tøm rimelig dataområde (A–N) fra rad 29 og ned 1000 rader
        for rr in range(START_ROW, START_ROW + 1000):
            for cc in range(1, 14 + 1):  # A..N
                ws.cell(rr, cc).value = None

        r = START_ROW
        data_row_ranges = []  # for DV/CF (utelukker seksjonsrader)

        def flush_range(start, end):
            if start is not None and end is not None and start <= end:
                data_row_ranges.append((start, end))

        # Bygg hver seksjon
        for section in SECTION_ORDER:
            # Seksjonsrad (alltid – for å speile UI)
            safe_merge_row(ws, r, 1, 14)  # A..N
            cell = ws.cell(r, 1)
            cell.value = section
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
            cell.alignment = SECTION_ALIGN
            r += 1

            # Datarader i denne seksjonen
            section_rows = [obj for obj in liste if ((obj.get("funksjonsvalg") or obj.get("integrert") or "Øvrig").strip()) == section]
            if section_rows:
                range_start = r
                for obj in section_rows:
                    system_txt = _sys(obj)
                    komponent  = obj.get("komponent", "")
                    test       = obj.get("testutfoerelse", "") or obj.get("test", "")
                    aksept     = obj.get("aksept", "") or obj.get("forventet_resultat", "")
                    status_val = obj.get("status") or "Ikke startet"

                    ws.cell(r, COL["A"], status_val)  # Status
                    ws.cell(r, COL["B"], system_txt)  # System
                    ws.cell(r, COL["C"], komponent)   # Komponent
                    ws.cell(r, COL["E"], test)        # Testutførelse
                    ws.cell(r, COL["G"], aksept)      # Akseptkriterie
                    r += 1
                flush_range(range_start, r - 1)

        # Legg datavalidering og CF kun på datarader (ikke seksjonsrader)
        apply_status_validation_and_cf_ranges(ws, data_row_ranges)

    # --- Fjern malarket helt til slutt ---
    try:
        if malark.title in wb.sheetnames:
            wb.remove(wb[malark.title])
    except Exception as e:
        log.warning("Kunne ikke fjerne malarket %r: %s", malark.title, e)

    # Sorter ark alfabetisk
    try:
        wb._sheets = [wb[n] for n in sorted(wb.sheetnames)]
    except Exception:
        pass

    # ---------- send fil ----------
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"Funksjonstest_Protokoll_{datetime.utcnow():%Y%m%d_%H%M%S}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )