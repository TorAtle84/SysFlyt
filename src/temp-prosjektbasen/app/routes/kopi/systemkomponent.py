import io
import json
import re
import fitz # PyMuPDF
import docx
from flask import Blueprint, request, jsonify, send_file, Response, stream_with_context, render_template
from flask_login import login_required
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook
from pathlib import Path
from app.models.project import Project
from app.models.systembygging import Systembygging
from app.models.komponentopptelling import Komponentopptelling
from app.models.db import db

systemkomponent_bp = Blueprint("systemkomponent", __name__, url_prefix="/systemkomponent")

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_ROOT = BASE_DIR / "temp"
TEMP_ROOT.mkdir(exist_ok=True)

TFM_DICT = {
  "AB": "Bjelke",
  "AD": "Dekker",
  "AE": "Elementer",
  "AF": "Fagverk",
  "AG": "Glassfelt",
  "AH": "Fundamenter",
  "AK": "Komplette konstruksjoner",
  "AL": "List - Beslistning",
  "AO": "Oppbygende - Utforende",
  "AP": "Plate",
  "AR": "Ramme - Oppheng",
  "AS": "Søyle",
  "AU": "Åpninger",
  "AV": "Vegg",
  "BA": "Armering Forsterkning",
  "BB": "Beskyttende - Stoppende",
  "BC": "Begrensende",
  "BF": "Fuging",
  "BG": "Pakning Tetning",
  "BI": "Isolasjon",
  "BP": "Avretningsmasse",
  "BS": "Spikerslag",
  "CB": "Balkong",
  "CC": "Baldakin",
  "CD": "Karnapp",
  "CG": "Rampe - Repos",
  "CK": "Komplett konstruksjon",
  "CM": "Kjemiske stoffer",
  "CO": "Kobling Overgang",
  "CP": "Pipe - Skostein",
  "CQ": "Festemateriell",
  "CR": "Rammeverk - Oppheng",
  "CT": "Trapp - Leider",
  "CX": "Tunnel - Bru (inne-ute)",
  "DB": "Dør med brannklasse",
  "DF": "Foldedør - Foldevegg",
  "DI": "Dør - innvendig",
  "DK": "Kjøretøyregulering",
  "DL": "Luke",
  "DP": "Port",
  "DT": "Dør - tilfluktsrom",
  "DU": "Dør - utvendig",
  "DV": "Vindu",
  "EB": "Overflatebekledning",
  "EC": "Overflatebehandling",
  "EH": "Himling",
  "FA": "Ventilert arbeidsplass",
  "FB": "Benk - Bord - Plate - Tavle",
  "FC": "Beslag",
  "FD": "Disk - Skranke",
  "FF": "Fryserom",
  "FH": "Hylle - Reol",
  "FI": "Kabinett",
  "FK": "Kjølerom - Svalrom",
  "FO": "Sittebenk - Sofa - Stol",
  "FR": "Rom",
  "FS": "Skap - Skuff",
  "FT": "Speil",
  "FV": "Vaskemaskin",
  "FX": "Krok - Knagg - Håndtak",
  "GA": "Automat - Maskin",
  "GB": "Benk - Bord - Plate - Tavle",
  "GD": "Dekontamator",
  "GE": "Autoklaver",
  "GF": "Fryseskap",
  "GG": "Gardiner - Forheng",
  "GH": "Hylle - Reol",
  "GK": "Kjøleskap - Kjøledisk",
  "GL": "Lås - Beslag",
  "GM": "Mattilbereding",
  "GN": "Nøkler",
  "GO": "Sofa - Sittebenk",
  "GP": "Stol",
  "GQ": "Seng - Liggebenk",
  "GS": "Skap - Skuffer",
  "GT": "Tørkeskap - Varmeskap",
  "GV": "Vaskemaskin",
  "GW": "Vekt",
  "GX": "Holder",
  "GY": "Avfallsbeholder",
  "HA": "Automobil - Bil",
  "HB": "Rullebord - Tralle",
  "HC": "Container - Vogn",
  "HM": "Maskin",
  "HS": "Rullestol",
  "HT": "Truck - kran",
  "HV": "Verktøy",
  "IB": "Brenner",
  "IC": "Solceller-solfangere",
  "ID": "Kjel for destruksjon",
  "IE": "Elektrokjel",
  "IF": "Kjel for fast-bio brensel",
  "IG": "Generator",
  "IK": "Kuldeaggregat",
  "IL": "Energibrønn",
  "IM": "Motor",
  "IO": "Oljekjel",
  "IP": "Gasskjel",
  "IT": "Trykkluftaggregat (enhet)",
  "IU": "Turbin",
  "IV": "Aggregatenhet",
  "JF": "Forsterker",
  "JK": "Kompressor",
  "JP": "Pumpe",
  "JQ": "Pumpe i VA-installasjoner",
  "JV": "Vifte",
  "JW": "Spesialvifte",
  "KA": "Aktuator",
  "KD": "Drivhjul - Drev",
  "KE": "Induktiv energioverføring",
  "KG": "Gjennomføring",
  "KH": "Transportenhet (hevende forflyttende)",
  "KJ": "Jordingskomponenter",
  "KK": "Kanal",
  "KM": "Mast - Antenne",
  "KN": "Nedløp",
  "KO": "Kraftoverføring",
  "KQ": "Rør - spesielt",
  "KR": "Rør - generelt",
  "KS": "Skinne - Bane - Spor",
  "KU": "Kombinert kabel",
  "KV": "Høyspenningskabel > 1000V",
  "KW": "Lavspenningskabel 50 til 1000V",
  "KX": "Lavspenningskabel < 50V",
  "KY": "Optisk kabel",
  "KZ": "Slange",
  "LB": "Varmeomformende med vifte",
  "LC": "Kjøleomformende med vifte",
  "LD": "Kjøleflater",
  "LE": "Kondensator",
  "LF": "Fordamper",
  "LG": "Gear clutch",
  "LH": "Varmeflate",
  "LI": "Varmeelement",
  "LK": "Kjøleomformende",
  "LL": "Lyskilde",
  "LN": "Likeretter",
  "LO": "Omformer",
  "LP": "Pens - Veksel - Sjalter",
  "LQ": "Vekselretter",
  "LR": "Frekvensomformer",
  "LS": "Strålevarme",
  "LU": "Luftfukter",
  "LV": "Varmeomformende",
  "LX": "Varmegjenvinner",
  "LZ": "Varmerkabel - Varmerør",
  "MA": "Absoluttfilter",
  "MB": "ABC-filter",
  "MC": "UV-filter",
  "ME": "Elektrostatiske filter",
  "MF": "Luftfilter",
  "MG": "Fettfilter",
  "MK": "Kondenspotte",
  "ML": "Luftutskiller",
  "MM": "Membran",
  "MO": "Utskiller",
  "MR": "Rist - Sil",
  "MS": "Syklon",
  "MT": "Tørke",
  "MU": "Filter for Lyd - Bilde - Frekvensutjevner",
  "MV": "Vannfilter",
  "MX": "Støyfilter",
  "NB": "Batteri - UPS",
  "NC": "Kondensator",
  "NI": "Informasjonslagring",
  "NK": "Kum",
  "NM": "Badekar - Basseng",
  "NO": "Åpen tank",
  "NT": "Tank med trykk",
  "NU": "Tank uten trykk",
  "NV": "Vekt Lodd",
  "NW": "Varmtvannsbereder",
  "NX": "Toalett",
  "NY": "Servant",
  "NZ": "Brannslukkingsapparat",
  "OA": "AV-maskiner",
  "OB": "Shuntgruppe",
  "OD": "Datamaskin",
  "OE": "Energimåler",
  "OF": "Systembærer",
  "OM": "Mottaker - Sender",
  "OP": "PBX",
  "OQ": "Dataprogramprogramvare",
  "OR": "Router Fordeler",
  "OS": "Sentralenhet Mikser i Lydsystem",
  "OT": "Telefonapparat",
  "OU": "Undersentral",
  "QB": "Belastningsvakt",
  "QD": "Differansetrykkvakt",
  "QE": "Elektrisk vern",
  "QF": "Strømningsvakt",
  "QG": "",
  "QH": "Fuktvakt",
  "QI": "",
  "QJ": "",
  "QK": "",
  "QL": "Lyddemper",
  "QM": "Mekanisk beskyttelse",
  "QN": "Nivåvakt",
  "QO": "Overtrykksventil",
  "QP": "Trykkvakt",
  "QQ": "Vibrasjonsvakt",
  "QR": "Rotasjonsvakt",
  "QS": "Strømvakt",
  "QT": "Temperaturvakt",
  "QU": "",
  "QV": "Sikkerhetsventil",
  "QW": "",
  "QX": "Solavskjerming",
  "QY": "Lynavleder",
  "QZ": "Brannvern",
  "RA": "AV-opptaker",
  "RB": "Bevegelse",
  "RC": "Seismometer",
  "RD": "Differansetrykkgiver",
  "RE": "Elektriske variabler",
  "RF": "Strømningsmåler",
  "RG": "Posisjon - Lengde",
  "RH": "Fuktighetsgiver",
  "RI": "Termometer",
  "RJ": "Fotocelle",
  "RK": "Kortleser",
  "RL": "",
  "RM": "Multifunksjonell Kombinert føler",
  "RN": "Nivågiver",
  "RO": "",
  "RP": "Trykkgiver",
  "RQ": "Manometer Trykkmåler",
  "RR": "Giver generelt",
  "RS": "Hastighetsmåler",
  "RT": "Temperaturgiver",
  "RU": "Ur",
  "RV": "Veieceller",
  "RW": "Virkningsgradsmåler",
  "RX": "Målepunkt",
  "RY": "Gassdetektor-Røykdetektor",
  "RZ": "Branndeteksjon",
  "SA": "Reguleringsventil manuell",
  "SB": "Reguleringsventil motorstyrt",
  "SC": "Stengeventil motorstyrt",
  "SD": "Alarmventil sprinkler",
  "SE": "Ekspansjonsventil",
  "SF": "Fraluftsventil",
  "SG": "Tilbakeslagsventil - Overtrykkspjeld",
  "SH": "Hurtigkobling",
  "SI": "Effektregulator",
  "SJ": "Jevntrykksventil",
  "SK": "Strømningsregulator - CAV",
  "SL": "",
  "SM": "Stengeventil manuell",
  "SN": "",
  "SO": "",
  "SP": "Trykkutjevningsventil",
  "SQ": "Strømningsregulator - VAV",
  "SR": "Reguleringsspjeld",
  "SS": "Stengespjeld",
  "ST": "Tilluftsventil",
  "SU": "Sugetrykksventil",
  "SV": "Strupeventil",
  "SW": "Plenumskammer",
  "SX": "Regulator",
  "SY": "",
  "SZ": "Brannspjeld - Røykspjeld",
  "UA": "Uttak alarm",
  "UB": "Blandebatteri",
  "UC": "",
  "UD": "Uttak data",
  "UE": "Uttak el",
  "UF": "Fellesuttak",
  "UG": "Uttak gass",
  "UH": "Høyttaler",
  "UI": "",
  "UJ": "Skriver",
  "UK": "Kontrollpanel - Tablå",
  "UL": "Uttak trykkluft",
  "UM": "Monitor - Display",
  "UN": "Nødbelysning",
  "UO": "Trommel",
  "UP": "Belysningsarmatur",
  "UQ": "",
  "UR": "Uttak radio",
  "US": "Stasjon",
  "UT": "Uttak telefon",
  "UU": "",
  "UV": "Uttak vann",
  "UW": "",
  "UX": "Koblingsboks",
  "UY": "Uttak antenne",
  "UZ": "Dyse - Spreder",
  "VA": "",
  "VB": "Bærelag",
  "VC": "",
  "VD": "Dekke",
  "VE": "",
  "VF": "",
  "VG": "Gress",
  "VH": "",
  "VI": "",
  "VJ": "",
  "VK": "Kantstein heller",
  "VL": "Masse",
  "VM": "Mekanisk beskyttelse",
  "VN": "",
  "VO": "",
  "VP": "Planterbuskertrær",
  "VQ": "",
  "VR": "",
  "VS": "Skilt",
  "VT": "",
  "VU": "",
  "VV": "",
  "VW": "",
  "VX": "",
  "VY": "",
  "VZ": "",
  "XA": "",
  "XB": "",
  "XC": "Kondensator",
  "XD": "Komp. for binærlogikk",
  "XE": "",
  "XF": "Komponenter for vern",
  "XG": "Komponenter for krafttilførsel",
  "XH": "Komponenter for signalering",
  "XI": "Potensiometer",
  "XJ": "",
  "XK": "Releer - Kontaktorer",
  "XL": "Induktiv komponenter",
  "XM": "Motor",
  "XN": "Integrerte kretser",
  "XO": "Urbryter - Timer",
  "XP": "Komponenter for måling og prøving",
  "XQ": "Effektbryter",
  "XR": "Motstand",
  "XS": "Bryter / Vender",
  "XT": "Transformator",
  "XU": "",
  "XV": "Halvlederkomponenter og elektronrør",
  "XW": "",
  "XX": "Rekkeklemmer - Samlesignal",
  "XY": "",
  "XZ": "Terminering og tilpasning",
  "YA": "",
  "YB": "",
  "YC": "",
  "YD": "",
  "YE": "",
  "YF": "",
  "YG": "",
  "YH": "",
  "YI": "",
  "YJ": "",
  "YK": "",
  "YL": "",
  "YM": "",
  "YN": "",
  "YO": "",
  "YP": "",
  "YQ": "",
  "YR": "",
  "YS": "",
  "YT": "",
  "YU": "",
  "YV": "",
  "YW": "",
  "YX": "",
  "YY": "",
  "YZ": "",
  "ZA": "",
  "ZB": "",
  "ZC": "",
  "ZD": "",
  "ZE": "",
  "ZF": "",
  "ZG": "",
  "ZH": "",
  "ZI": "",
  "ZJ": "",
  "ZK": "",
  "ZL": "",
  "ZM": "",
  "ZN": "",
  "ZO": "",
  "ZP": "",
  "ZQ": "",
  "ZR": "",
  "ZS": "",
  "ZT": "",
  "ZU": "",
  "ZV": "",
  "ZW": "",
  "ZX": "",
  "ZY": "",
  "ZZ": ""
}

@systemkomponent_bp.route("/")
@login_required
def index():
    prosjekter = Project.query.order_by(Project.project_name.asc()).all()
    return render_template("systemkomponent.html", prosjekter=prosjekter)

def extract_text(f):
    """
    Ekstraherer tekst fra ulike filtyper (PDF, DOCX, XLSX, TXT).
    Støtter feilhåndtering for ulike filformater.
    """
    name = f.filename.lower()
    data = f.read()
    bio = io.BytesIO(data)

    if name.endswith(".pdf"):
        try:
            doc = fitz.open(stream=data, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            return text
        except Exception as e:
            print(f"Feil ved lesing av PDF {f.filename}: {e}")
            return ""

    elif name.endswith(".docx"):
        try:
            docx_file = docx.Document(bio)
            return "\n".join([p.text for p in docx_file.paragraphs])
        except Exception as e:
            print(f"Feil ved lesing av DOCX {f.filename}: {e}")
            return ""

    elif name.endswith(".xlsx"):
        try:
            wb = load_workbook(filename=bio, data_only=True)
            # Deaktiver alle aktive filtre i Excel-filen
            for ws in wb.worksheets:
                if ws.auto_filter.ref: # Check if autofilter is set
                    ws.auto_filter.ref = None # Deactivate autofilter
            
            text = ""
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    # Filter ut None verdier for å unngå "None" i teksten
                    line = " ".join([str(cell) if cell is not None else "" for cell in row])
                    text += line + "\n"
            return text
        except Exception as e:
            print(f"Feil ved lesing av XLSX {f.filename}: {e}")
            return ""

    elif name.endswith(".txt"):
        try:
            return data.decode(errors="ignore")
        except Exception as e:
            print(f"Feil ved lesing av TXT {f.filename}: {e}")
            return ""
    else:
        print(f"Ukjent filtype: {f.filename}. Hopper over.")
        return ""


# Hjelpefunksjon for å bygge det regulære uttrykket
def build_dynamic_regex(format_string):
    """
    Bygger et regulært uttrykk fra en formatstreng (f.eks. "{byggnr}{system}{komponent}")
    ved å erstatte plassholdere med deres tilsvarende regex-mønstre, inkludert prefikser.
    System- og Typekode-segmentene blir gjort valgfrie.
    """
    # Define regex patterns for the *content* inside each placeholder, WITHOUT prefixes.
    _content_regex_patterns = {
        "{byggnr}": r"(?P<byggnr>[A-Za-z0-9]+)",
        "{system}": r"(?P<system>[^- %\n]*)", # Content can be empty
        "{komponent}": r"(?P<komponent>[^ %\n]+)", # Content must exist (at least one character)
        "{typekode}": r"(?P<typekode>[^\s]*)" # Content can be empty
    }

    final_regex_parts = []
    
    current_string_index = 0
    
    while current_string_index < len(format_string):
        matched_placeholder = False
        
        if format_string[current_string_index:].startswith("{byggnr}"):
            # Modified: Allow one or more '+' signs for byggnr
            final_regex_parts.append(r"\++" + _content_regex_patterns["{byggnr}"])
            current_string_index += len("{byggnr}")
            matched_placeholder = True
        elif format_string[current_string_index:].startswith("{system}"):
            final_regex_parts.append(r"(?:" + re.escape('=') + _content_regex_patterns["{system}"] + ")?")
            current_string_index += len("{system}")
            matched_placeholder = True
        elif format_string[current_string_index:].startswith("{komponent}"):
            final_regex_parts.append(re.escape('-') + _content_regex_patterns["{komponent}"])
            current_string_index += len("{komponent}")
            matched_placeholder = True
        elif format_string[current_string_index:].startswith("{typekode}"):
            final_regex_parts.append(r"(?:" + re.escape('%') + _content_regex_patterns["{typekode}"] + ")?")
            current_string_index += len("{typekode}")
            matched_placeholder = True
        
        if not matched_placeholder:
            final_regex_parts.append(re.escape(format_string[current_string_index]))
            current_string_index += 1
            
    return "".join(final_regex_parts)


# === SYSTEMBYGGING ===
@systemkomponent_bp.route("/systembygging", methods=["POST"])
def systembygging():
    """
    Behandler opplasting av filer for systembygging.
    Identifiserer og lister opp komponent-ID-er basert på et gitt format og kriterier.
    Hvert søketreff vil returneres som en egen rad.
    """
    files = request.files.getlist("files")
    kriterier_str = request.form.get("kriterier", "")
    system_valg = [s.strip() for s in kriterier_str.split(',') if s.strip()]
    formatval = request.form.get("format", "") # F.g.s. "{byggnr}{system}{komponent}{typekode}" (fra data-placeholder)

    # Bygg det endelige regulære uttrykket ved hjelp av hjelpefunksjonen
    regex_str = build_dynamic_regex(formatval)
    pattern = re.compile(regex_str)

    def generate():
        rows = []
        for f in files:
            yield json.dumps({"currentFile": f.filename}) + "\n"
            text = extract_text(f)

            if not text:
                continue # Hopp over tomme eller uleste filer

            # Itererer over alle funne treff i teksten
            for m in pattern.finditer(text):
                gr = m.groupdict() # Hent ut de navngitte gruppene fra treffet

                # Hent og strip bort whitespace fra de ekstraherte verdiene.
                byggnr = (gr.get("byggnr") or "").strip()
                system = (gr.get("system") or "").strip()
                komponent = (gr.get("komponent") or "").strip()
                typekode = (gr.get("typekode") or "").strip() 

                # NY LOGIKK FOR FILTERING:
                # Inkluder raden KUN hvis:
                # 1. Komponent er til stede (alltid obligatorisk for en meningsfull rad).
                # 2. HVIS {system} er inkludert i formatet:
                #    Da må ENTEN system_ID (verdien) være til stede, ELLER (system_ID er tom *men* prefikset '=' ble funnet).
                #    Dette forhindrer rader med bare "=" og ingen komponent.
                #    Hvis {system} IKKE er inkludert i formatet, da gjelder ikke denne sjekken.

                # Sjekk 1: Komponent må alltid være til stede for at raden skal være meningsfull
                if not komponent:
                    continue

                # Sjekk 2: Hvis {system} er en del av formatet, men ingen systemverdi ble funnet,
                # og det heller ikke er en tom streng fra en '='-match (alltså, '=' var ikke til stede heller).
                # Vi antar at 'system' i `gr` kun er tomt hvis '=' ble matchet men innholdet var tomt.
                # Hvis 'system' *ikke* er i `gr`, betyr det at hele segmentet (?:=...) ikke ble matchet.
                if '{system}' in formatval and not system and 'system' not in gr:
                    # Dette fanger tilfellet der formatet forventet et system,
                    # men ingen del av systemsegmentet ble funnet i teksten.
                    # Vi vil ikke vise rader som '+BYGGNR-KOMPONENT' hvis formatet var '+{byggnr}={system}-{komponent}'
                    # og det ikke var noe '=SYSTEM' i teksten.
                    continue


                # Fjerner '=' fra starten av system-strengen for matching mot kriterier
                system_id_for_criteria = system
                if system_valg:
                    if not system_id_for_criteria or system_id_for_criteria[:2] not in system_valg:
                        continue 


                # REKONSTRUER current_full_id basert på formatval og faktiske fangede verdier
                final_id_builder = []
                idx = 0
                while idx < len(formatval):
                    substituted = False
                    
                    if formatval[idx:].startswith("{byggnr}"):
                        # Find the actual number of '+' characters from the start of the match
                        num_plus_signs = 0
                        if m.group(0).startswith('+'):
                            part_before_byggnr = m.group(0).split(byggnr)[0] if byggnr else m.group(0) 
                            num_plus_signs = part_before_byggnr.count('+')
                        
                        final_id_builder.append("+" * num_plus_signs + byggnr)
                        idx += len("{byggnr}")
                        substituted = True
                    elif formatval[idx:].startswith("{system}"):
                        # Append "={system}" if system was captured by the regex (i.e., 'system' is a key in gr)
                        if "system" in gr: 
                             final_id_builder.append("=" + system) 
                        idx += len("{system}")
                        substituted = True
                    elif formatval[idx:].startswith("{komponent}"):
                        final_id_builder.append("-" + komponent)
                        idx += len("{komponent}")
                        substituted = True
                    elif formatval[idx:].startswith("{typekode}"):
                        if "typekode" in gr: 
                            final_id_builder.append("%" + typekode) 
                        idx += len("{typekode}")
                        substituted = True
                    
                    if not substituted:
                        final_id_builder.append(formatval[idx])
                        idx += 1
                
                current_full_id = "".join(final_id_builder)


                # Bestem beskrivelse basert på komponentens TFM-prefix
                actual_komponent_id = komponent.lstrip('-') 
                prefix = actual_komponent_id[0:2] if len(actual_komponent_id) >= 2 else "" 
                desc = TFM_DICT.get(prefix, "Ukjent")

                # Legg til hvert enkelt funnet treff som en separat rad, inkludert unique_system_key
                rows.append({
                    "full_id": current_full_id,
                    "desc": desc,
                    "system": system, # The full captured system string (e.g., '3601.0005:02')
                    "component": actual_komponent_id, 
                    "files": [f.filename],
                    "unique_system_key": get_unique_system_id(system) # NEW: Add the unique key for frontend/Excel grouping
                })
        
        # Sorter de genererte radene for konsistent visning på klienten
        rows_sorted = sorted(rows, key=lambda x: (x["system"], x["component"]))
        yield json.dumps({"rows": rows_sorted})

    return Response(stream_with_context(generate()), mimetype="text/plain")

# Hjelpefunksjon for å ekstrahere den unike system-ID-en for gruppering
# Skal hente systemnavnet frem til første spesialtegn som ':' eller '*'
# eller frem til '-' hvis det er det første spesialtegnet.
def get_unique_system_id(system_string):
    if not system_string:
        return "" # For tomme system-strenger (missing system)
    
    # Define a set of delimiters that mark the end of the "unique" part of the system ID
    # Order matters: check for ':' first, then '*', then '-'
    delimiters_regex = r'[:*\-]' # Match any of ':', '*', '-'

    match = re.search(delimiters_regex, system_string)
    if match:
        return system_string[:match.start()] # Return substring before the delimiter
    
    return system_string # If no delimiter is found, return the whole string


# Function to adjust column widths for a given worksheet
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter (e.g., 'A')
        for cell in col:
            try:
                # Get the string representation of the cell value
                cell_value_str = str(cell.value)
                if cell_value_str:
                    max_length = max(max_length, len(cell_value_str))
            except TypeError: 
                pass
        
        adjusted_width = (max_length + 2) # Add a small padding
        ws.column_dimensions[column].width = adjusted_width

# Function to apply cell formatting (text format and wrap text)
def apply_cell_format(cell):
    cell.number_format = '@'  # Text format
    cell.alignment = Alignment(wrap_text=True)


# === KOMPONENTOPPTELLING ===
@systemkomponent_bp.route("/komponentopptelling", methods=["POST"])
def komponentopptelling():
    """
    Behandler opplasting av filer for komponentopptelling.
    Teller forekomster av komponenter basert på TFM-prefix og samler dem.
    """
    files = request.files.getlist("files")
    kriterier_str = request.form.get("kriterier", "")
    system_valg = [s.strip() for s in kriterier_str.split(',') if s.strip()]
    formatval = request.form.get("format", "") # Format for gjenkjenning av komponenter

    # Bygg det endelige regulære uttrykket ved hjelp av hjelpefunksjonen
    regex_str = build_dynamic_regex(formatval)
    pattern = re.compile(regex_str)

    # Dictionaries for å lagre opptellingen, skilt på om komponenten har system-ID eller ikke
    unique_component_with_system = {} 
    unique_component_without_system = {} 

    def generate():
        for f in files:
            yield json.dumps({"currentFile": f.filename}) + "\n"
            text = extract_text(f)

            if not text:
                continue 

            for m in pattern.finditer(text):
                gr = m.groupdict()
                system = (gr.get("system") or "").strip()
                komponent = (gr.get("komponent") or "").strip() 
                
                # Bestem om komponenten har en system-ID
                has_system = bool(system) 

                # Ekstraherer den faktiske komponent-IDen (f.eks. "KA001", "KAB")
                actual_komponent_id = komponent.lstrip('-')

                # NY OG FORBEDRET LOGIKK for aggregeringsnøkkel for komponentopptelling
                # Mål: "SQ012T/003" -> "SQ", "SQZ012T/003" -> "SQZ"
                # Kun tell komponenter der de to første bokstavene (TFM-prefix) finnes i TFM_DICT.
                
                tfm_prefix = actual_komponent_id[0:2] if len(actual_komponent_id) >= 2 else ""

                # Filtrer: Kun prosesser hvis TFM-prefixet finnes i TFM_DICT og består av bokstaver
                if not tfm_prefix.isalpha() or tfm_prefix not in TFM_DICT:
                    continue # Hopp over komponenter som ikke har gyldig TFM-prefix
                
                # Nå som vi vet vi har et gyldig TFM-prefix, bestemmer vi aggregeringsnøkkelen.
                # Nøkkelen skal være TFM-prefixet pluss eventuelle bokstaver som følger, frem til første tall.
                
                comp_aggregation_key = tfm_prefix # Start med de to første bokstavene (TFM)

                # Regex for å fange alfanumeriske tegn (bokstaver og tall) umiddelbart etter TFM-prefixet,
                # frem til første spesialtegn som ikke er tall eller bokstav.
                # Dette skal håndtere 'SQ012T' og 'SQZ012T' for å få 'SQ' og 'SQZ'.
                # Vi ignorerer tall helt i aggregeringsnøkkelen.

                # Finn den delen av actual_komponent_id som starter med tfm_prefix
                # og deretter bare fortsetter med bokstaver (og ev. bindestrek) til et tall eller annet spesialtegn
                
                # Use re.match to find the longest alphanumeric prefix
                # This should capture "SQ" from "SQ012", "SQZ" from "SQZ012", "KAA" from "KAA"
                # Pattern: ^(TFM_prefix)([A-Za-z\-]*).* (capture TFM + following letters/dashes)
                # Then check if this full matched prefix is present at the start of actual_komponent_id
                
                # Example: "SQ012T/003"
                # tfm_prefix = "SQ"
                # We want the key "SQ".
                # If "SQZ012T/003"
                # tfm_prefix = "SQ"
                # We want the key "SQZ".
                
                # Let's try to extract the non-numeric part immediately following the tfm_prefix.
                # Find the first character that is a digit or a non-alphanumeric special character (excluding dot and dash if allowed in name)
                
                # Simplified approach: Capture the TFM prefix and any subsequent letters, until a digit or another non-letter/non-dash character.
                # The assumption is that component names like "SQZ" are formed by TFM + additional letters.
                
                # Use a more explicit regex on the actual_komponent_id to capture the aggregation key:
                # Capture two letters (TFM), then any number of letters or dashes, stopping at the first digit or other special character.
                # This ensures "SQ" from "SQ012T", "SQZ" from "SQZ012T", "KA-ABC" from "KA-ABC123".
                
                key_match = re.match(r'([A-Za-z]{2}[A-Za-z\-]*)', actual_komponent_id)
                if key_match:
                    extracted_key = key_match.group(1)
                    # Now validate if this extracted key starts with the actual tfm_prefix.
                    if extracted_key.startswith(tfm_prefix):
                        comp_aggregation_key = extracted_key
                    else: # Fallback if regex match doesn't align with tfm_prefix
                        comp_aggregation_key = tfm_prefix
                else: # No match found (e.g. "0123" or "!!AB")
                    comp_aggregation_key = tfm_prefix # Use just the TFM prefix as fallback
                
                # Final check to ensure the comp_aggregation_key is not empty or purely numeric if it shouldn't be.
                # After the above logic, it should contain a valid TFM prefix.
                if not comp_aggregation_key:
                    comp_aggregation_key = "Ukjent Komponent" # Should not typically be hit if tfm_prefix check passed


                desc = TFM_DICT.get(tfm_prefix, "Ukjent") 


                # Filtrer basert på systemkriterier, kun hvis komponenten har en system-ID
                if system_valg and has_system:
                    system_id_for_criteria = system 
                    if not system_id_for_criteria or system_id_for_criteria[:2] not in system_valg:
                        continue
                
                # Velg riktig target dictionary basert på om system-ID er tilstede
                target = unique_component_with_system if has_system else unique_component_without_system

                # Bruker den nye aggregeringsnøkkelen
                if comp_aggregation_key not in target:
                    target[comp_aggregation_key] = {"per_file": {}, "files": set()}
                
                # Oppdater tellingen for den aktuelle filen for denne aggregeringsnøkkelen
                if f.filename not in target[comp_aggregation_key]["per_file"]:
                    target[comp_aggregation_key]["per_file"][f.filename] = 0
                target[comp_aggregation_key]["per_file"][f.filename] += 1
                
                # Legg til filnavnet i settet av filer der denne komponenten er funnet
                target[comp_aggregation_key]["files"].add(f.filename)

        # Konverter de aggregerte dataene til en liste av rader for frontend
        rows_to_send = []
        for target_dict, has_system_flag in [(unique_component_with_system, "Ja"), (unique_component_without_system, "Nei")]:
            # Sorterer nøklene alfabetisk for konsistent rekkefølge i tabellen
            for agg_id in sorted(target_dict.keys()): 
                data = target_dict[agg_id]
                desc = TFM_DICT.get(agg_id[0:2] if len(agg_id) >=2 else "", "Ukjent") 
                
                rows_to_send.append({
                    "id": agg_id, 
                    "desc": desc,
                    "per_file": data["per_file"],
                    "files": list(data["files"]), 
                    "has_system": has_system_flag
                })
        
        # Sorter radene for konsistent visning (primær sortering på 'has_system', sekundær på 'id')
        rows_to_send.sort(key=lambda x: (x["has_system"], x["id"]))

        yield json.dumps({"rows": rows_to_send})

    return Response(stream_with_context(generate()), mimetype="text/plain")


# Function to adjust column widths for a given worksheet
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                cell_value_str = str(cell.value)
                if cell_value_str:
                    max_length = max(max_length, len(cell_value_str))
            except TypeError: 
                pass
        
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column].width = adjusted_width

# Function to apply cell formatting (text format and wrap text)
def apply_cell_format(cell):
    cell.number_format = '@' 
    cell.alignment = Alignment(wrap_text=True)


# === GENERER EXCEL ===
@systemkomponent_bp.route("/excel", methods=["POST"])
def generate_excel():
    """
    Genererer en Excel-fil basert på de aggregerte resultatene.
    Oppretter separate ark for hvert system hvis det er systembygging,
    ellers et enkelt ark for komponentopptelling.
    """
    data = request.json or {}
    rows = data.get("rows", [])

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"]) 

    is_system_building = rows and "full_id" in rows[0]
    download_filename = "Systembygging.xlsx" if is_system_building else "Komponentopptelling.xlsx"

    if is_system_building:
        systems = {}
        missing_system_rows = [] 

        for row in rows:
            sys_for_grouping = get_unique_system_id(row.get("system", "")) 

            if not row.get("system"): 
                missing_system_rows.append(row) 
            else:
                if sys_for_grouping not in systems:
                    systems[sys_for_grouping] = []
                systems[sys_for_grouping].append(row) 
                
        for sys_name, comp_data_list in sorted(systems.items()): 
            sheet_name_base = sys_name[:31] 
            invalid_sheet_chars = [":", "/", "\\", "?", "*", "[", "]"]
            cleaned_sheet_name = sheet_name_base
            for char in invalid_sheet_chars:
                cleaned_sheet_name = cleaned_sheet_name.replace(char, "_")
            sheet_name = cleaned_sheet_name

            if not sheet_name:
                sheet_name = "Ukjent System"
            
            counter = 1
            original_sheet_name = sheet_name 
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_sheet_name[:28]}_{counter}" 
                counter += 1

            ws = wb.create_sheet(title=sheet_name) 
            headers = ["Full ID", "Beskrivelse", "Filer"]
            ws.append(headers)
            ws.auto_filter.ref = ws.dimensions 

            for cell in ws[1]:
                apply_cell_format(cell)

            for data_row in comp_data_list:
                excel_full_id = data_row["full_id"]
                
                if data_row["system"]: 
                    segment_to_replace = "=" + data_row["system"]
                    excel_full_id = excel_full_id.replace(segment_to_replace, data_row["system"])


                row_values = [excel_full_id, data_row["desc"], ", ".join(data_row["files"])]
                ws.append(row_values)
                for cell in ws[ws.max_row]:
                    apply_cell_format(cell)
            
            adjust_column_widths(ws) 

        if missing_system_rows:
            missing_sheet_name = "Systemnummer mangler"
            counter = 1
            original_missing_sheet_name = missing_sheet_name
            while missing_sheet_name in wb.sheetnames:
                missing_sheet_name = f"{original_missing_sheet_name[:28]}_{counter}"
                counter += 1

            ws_missing = wb.create_sheet(title=missing_sheet_name)
            headers = ["Full ID", "Beskrivelse", "Filer"]
            ws_missing.append(headers)
            ws_missing.auto_filter.ref = ws_missing.dimensions 

            for cell in ws_missing[1]:
                apply_cell_format(cell)

            for data_row in missing_system_rows:
                row_values = [data_row["full_id"], data_row["desc"], ", ".join(data_row["files"])]
                ws_missing.append(row_values)
                for cell in ws_missing[ws_missing.max_row]:
                    apply_cell_format(cell)
            
            adjust_column_widths(ws_missing) 

    else:
        ws = wb.active 
        if ws is None: 
             ws = wb.create_sheet(title="Komponentopptelling")
        else: 
            ws.title = "Komponentopptelling"
        
        headers = ["Med system?", "Komponent", "Beskrivelse", "Antall", "Filer"]
        ws.append(headers)
        ws.auto_filter.ref = ws.dimensions 

        for cell in ws[1]:
            apply_cell_format(cell)

        for r in rows:
            row_values = [r["has_system"], r["id"], r["desc"], r["count"], ", ".join(r["files"])]
            ws.append(row_values)
            for cell in ws[ws.max_row]:
                apply_cell_format(cell)
        
        adjust_column_widths(ws) 

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name=download_filename, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@systemkomponent_bp.route("/send/systembygging", methods=["POST"])
@login_required
def send_systembygging():
    print("✅ Route /send/systembygging ble kalt!")

    data = request.get_json()
    if not data:
        print("🚫 Ingen JSON-data mottatt")
        return "Ingen JSON-data", 400

    project_id = data.get("project_id")
    rows = data.get("rows", [])

    print("📦 project_id:", project_id)
    print("📦 rows size:", len(rows))

    if not project_id or not rows:
        return "Mangler prosjekt eller data", 400

    Systembygging.query.filter_by(project_id=project_id).delete()

    for r in rows:
        db.session.add(Systembygging(
            project_id=project_id,
            full_id=r["full_id"],
            desc=r["desc"],
            files=", ".join(r["files"])
        ))
    db.session.commit()
    return "OK", 200


@systemkomponent_bp.route("/send/komponentopptelling", methods=["POST"])
@login_required
def send_komponentopptelling():
    data = request.get_json()
    project_id = data.get("project_id")
    rows = data.get("rows", [])

    if not project_id or not rows:
        return "Mangler prosjekt eller data", 400

    Komponentopptelling.query.filter_by(project_id=project_id).delete()

    for r in rows:
        per_file = r.get("per_file", {})
        count = sum(per_file.values()) if isinstance(per_file, dict) else 0

        db.session.add(Komponentopptelling(
            project_id=project_id,
            komponent=r.get("id"),
            desc=r.get("desc"),
            count=count,
            has_system=r.get("has_system") == "Ja",
            files=", ".join(r.get("files", []))
        ))

    db.session.commit()
    return "OK", 200


    
@systemkomponent_bp.route("/tab/system/<int:project_id>")
@login_required
def vis_systembygging(project_id):
    prosjekt = Project.query.get_or_404(project_id)
    rader = Systembygging.query.filter_by(project_id=project_id).all()
    print(f"🔍 Henter systembygging for prosjekt {project_id}: {len(rader)} rader")
    for r in rader[:3]:  # Skriv ut de første 3
        print("➡️", r.full_id, "|", r.desc, "|", r.files)
    return render_template("partials/tab_system.html", prosjekt=prosjekt, rader=rader)



@systemkomponent_bp.route("/tab/komponent/<int:project_id>")
@login_required
def vis_komponentopptelling(project_id):
    prosjekt = Project.query.get_or_404(project_id)
    rader = Komponentopptelling.query.filter_by(project_id=project_id).all()
    return render_template("tab_komponent.html", prosjekt=prosjekt, rader=rader)
    
