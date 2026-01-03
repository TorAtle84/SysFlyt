import os
import tempfile
import traceback
import re
import json
from pathlib import Path
import logging

from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from app.models.db import db
from app.models.merkeskilt import Merkskilt
from app.models import User
from app.models.project import Project

merkeskilt_bp = Blueprint('merkeskilt', __name__)
logger = logging.getLogger(__name__)

# ────────────────────────────────────────────────────────────────────────────────
# Hent kode→beskrivelse‐mapping (tfm.xlsx)
# ────────────────────────────────────────────────────────────────────────────────
BASEDIR = Path(__file__).resolve().parents[1]            # …/app
DATA_DIR = BASEDIR / "data"
TFM_XLSX = DATA_DIR / "tfm.xlsx"
TFM_SETTINGS = DATA_DIR / "tfm-settings.json"

def _load_tfm_settings_dict() -> dict[str, bool]:
    """
    Leser tfm-settings.json og returnerer {kode: bool}. Støtter både
    flatt dict og {innstillinger:[{kode,aktiv},...]}.
    """
    try:
        if not TFM_SETTINGS.exists() or TFM_SETTINGS.stat().st_size == 0:
            return {}
        with TFM_SETTINGS.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and "innstillinger" in data:
            data = {i["kode"]: bool(i["aktiv"]) for i in data.get("innstillinger", []) if "kode" in i}
        return data if isinstance(data, dict) else {}
    except Exception as e:
        logger.warning("TFM settings kunne ikke lastes, bruker tomt oppsett: %s", e)
        return {}

def _save_tfm_settings(items: list[dict]) -> None:
    """Lagrer {kode: bool} for enkelhet og robusthet."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    flat = {i["kode"]: bool(i["aktiv"]) for i in items if "kode" in i}
    with TFM_SETTINGS.open("w", encoding="utf-8") as f:
        json.dump(flat, f, ensure_ascii=False, indent=2)

def _load_tfm_mapping() -> tuple[pd.DataFrame, dict[str, str]]:
    """Laster TFM-mapping fra Excel. Prøver ark 'TFM', faller tilbake til første ark."""
    try:
        if not TFM_XLSX.exists():
            raise FileNotFoundError(f"Mangler fil: {TFM_XLSX}")
        try:
            df = pd.read_excel(TFM_XLSX, sheet_name="TFM", header=None, dtype=str)
        except Exception:
            # Fallback: første ark
            with pd.ExcelFile(TFM_XLSX) as xls:
                first = xls.sheet_names[0]
            df = pd.read_excel(TFM_XLSX, sheet_name=first, header=None, dtype=str)
        df = df.fillna("Ikke i bruk")
        # Filtrer ut rader uten kode (kol A)
        df = df[df[0].notna() & (df[0].astype(str).str.strip() != "")]
        return df, dict(zip(df[0].astype(str).str.strip(), df[1].astype(str).str.strip()))
    except Exception as e:
        logger.error("Kunne ikke lese %s: %s", TFM_XLSX, e)
        return pd.DataFrame(columns=[0, 1]), {}

# Ikke last ved import – last ved behov i hver route (unngå tom cache/støtte hot-reload)
kode_mapping: dict[str, str] = {}

def hent_teknikere_og_admin():
    brukere = User.query.filter(User.role.in_(["tekniker", "admin"])).all()
    resultat_liste = []
    for b in brukere:
        navn = f"{getattr(b, 'first_name', '')} {getattr(b, 'last_name', '')}".strip() or getattr(b, "email", "")
        resultat_liste.append({"id": b.id, "navn": navn})
    return resultat_liste

@merkeskilt_bp.route("/merkeskilt")
def vis_merkeskilt():
    # 1) Teknikere/admin for dropdown (dersom du trenger det)
    teknikere = hent_teknikere_og_admin()

    # 2) TFM-innstillinger
    tfm_settings = _load_tfm_settings_dict()
    mapping_df, _lokal_map = _load_tfm_mapping()
    tfm_liste = []
    if mapping_df.empty:
        logger.warning("TFM-liste tom – sjekk %s og arknavn.", TFM_XLSX)
    for _, row in mapping_df.iterrows():
        kode         = str(row[0]).strip()
        beskrivelse  = str(row[1]).strip()
        aktiv        = tfm_settings.get(kode, True)
        tfm_liste.append((kode, beskrivelse, aktiv))

    # 3) Hent alle prosjekter for «Send til prosjekt»
    projects = Project.query.order_by(Project.project_name).all()

    # 4) Render malen med alle nødvendige parametre
    return render_template(
        "merkeskilt.html",
        teknikere=teknikere,
        tfm_liste=tfm_liste,
        projects=projects
    )

@merkeskilt_bp.route("/api/generer", methods=["POST"])
def generer_komponenter():
    data = request.json or {}
    taglinjer = data.get("tags", [])
    formatvalg = data.get("format", "")
    resultat, feilmeldinger = [], []

    tfm_settings = _load_tfm_settings_dict()
    mapping_df, lok_map = _load_tfm_mapping()
    global kode_mapping
    kode_mapping = lok_map or {}

    if not formatvalg:
        return jsonify({"feil": ["Ingen format-streng valgt."]}), 400

    for tag in taglinjer:
        tag = tag.strip()
        eq_pos = tag.find("=")
        dash_pos = tag.find("-", eq_pos + 1) if eq_pos != -1 else -1
        percent_pos = tag.find("%", dash_pos + 1) if dash_pos != -1 else -1

        if eq_pos == -1 or dash_pos == -1:
            feilmeldinger.append(f"Feil på rad '{tag}' – mangler '=' eller '-'")
            continue

        byggnr = tag[:eq_pos]
        system = tag[eq_pos:dash_pos]
        komponent = tag[dash_pos:percent_pos] if percent_pos != -1 else tag[dash_pos:]
        typekode = tag[percent_pos:] if percent_pos != -1 else ""

        kode2 = komponent[1:3] if len(komponent) >= 3 else ""
        aktiv = tfm_settings.get(kode2, True)
        if not aktiv and not komponent.startswith("-STB"):
            continue

        try:
            komp_str = (formatvalg
                .replace("{byggnr}", byggnr)
                .replace("{system}", system)
                .replace("{komponent}", komponent)
                .replace("{typekode}", typekode))
        except:
            feilmeldinger.append(f"Feil i format '{formatvalg}' for '{tag}'")
            continue

        beskrivelse = (kode_mapping or lok_map).get(kode2, "Ikke i bruk")
        resultat.append({
            "komponent": komp_str,
            "beskrivelse": beskrivelse,
            "himlingsskilt": "Nei",
            "stripsskilt": "Nei"
        })

    if feilmeldinger:
        return jsonify({"feil": feilmeldinger}), 400
    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/scan-og-generer", methods=["POST"])
def scan_og_generer():
    formatvalg = request.form.get("format", "").strip()
    if not formatvalg:
        return jsonify({"feil": ["Ingen format-streng valgt."]}), 400

    kriteriestr = request.form.get("system_kriterier", "").strip()
    system_kriterier = [k for k in kriteriestr.split(",") if k.strip().isdigit()] if kriteriestr else []

    filer = request.files.getlist("files")
    if not filer:
        return jsonify({"feil": ["Ingen filer funnet."]}), 400

    tfm_settings = _load_tfm_settings_dict()
    mapping_df, lok_map = _load_tfm_mapping()
    global kode_mapping
    kode_mapping = lok_map or {}

    placeholders = {
        "{byggnr}": r"(?P<byggnr>\+[A-Za-z0-9]+)",
        "{system}": r"(?P<system>=[^-]+)",
        "{komponent}": r"(?P<komponent>-[^%]+)",
        "{typekode}": r"(?P<typekode>[%/].+)"
    }
    regex_str = formatvalg
    for ph, subpat in placeholders.items():
        regex_str = regex_str.replace(ph, subpat)
    try:
        mønster = re.compile(regex_str)
    except re.error as e:
        return jsonify({"feil": [f"Regex-feil: {e}"]}), 400

    funn_set = set()
    for f in filer:
        try:
            data_ark = pd.read_excel(f, sheet_name=None, dtype=str)
        except Exception as e:
            return jsonify({"feil": [f"Feil i fil '{f.filename}': {e}"]}), 400
        for df in data_ark.values():
            for verdi in df.astype(str).values.flatten():
                if not isinstance(verdi, str):
                    continue
                for match in mønster.finditer(verdi):
                    system_val = match.groupdict().get("system", "")
                    if "{system}" in formatvalg and system_kriterier:
                        if system_val[1:3] not in system_kriterier:
                            continue
                    full_komponent = match.groupdict().get("komponent", "")
                    kode2 = full_komponent[1:3] if len(full_komponent) >= 3 else ""
                    aktiv = tfm_settings.get(kode2, True)

                    # QUICK FIX: Hopp over hvis inaktiv, MED UNNTAK for -STB
                    if not aktiv and not full_komponent.startswith("-STB"):
                        continue

                    funn_set.add(match.group(0))

    resultat = []
    for tag in funn_set:
        match = mønster.match(tag)
        if not match:
            continue
        gr = match.groupdict()
        byggnr = gr.get("byggnr", "")
        system = gr.get("system", "")
        komponent = gr.get("komponent", "")
        typekode = gr.get("typekode", "")
        komponent_val = match.groupdict().get("komponent", "")
        kode2 = komponent_val[1:3] if komponent_val else ""
        beskrivelse = (kode_mapping or lok_map).get(kode2, "Ikke i bruk")
        komp_str = (
            formatvalg.replace("{byggnr}", byggnr)
                      .replace("{system}", system)
                      .replace("{komponent}", komponent)
                      .replace("{typekode}", typekode)
        )
        resultat.append({
            "komponent": komp_str,
            "beskrivelse": beskrivelse,
            "himlingsskilt": "Nei",
            "stripsskilt": "Nei"
        })

    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/lastned", methods=["POST"])
def last_ned_excel():
    data = request.json or {}
    rader = data.get("rader", [])
    ordrenummer = data.get("ordrenummer")
    if not ordrenummer:
        return jsonify({"error": "Ordrenummer er påkrevd."}), 400

    wb = Workbook()
    ws = wb.active

    # 1) Skriv overskrifter + autofilter
    headers = ["Komponent-ID", "Beskrivelse", "Himlingsskilt", "Stripsskilt"]
    ws.append(headers)
    ws.auto_filter.ref = "A1:D1"  # Autofilter

    # 2) Legg inn radene med riktig datatype
    for rad in rader:
        komponent = rad.get("komponent", "")
        beskrivelse = rad.get("beskrivelse", "")
        himling = rad.get("himlingsskilt", "Nei")
        strips = rad.get("stripsskilt", "Nei")

        radnr = ws.max_row + 1
        c_komp = ws.cell(row=radnr, column=1, value=komponent)
        c_komp.number_format = "@"  # Eksplisitt tekstformat
        c_komp.data_type = "s"

        ws.cell(row=radnr, column=2, value=beskrivelse)
        ws.cell(row=radnr, column=3, value=himling)
        ws.cell(row=radnr, column=4, value=strips)

    # 3) Autojuster kolonnebredder
    for kol in ws.columns:
        maks = max(len(str(c.value)) if c.value else 0 for c in kol)
        ws.column_dimensions[kol[0].column_letter].width = maks + 2

    # 4) Lagre og send
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name=f"{ordrenummer}-Merkeskilt.xlsx")


@merkeskilt_bp.route("/api/tfm-liste", methods=["GET"])
def hent_tfm_liste():
    tfm_settings = _load_tfm_settings_dict()
    mapping_df, _lokal_map = _load_tfm_mapping()

    er_admin_flag = current_user.is_authenticated and getattr(current_user, "role", "") == "admin"

    resultat = []
    if mapping_df.empty:
        # Tom liste → returnér 200 med tom array; frontend viser tydelig beskjed
        return jsonify([]), 200
    for idx, row in mapping_df.iterrows():
        kode = str(row[0]).strip()
        beskrivelse = str(row[1]).strip()
        aktiv = tfm_settings.get(kode, True)
        resultat.append({
            "kode": kode,
            "beskrivelse": beskrivelse,
            "aktiv": aktiv,
            "er_admin": er_admin_flag
        })

    return jsonify(resultat), 200

@merkeskilt_bp.route("/api/tfm-liste/save", methods=["POST"])
def lagre_tfm_innstillinger():
    data = request.get_json(silent=True) or {}
    innstillinger = data.get("innstillinger", [])
    if not isinstance(innstillinger, list):
        return jsonify({"feil": ["Ugyldig format: forventer liste 'innstillinger'"]}), 400
    try:
        _save_tfm_settings(innstillinger)
        return jsonify({"success": True}), 200
    except Exception as e:
        logger.error("Kunne ikke lagre tfm-settings.json: %s", e)
        return jsonify({"feil": ["Klarte ikke lagre."]}), 500

@merkeskilt_bp.route("/send_merkeskilt", methods=["POST"])
@login_required
def send_merkeskilt():
    data = request.get_json()
    project_id = data.get("project_id")
    rows = data.get("rows")

    if not project_id or not rows:
        return jsonify({"error": "Mangler data"}), 400

    # 1. Slett eksisterende rader for dette prosjektet
    Merkskilt.query.filter_by(project_id=project_id).delete()

    # 2. Legg inn nye rader
    for row in rows:
        ny_rad = Merkskilt(
            project_id=project_id,
            komponent_id=row.get("komponent_id"),
            beskrivelse=row.get("beskrivelse"),
            aktiv=row.get("aktiv") == "Ja",
            inaktiv=row.get("inaktiv") == "Ja"
        )
        db.session.add(ny_rad)

    db.session.commit()
    return jsonify({"success": True})
