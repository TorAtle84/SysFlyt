import os, re, json, tempfile, subprocess, traceback
from io import BytesIO
from collections import OrderedDict

import pandas as pd
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

masseliste_bp = Blueprint("masseliste", __name__)

# ────────────────────────────────────────────────────────────────────────────────
# UI
# ────────────────────────────────────────────────────────────────────────────────
@masseliste_bp.route("/masseliste", methods=["GET"])
def vis_masseliste():
    return render_template("masseliste.html")

# ────────────────────────────────────────────────────────────────────────────────
# EKSISTERENDE: Søk i masseliste (Excel → JSON)
# ────────────────────────────────────────────────────────────────────────────────
@masseliste_bp.route("/api/parse-masseliste", methods=["POST"])
def parse_masseliste_api():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Ingen fil lastet opp"}), 400

    try:
        df = pd.read_excel(file, dtype=str).fillna("")
        data = df.to_dict(orient="records")
        current_app.logger.info(f"[INFO] Lest inn {len(data)} rader fra fil {file.filename}")
        return jsonify({"rows": data, "columns": list(df.columns)})
    except Exception as e:
        current_app.logger.exception("Under parsing")
        return jsonify({"error": f"Feil under innlasting: {str(e)}"}), 500

# ────────────────────────────────────────────────────────────────────────────────
# NY: Eksport av “Søk i masseliste” → Excel
# ────────────────────────────────────────────────────────────────────────────────
@masseliste_bp.route("/api/masseliste/export", methods=["POST"])
def export_parsed_masseliste():
    payload = request.get_json() or {}
    rows = payload.get("rows") or []
    columns = payload.get("columns")  # valgfri, hvis du vil styre kolonnerekkefølge
    if not rows:
        return jsonify({"error": "Ingen rader å eksportere."}), 400

    # Bestem kolonner
    if not columns:
        keys = set()
        for r in rows:
            keys.update(r.keys())
        columns = sorted(keys)

    wb = Workbook()
    ws = wb.active
    ws.title = "Masseliste"
    ws.append(columns)

    # Skriv rader
    for r in rows:
        ws.append([r.get(c, "") for c in columns])

    # Autofilter
    try:
        last_col_letter = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"
    except Exception:
        pass

    # Tekstformat for potensielle ID-kolonner (prøv å finne)
    id_candidates = [c for c in columns if c.lower() in ("komponent-id", "komponent", "id")]
    for id_col in id_candidates:
        try:
            idx = columns.index(id_col) + 1
            for cell in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell[0].number_format = "@"
                cell[0].data_type = "s"
        except Exception:
            pass

    # Autojuster bredder
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    navn = f"{getattr(current_user, 'first_name', 'bruker')}_{getattr(current_user, 'last_name', '')}".strip() or "bruker"
    return send_file(tmp.name, as_attachment=True, download_name=f"{navn}-Masseliste.xlsx")

# ────────────────────────────────────────────────────────────────────────────────
# EKSISTERENDE: Versjonssammenligning (unik diff)
# ────────────────────────────────────────────────────────────────────────────────
@masseliste_bp.route("/versjonssammenligning", methods=["POST"])
def versjonssammenligning_api():
    current_app.logger.info("Starter generering av unike endringer-rapport...")
    ref_file = request.files.get("ref_file")
    orig_file = request.files.get("orig_file")

    if not ref_file or not orig_file:
        return "Begge filer må lastes opp", 400

    df_new = pd.read_excel(ref_file, dtype=str).fillna("")
    df_old = pd.read_excel(orig_file, dtype=str).fillna("")

    # Finn komponentkolonne
    prefererte_feltnavn = ["komponent", "type", "utstyr"]
    komponentfelt = None
    for navn in prefererte_feltnavn:
        treff = [col for col in df_new.columns if navn in col.lower()]
        if treff:
            komponentfelt = treff[0]
            break
    if not komponentfelt:
        return "Fant ingen komponent-, type- eller utstyr-kolonne", 400

    current_app.logger.info(f"Bruker komponentfelt: {komponentfelt}")

    df_new['__komponent__'] = df_new[komponentfelt].astype(str)
    df_old['__komponent__'] = df_old[komponentfelt].astype(str)

    new_counts = df_new['__komponent__'].value_counts().to_dict()
    old_counts = df_old['__komponent__'].value_counts().to_dict()

    telle_kandidater = ['antall', 'kvantitet', 'mengde', 'qt', 'stk', 'stykk']
    tellefelt = None
    for kol in df_new.columns:
        if any(k in kol.lower() for k in telle_kandidater):
            if pd.to_numeric(df_new[kol], errors='coerce').notna().any():
                tellefelt = kol
                break

    if tellefelt:
        current_app.logger.info(f"Fant tellefelt: {tellefelt}, vurderer opptelling...")
        df_new[tellefelt] = pd.to_numeric(df_new[tellefelt], errors='coerce').fillna(0)
        df_old[tellefelt] = pd.to_numeric(df_old[tellefelt], errors='coerce').fillna(0)
        if df_new[tellefelt].max() > 1:
            new_counts = df_new.groupby('__komponent__')[tellefelt].sum().to_dict()
            old_counts = df_old.groupby('__komponent__')[tellefelt].sum().to_dict()

    recs = []
    for k in sorted(set(new_counts.keys()) | set(old_counts.keys())):
        ny = new_counts.get(k, 0)
        gammel = old_counts.get(k, 0)
        diff = ny - gammel
        recs.append({'Komponent': k, 'Ny': ny, 'Gammel': gammel, 'Differanse': diff})

    df_diff = pd.DataFrame(recs)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_diff.to_excel(writer, index=False, sheet_name='Unike')
        ws = writer.sheets['Unike']
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 80)

        header = [c.value for c in ws[1]]
        if 'Differanse' in header:
            diff_col_idx = header.index('Differanse') + 1
            for row in ws.iter_rows(min_row=2, min_col=diff_col_idx, max_col=diff_col_idx):
                cell = row[0]
                try:
                    if float(cell.value) > 0:
                        cell.fill = PatternFill('solid', fgColor='BDD7EE')  # blå
                    elif float(cell.value) < 0:
                        cell.fill = PatternFill('solid', fgColor='F8CECC')  # rød
                except Exception:
                    pass

    buf.seek(0)
    brukernavn = f"{getattr(current_user, 'first_name', 'bruker')}_{getattr(current_user, 'last_name', '')}".strip() or "bruker"
    current_app.logger.info("Excel med unike endringer generert")

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{brukernavn}-Unikeendringer.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ────────────────────────────────────────────────────────────────────────────────
# IFC-masseliste (skann + eksport med Pset-valg)
# ────────────────────────────────────────────────────────────────────────────────
BASEDIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.abspath(os.path.join(BASEDIR, "..", "data"))
TFM_SETTINGS_PATH = os.path.join(DATA_DIR, "tfm-settings.json")

def load_tfm_settings():
    try:
        with open(TFM_SETTINGS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

PLACEHOLDERS = {
    "{byggnr}":   r"(?P<byggnr>\+[A-Za-z0-9]+)",
    "{system}":   r"(?P<system>=[^-]+)",
    "{komponent}":r"(?P<komponent>-[^%]+)",
    "{typekode}": r"(?P<typekode>[%/].+)"
}

def format_to_regex(format_str: str) -> re.Pattern:
    rx = format_str
    for ph, sub in PLACEHOLDERS.items():
        rx = rx.replace(ph, sub)
    return re.compile(rx)

def try_import_ifcopenshell():
    try:
        import ifcopenshell
        from ifcopenshell.util.element import get_psets
        return ifcopenshell, get_psets
    except Exception as e:
        current_app.logger.warning(f"ifcopenshell ikke tilgjengelig: {e}")
        return None, None

def ifc_to_json_via_cli(ifc_path: str) -> dict:
    out_json = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
    out_json.close()
    cmd = ["ifcconvert", ifc_path, out_json.name, "--json"]
    try:
        subprocess.check_call(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        with open(out_json.name, "r", encoding="utf-8") as f:
            data = json.load(f)
        os.unlink(out_json.name)
        return data
    except Exception as e:
        try:
            os.unlink(out_json.name)
        except Exception:
            pass
        raise RuntimeError(f"Kunne ikke kjøre ifcconvert: {e}")

def build_candidate_strings(elem, psets_dict: dict) -> list:
    cands = []
    for attr in ("Name", "Tag", "ObjectType", "GlobalId", "PredefinedType"):
        val = getattr(elem, attr, None)
        if val:
            cands.append(str(val))
    try:
        type_obj = getattr(elem, "IsTypedBy", None)
        if type_obj:
            for rel in type_obj:
                t = getattr(rel, "RelatingType", None)
                if t and getattr(t, "Name", None):
                    cands.append(str(t.Name))
                if t and getattr(t, "PredefinedType", None):
                    cands.append(str(t.PredefinedType))
    except Exception:
        pass
    for pset_name, props in (psets_dict or {}).items():
        cands.append(pset_name)
        if isinstance(props, dict):
            for k, v in props.items():
                if isinstance(v, dict) and "value" in v:
                    val = v["value"]
                else:
                    val = v
                if val is None:
                    continue
                cands.append(f"{k}={val}")
    return [s for s in cands if isinstance(s, str) and s]

def tfm_active(komponent_str: str, tfm_settings: dict) -> bool:
    m = re.search(r"-([A-Za-z]{2})", komponent_str or "")
    if not m:
        return True
    kode2 = m.group(1).upper()
    return bool(tfm_settings.get(kode2, True))

@masseliste_bp.route("/api/ifc/scan", methods=["POST"])
def api_ifc_scan():
    fmt = (request.form.get("format") or "").strip()
    kriterier_str = (request.form.get("system_kriterier") or "").strip()
    system_kriterier = [k for k in kriterier_str.split(",") if k.strip().isdigit()]
    f = request.files.get("ifc_file")

    if not fmt:
        return jsonify({"feil": ["Ingen format-streng valgt."]}), 400
    if not f:
        return jsonify({"feil": ["Ingen IFC-fil lastet opp."]}), 400

    try:
        pattern = format_to_regex(fmt)
    except re.error as e:
        return jsonify({"feil": [f"Regex-feil i format: {e}"]}), 400

    tmp_ifc = tempfile.NamedTemporaryFile(delete=False, suffix=".ifc")
    f.stream.seek(0)
    tmp_ifc.write(f.read())
    tmp_ifc.close()

    tfm_settings = load_tfm_settings()
    treff = []

    try:
        ifcopenshell, get_psets = try_import_ifcopenshell()
        if ifcopenshell:
            model = ifcopenshell.open(tmp_ifc.name)
            for elem in model.by_type("IfcElement"):
                try:
                    psets = (get_psets(elem) or {})
                except Exception:
                    psets = {}
                candidates = build_candidate_strings(elem, psets)

                if "{system}" in fmt and not any("=" in c for c in candidates):
                    continue

                matched_value, groups = None, None
                for c in candidates:
                    m = pattern.search(str(c))
                    if m:
                        matched_value = m.group(0)
                        groups = m.groupdict()
                        break
                if not matched_value:
                    continue

                if "{system}" in fmt and system_kriterier and groups.get("system"):
                    if groups["system"][1:3] not in system_kriterier:
                        continue
                if not tfm_active(groups.get("komponent", ""), tfm_settings):
                    continue

                row = OrderedDict()
                row["Komponent-ID"] = matched_value
                row["IFC Class"] = getattr(elem, "is_a", lambda: "IfcElement")()
                row["Name"] = getattr(elem, "Name", "")
                row["Tag"] = getattr(elem, "Tag", "")
                row["ObjectType"] = getattr(elem, "ObjectType", "")
                row["GlobalId"] = getattr(elem, "GlobalId", "")
                try:
                    if hasattr(elem, "IsTypedBy") and elem.IsTypedBy:
                        t = elem.IsTypedBy[0].RelatingType
                        row["Type.Name"] = getattr(t, "Name", "")
                        row["Type.PredefinedType"] = getattr(t, "PredefinedType", "")
                except Exception:
                    pass
                for pset_name, props in (psets or {}).items():
                    if isinstance(props, dict):
                        for k, v in props.items():
                            val = v.get("value") if isinstance(v, dict) else v
                            if val is not None:
                                row[f"Pset:{pset_name}.{k}"] = str(val)
                treff.append(row)
        else:
            data = ifc_to_json_via_cli(tmp_ifc.name)
            objects = data.get("objects") or []
            for obj in objects:
                candidates = []
                for key in ("Name", "Tag", "ObjectType", "GlobalId", "PredefinedType"):
                    v = obj.get(key)
                    if v:
                        candidates.append(str(v))
                t = obj.get("type") or {}
                if isinstance(t, dict):
                    if t.get("Name"): candidates.append(str(t["Name"]))
                    if t.get("PredefinedType"): candidates.append(str(t["PredefinedType"]))
                psets = obj.get("psets") or {}
                for pset, props in psets.items():
                    candidates.append(pset)
                    if isinstance(props, dict):
                        for k, v in props.items():
                            candidates.append(f"{k}={v}")

                if "{system}" in fmt and not any("=" in c for c in candidates):
                    continue

                matched_value, groups = None, None
                for c in candidates:
                    m = pattern.search(str(c))
                    if m:
                        matched_value = m.group(0)
                        groups = m.groupdict()
                        break
                if not matched_value:
                    continue

                if "{system}" in fmt and system_kriterier and groups.get("system"):
                    if groups["system"][1:3] not in system_kriterier:
                        continue
                if not tfm_active(groups.get("komponent", ""), tfm_settings):
                    continue

                row = OrderedDict()
                row["Komponent-ID"] = matched_value
                row["IFC Class"] = obj.get("ifc_class") or obj.get("typeName") or ""
                row["Name"] = obj.get("Name", "")
                row["Tag"] = obj.get("Tag", "")
                row["ObjectType"] = obj.get("ObjectType", "")
                row["GlobalId"] = obj.get("GlobalId", "")
                if isinstance(t, dict):
                    row["Type.Name"] = t.get("Name", "")
                    row["Type.PredefinedType"] = t.get("PredefinedType", "")
                for pset, props in (psets or {}).items():
                    if isinstance(props, dict):
                        for k, v in props.items():
                            row[f"Pset:{pset}.{k}"] = str(v)
                treff.append(row)

    except Exception as e:
        current_app.logger.exception("Feil under IFC-scan")
        return jsonify({"feil": [f"Feil under IFC-skanning: {e}"]}), 500
    finally:
        try:
            os.unlink(tmp_ifc.name)
        except Exception:
            pass

    # Returner også en liste over alle Pset-kolonner (for checkbox-UI)
    pset_cols = sorted({k for r in treff for k in r.keys() if str(k).startswith("Pset:")})
    return jsonify({"rows": treff, "pset_columns": pset_cols}), 200

@masseliste_bp.route("/api/ifc/export", methods=["POST"])
def api_ifc_export():
    data = request.get_json() or {}
    rows = data.get("rows") or []
    pset_selected = data.get("pset_columns")  # liste eller None
    if not rows:
        return jsonify({"error": "Ingen rader å eksportere."}), 400

    core_first = ["Komponent-ID", "IFC Class", "Name", "Tag", "ObjectType", "GlobalId", "Type.Name", "Type.PredefinedType"]

    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())

    if pset_selected is not None:
        tail = sorted([
            k for k in all_keys
            if k not in core_first and (not str(k).startswith("Pset:") or k in set(pset_selected))
        ])
    else:
        tail = sorted([k for k in all_keys if k not in core_first])

    cols = [k for k in core_first if k in all_keys] + tail

    wb = Workbook()
    ws = wb.active
    ws.title = "IFC-masseliste"
    ws.append(cols)

    try:
        last_col_letter = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"
    except Exception:
        pass

    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    # Tekstformat på ID
    try:
        id_col_idx = cols.index("Komponent-ID") + 1
        for cell in ws.iter_rows(min_row=2, min_col=id_col_idx, max_col=id_col_idx):
            cell[0].number_format = "@"
            cell[0].data_type = "s"
    except ValueError:
        pass

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    navn = f"{getattr(current_user, 'first_name', 'bruker')}_{getattr(current_user, 'last_name', '')}".strip() or "bruker"
    return send_file(tmp.name, as_attachment=True, download_name=f"{navn}-IFC-masseliste.xlsx")
