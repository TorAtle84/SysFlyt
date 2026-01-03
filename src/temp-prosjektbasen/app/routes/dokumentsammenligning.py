# app/routes/dokumentsammenligning.py
import os
import re
import json
import colorsys
import logging
import difflib
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Set, Tuple, Optional

from flask import Blueprint, render_template, request, send_file, abort
from flask_login import current_user
from werkzeug.utils import secure_filename

# Dokument/tekst
from fitz import open as fitz_open  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Diff-motor (for "Syntaktisk → diff" Excel-rapport)
from app.utils import generate_report
from rapidfuzz import fuzz

# OCR (valgfritt – vi prøver å importere, men appen fungerer uten)
try:
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False


# ==============================
# Blueprint (matcher base.html)
# ==============================
dokumentsammenligning_bp = Blueprint("dokumentsammenligning", __name__)

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_ROOT = BASE_DIR / "temp"
os.makedirs(TEMP_ROOT, exist_ok=True)

# ==============================
# Hjelpere
# ==============================
META_CHARS = set(r".^$*+?{}[]\\|()")

TFM_MAP = {"AB": "Bjelke", "AD": "Dekker", "AE": "Elementer", "AF": "Fagverk", "AG": "Glassfelt", "AH": "Fundamenter", "AK": "Komplette konstruksjoner", "AL": "List - Beslistning", "AO": "Oppbygende - Utforende", "AP": "Plate", "AR": "Ramme - Oppheng", "AS": "Søyle", "AU": "Åpninger", "AV": "Vegg", "BA": "Armering Forsterkning", "BB": "Beskyttende - Stoppende", "BC": "Begrensende", "BF": "Fuging", "BG": "Pakning Tetning", "BI": "Isolasjon", "BP": "Avretningsmasse", "BS": "Spikerslag", "CB": "Balkong", "CC": "Baldakin", "CD": "Karnapp", "CG": "Rampe - Repos", "CK": "Komplett konstruksjon", "CM": "Kjemiske stoffer", "CO": "Kobling Overgang", "CP": "Pipe - Skostein", "CQ": "Festemateriell", "CR": "Rammeverk - Oppheng", "CT": "Trapp - Leider", "CX": "Tunnel - Bru (inne-ute)", "DB": "Dør med brannklasse", "DF": "Foldedør - Foldevegg", "DI": "Dør - innvendig", "DK": "Kjøretøyregulering", "DL": "Luke", "DP": "Port", "DT": "Dør - tilfluktsrom", "DU": "Dør - utvendig", "DV": "Vindu", "EB": "Overflatebekledning", "EC": "Overflatebehandling", "EH": "Himling", "FA": "Ventilert arbeidsplass", "FB": "Benk - Bord - Plate - Tavle", "FC": "Beslag", "FD": "Disk - Skranke", "FF": "Fryserom", "FH": "Hylle - Reol", "FI": "Kabinett", "FK": "Kjølerom - Svalrom", "FO": "Sittebenk - Sofa - Stol", "FR": "Rom", "FS": "Skap - Skuff", "FT": "Speil", "FV": "Vaskemaskin", "FX": "Krok - Knagg - Håndtak", "GA": "Automat - Maskin", "GB": "Benk - Bord - Plate - Tavle", "GD": "Dekontamator", "GE": "Autoklaver", "GF": "Fryseskap", "GG": "Gardiner - Forheng", "GH": "Hylle - Reol", "GK": "Kjøleskap - Kjøledisk", "GL": "Lås - Beslag", "GM": "Mattilbereding", "GN": "Nøkler", "GO": "Sofa - Sittebenk", "GP": "Stol", "GQ": "Seng - Liggebenk", "GS": "Skap - Skuffer", "GT": "Tørkeskap - Varmeskap", "GV": "Vaskemaskin", "GW": "Vekt", "GX": "Holder", "GY": "Avfallsbeholder", "HA": "Automobil - Bil", "HB": "Rullebord - Tralle", "HC": "Container - Vogn", "HM": "Maskin", "HS": "Rullestol", "HT": "Truck - kran", "HV": "Verktøy", "IB": "Brenner", "IC": "Solceller-solfangere", "ID": "Kjel for destruksjon", "IE": "Elektrokjel", "IF": "Kjel for fast-bio brensel", "IG": "Generator", "IK": "Kuldeaggregat", "IL": "Energibrønn", "IM": "Motor", "IO": "Oljekjel", "IP": "Gasskjel", "IT": "Trykkluftaggregat (enhet)", "IU": "Turbin", "IV": "Aggregatenhet", "JF": "Forsterker", "JK": "Kompressor", "JP": "Pumpe", "JQ": "Pumpe i VA-installasjoner", "JV": "Vifte", "JW": "Spesialvifte", "KA": "Aktuator", "KD": "Drivhjul - Drev", "KE": "Induktiv energioverføring", "KG": "Gjennomføring", "KH": "Transportenhet (hevende forflyttende)", "KJ": "Jordingskomponenter", "KK": "Kanal", "KM": "Mast - Antenne", "KN": "Nedløp", "KO": "Kraftoverføring", "KQ": "Rør - spesielt", "KR": "Rør - generelt", "KS": "Skinne - Bane - Spor", "KU": "Kombinert kabel", "KV": "Høyspenningskabel > 1000V", "KW": "Lavspenningskabel 50 til 1000V", "KX": "Lavspenningskabel < 50V", "KY": "Optisk kabel", "KZ": "Slange", "LB": "Varmeomformende med vifte", "LC": "Kjøleomformende med vifte", "LD": "Kjøleflater", "LE": "Kondensator", "LF": "Fordamper", "LG": "Gear clutch", "LH": "Varmeflate", "LI": "Varmeelement", "LK": "Kjøleomformende", "LL": "Lyskilde", "LN": "Likeretter", "LO": "Omformer", "LP": "Pens - Veksel - Sjalter", "LQ": "Vekselretter", "LR": "Frekvensomformer", "LS": "Strålevarme", "LU": "Luftfukter", "LV": "Varmeomformende", "LX": "Varmegjenvinner", "LZ": "Varmerkabel - Varmerør", "MA": "Absoluttfilter", "MB": "ABC-filter", "MC": "UV-filter", "ME": "Elektrostatiske filter", "MF": "Luftfilter", "MG": "Fettfilter", "MK": "Kondenspotte", "ML": "Luftutskiller", "MM": "Membran", "MO": "Utskiller", "MR": "Rist - Sil", "MS": "Syklon", "MT": "Tørke", "MU": "Filter for Lyd - Bilde - Frekvensutjevner", "MV": "Vannfilter", "MX": "Støyfilter", "NB": "Batteri - UPS", "NC": "Kondensator", "NI": "Informasjonslagring", "NK": "Kum", "NM": "Badekar - Basseng", "NO": "Åpen tank", "NT": "Tank med trykk", "NU": "Tank uten trykk", "NV": "Vekt Lodd", "NW": "Varmtvannsbereder", "NX": "Toalett", "NY": "Servant", "NZ": "Brannslukkingsapparat", "OA": "AV-maskiner", "OB": "Shuntgruppe", "OD": "Datamaskin", "OE": "Energimåler", "OF": "Systembærer", "OM": "Mottaker - Sender", "OP": "PBX", "OQ": "Dataprogramprogramvare", "OR": "Router Fordeler", "OS": "Sentralenhet Mikser i Lydsystem", "OT": "Telefonapparat", "OU": "Undersentral", "QB": "Belastningsvakt", "QD": "Differansetrykkvakt", "QE": "Elektrisk vern", "QF": "Strømningsvakt", "QG": "", "QH": "Fuktvakt", "QI": "", "QJ": "", "QK": "", "QL": "Lyddemper", "QM": "Mekanisk beskyttelse", "QN": "Nivåvakt", "QO": "Overtrykksventil", "QP": "Trykkvakt", "QQ": "Vibrasjonsvakt", "QR": "Rotasjonsvakt", "QS": "Strømvakt", "QT": "Temperaturvakt", "QU": "", "QV": "Sikkerhetsventil", "QW": "", "QX": "Solavskjerming", "QY": "Lynavleder", "QZ": "Brannvern", "RA": "AV-opptaker", "RB": "Bevegelse", "RC": "Seismometer", "RD": "Differansetrykkgiver", "RE": "Elektriske variabler", "RF": "Strømningsmåler", "RG": "Posisjon - Lengde", "RH": "Fuktighetsgiver", "RI": "Termometer", "RJ": "Fotocelle", "RK": "Kortleser", "RL": "", "RM": "Multifunksjonell Kombinert føler", "RN": "Nivågiver", "RO": "", "RP": "Trykkgiver", "RQ": "Manometer Trykkmåler", "RR": "Giver generelt", "RS": "Hastighetsmåler", "RT": "Temperaturgiver", "RU": "Ur", "RV": "Veieceller", "RW": "Virkningsgradsmåler", "RX": "Målepunkt", "RY": "Gassdetektor-Røykdetektor", "RZ": "Branndeteksjon", "SA": "Reguleringsventil manuell", "SB": "Reguleringsventil motorstyrt", "SC": "Stengeventil motorstyrt", "SD": "Alarmventil sprinkler", "SE": "Ekspansjonsventil", "SF": "Fraluftsventil", "SG": "Tilbakeslagsventil - Overtrykkspjeld", "SH": "Hurtigkobling", "SI": "Effektregulator", "SJ": "Jevntrykksventil", "SK": "Strømningsregulator - CAV", "SL": "", "SM": "Stengeventil manuell", "SN": "", "SO": "", "SP": "Trykkutjevningsventil", "SQ": "Strømningsregulator - VAV", "SR": "Reguleringsspjeld", "SS": "Stengespjeld", "ST": "Tilluftsventil", "SU": "Sugetrykksventil", "SV": "Strupeventil", "SW": "Plenumskammer", "SX": "Regulator", "SY": "", "SZ": "Brannspjeld - Røykspjeld", "UA": "Uttak alarm", "UB": "Blandebatteri", "UC": "", "UD": "Uttak data", "UE": "Uttak el", "UF": "Fellesuttak", "UG": "Uttak gass", "UH": "Høyttaler", "UI": "", "UJ": "Skriver", "UK": "Kontrollpanel - Tablå", "UL": "Uttak trykkluft", "UM": "Monitor - Display", "UN": "Nødbelysning", "UO": "Trommel", "UP": "Belysningsarmatur", "UQ": "", "UR": "Uttak radio", "US": "Stasjon", "UT": "Uttak telefon", "UU": "", "UV": "Uttak vann", "UW": "", "UX": "Koblingsboks", "UY": "Uttak antenne", "UZ": "Dyse - Spreder", "VA": "", "VB": "Bærelag", "VC": "", "VD": "Dekke", "VE": "", "VF": "", "VG": "Gress", "VH": "", "VI": "", "VJ": "", "VK": "Kantstein heller", "VL": "Masse", "VM": "Mekanisk beskyttelse", "VN": "", "VO": "", "VP": "Planterbuskertrær", "VQ": "", "VR": "", "VS": "Skilt", "VT": "", "VU": "", "VV": "", "VW": "", "VX": "", "VY": "", "VZ": "", "XA": "", "XB": "", "XC": "Kondensator", "XD": "Komp. for binærlogikk", "XE": "", "XF": "Komponenter for vern", "XG": "Komponenter for krafttilførsel", "XH": "Komponenter for signalering", "XI": "Potensiometer", "XJ": "", "XK": "Releer - Kontaktorer", "XL": "Induktiv komponenter", "XM": "Motor", "XN": "Integrerte kretser", "XO": "Urbryter - Timer", "XP": "Komponenter for måling og prøving", "XQ": "Effektbryter", "XR": "Motstand", "XS": "Bryter / Vender", "XT": "Transformator", "XU": "", "XV": "Halvlederkomponenter og elektronrør", "XW": "", "XX": "Rekkeklemmer - Samlesignal", "XY": "", "XZ": "Terminering og tilpasning", "YA": "", "YB": "", "YC": "", "YD": "", "YE": "", "YF": "", "YG": "", "YH": "", "YI": "", "YJ": "", "YK": "", "YL": "", "YM": "", "YN": "", "YO": "", "YP": "", "YQ": "", "YR": "", "YS": "", "YT": "", "YU": "", "YV": "", "YW": "", "YX": "", "YY": "", "YZ": "", "ZA": "", "ZB": "", "ZC": "", "ZD": "", "ZE": "", "ZF": "", "ZG": "", "ZH": "", "ZI": "", "ZJ": "", "ZK": "", "ZL": "", "ZM": "", "ZN": "", "ZO": "", "ZP": "", "ZQ": "", "ZR": "", "ZS": "", "ZT": "", "ZU": "", "ZV": "", "ZW": "", "ZX": "", "ZY": "", "ZZ": ""}

def is_regex(pat: str) -> bool:
    return any(ch in META_CHARS for ch in (pat or ""))

def compile_regex(pat: str, flags=re.IGNORECASE | re.MULTILINE | re.DOTALL):
    try:
        return re.compile(pat, flags)
    except re.error:
        return re.compile(re.escape(pat or ""), flags)

def tfm_desc(match_text: str) -> str:
    m = re.match(r"([A-Za-z]+)", match_text or "")
    if not m:
        return "Uspesifisert"
    key = (m.group(1) or "")[:2].upper()
    return TFM_MAP.get(key, f"Uspesifisert ({key})")

def compact_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s or "")

def join_word_whitespace(s: str) -> str:
    """Fjern whitespace KUN mellom word-tegn: 'K A 4 0 1' → 'KA401'."""
    return re.sub(r"(?<=\w)\s+(?=\w)", "", s or "")

def strip_word_boundaries(pattern: str) -> str:
    return re.sub(r"\\b", "", pattern or "")

def canonical(val: str) -> str:
    """Kanoniser for dedup: fjern whitespace og -_/."""
    return re.sub(r"[\s\-\_/]+", "", val or "")

def make_flex_from_sample(sample: str) -> str:
    """
    Gjør et format-eksempel (f.eks. 'JPA5001') om til elastisk regex som tåler
    mellomrom mellom hvert tegn og valgfri ledende '-'/'='.
    Håndterer også blandede tegn (f.eks. 'JP-5001').
    """
    s = re.sub(r"\s+", "", sample or "")
    if not s:
        return re.escape(sample or "")

    chunks = []
    for ch in s:
        if ch.isalpha():
            chunks.append(r"[A-Za-z]\s*")
        elif ch.isdigit():
            chunks.append(r"\d\s*")
        else:
            chunks.append(re.escape(ch) + r"\s*")

    core = "".join(chunks)
    return rf"(?<!\w)[\-=]?\s*{core}(?!\w)"

# ==============================
# Tekstekstraksjon – backup + robust ord-variant for PDF
# ==============================
def extract_lines_backup(path: str) -> list[str]:
    """
    Samme strategi som i backup/diff:
      - PDF: page.get_text() → splitlines
      - DOCX: avsnitt
      - XLSX/XLS: celler pr. rad
      - TXT: linjer
    + punktum/komma-varianter pr linje
    """
    ext = Path(path).suffix.lower()
    raw_lines: list[str] = []
    try:
        if ext == '.pdf':
            doc = fitz_open(path)
            raw_lines = sum([page.get_text().splitlines() for page in doc], [])
        elif ext == '.docx':
            doc = DocxDocument(path)
            raw_lines = [p.text for p in doc.paragraphs]
        elif ext in ('.xlsx', '.xls'):
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    raw_lines.append(' '.join(str(c) for c in row if c is not None))
        else:
            with open(path, 'r', errors='ignore') as f:
                raw_lines = f.read().splitlines()
    except Exception:
        logging.warning(f"Kunne ikke lese fil: {path}")

    lines: list[str] = []
    for line in raw_lines:
        s = (line or "").strip()
        if not s:
            continue
        lines.append(s)                 # original
        alt = s.replace('.', ',')       # punktum → komma
        if alt != s:
            lines.append(alt)
        alt2 = s.replace(',', '.')      # komma → punktum
        if alt2 != s:
            lines.append(alt2)
    return lines

def extract_lines_pdf_words(path: str) -> list[str]:
    """
    Robust PDF-variant: bruk page.get_text('words') og grupper (block,line).
    Gir ofte 'KA401' som ett ord selv når PDF lagrer tegn separat.
    """
    if Path(path).suffix.lower() != '.pdf':
        return []

    lines: list[str] = []
    try:
        doc = fitz_open(path)
        for page in doc:
            words = page.get_text("words")  # x0,y0,x1,y1,text,block,line,word
            if not words:
                txt = page.get_text() or ""
                if txt:
                    lines.extend([ln for ln in txt.splitlines() if ln.strip()])
                continue
            words.sort(key=lambda w: (w[5], w[6], w[7]))
            curr = None
            buff = []
            for w in words:
                key = (w[5], w[6])
                tok = (w[4] or "").strip()
                if not tok:
                    continue
                if key != curr and buff:
                    lines.append(" ".join(buff))
                    buff = []
                curr = key
                buff.append(tok)
            if buff:
                lines.append(" ".join(buff))
    except Exception:
        logging.exception("PDF word-ekstraksjon feilet (fall-back til backup)")

    # samme punktum/komma-varianter
    out: list[str] = []
    for s in lines:
        s = s.strip()
        if not s:
            continue
        out.append(s)
        alt = s.replace('.', ',')
        if alt != s:
            out.append(alt)
        alt2 = s.replace(',', '.')
        if alt2 != s:
            out.append(alt2)
    return out

def merged_lines_for_search(path: str) -> list[str]:
    """For SØK: slå sammen backup-linjer og ordbaserte PDF-linjer (hvis PDF)."""
    base = extract_lines_backup(path)
    ordlinjer = extract_lines_pdf_words(path)
    if not ordlinjer:
        return base
    seen = set()
    out = []
    for ln in base + ordlinjer:
        if ln not in seen:
            seen.add(ln)
            out.append(ln)
    return out

# ==============================
# Mønster-normalisering (diff-lignende elastisitet)
# ==============================
def make_search_patterns(user_pat: str) -> List[re.Pattern]:
    """
    - Hvis brukeren skrev et *format-eksempel* (ingen regex-metategn) → elastisk format-regex.
    - Hvis brukeren skrev regex → bruk som det er.
    I begge tilfeller lager vi også en 'joined' variant (uten \b).
    """
    patterns: List[str] = []
    if not is_regex(user_pat):
        flex = make_flex_from_sample(user_pat)
        patterns.extend([flex, re.escape(user_pat or "")])
    else:
        patterns.append(user_pat or "")

    joined = [strip_word_boundaries(p) for p in patterns]
    out: List[re.Pattern] = []
    seen = set()
    for p in patterns + joined:
        if p and p not in seen:
            seen.add(p)
            out.append(compile_regex(p))
    return out

# ==============================
# Søk – bygg rader
# ==============================
def find_matches_build_rows(filename: str, lines: List[str], user_pat: str) -> List[Dict]:
    """
    Søk linje-for-linje (og i 'joined'):
      - brukerens mønster(e) (inkl. elastisk)
      - joined-linje (K A 4 0 1 → KA401)
      - dokument-nivå 'joined'
    Bygg 'Komplett navn' fra ev. prefiks rett foran (360.001- / 360.001/).
    Dedupliser på kanonisk verdi (senere gjort per fil i dedupe_keep_longest()).
    """
    rows: List[Dict] = []
    seen_keys: Set[Tuple[str, str]] = set()
    seen_vals: Set[str] = set()

    rxs = make_search_patterns(user_pat)

    # 1) linje-for-linje
    for raw in lines:
        line = compact_spaces(raw)
        if not line:
            continue

        # direkte søk i linjen
        for rx in rxs:
            start = 0
            while True:
                m = rx.search(line, pos=start)
                if not m:
                    break
                val_raw = m.group(0)
                val_can = canonical(val_raw)

                pre = line[:m.start()]
                pref = ""
                if pre:
                    k = re.search(r"([A-Za-z0-9_.]+[-/])\s*$", pre) or re.search(r"([A-Za-z0-9_.]{2,})\s*$", pre)
                    if k:
                        pref = k.group(1)
                komplett = (pref + val_raw).strip() if pref else val_raw

                key = (val_can, komplett)
                if key not in seen_keys:
                    seen_keys.add(key)
                    seen_vals.add(val_can)
                    rows.append({
                        "Filnavn": filename,
                        "Komplett navn": komplett,
                        "Regex-treff": val_raw,
                        "Komponenttekst (TFM)": tfm_desc(val_can),
                    })
                start = m.end()

        # joined-variant for linjen
        jline = join_word_whitespace(line)
        if jline != line:
            for rx in rxs:
                start = 0
                while True:
                    m = rx.search(jline, pos=start)
                    if not m:
                        break
                    val_raw = m.group(0)
                    val_can = canonical(val_raw)
                    key = (val_can, val_raw)
                    if key not in seen_keys:
                        seen_keys.add(key)
                        if val_can not in seen_vals:
                            seen_vals.add(val_can)
                        rows.append({
                            "Filnavn": filename,
                            "Komplett navn": val_raw,
                            "Regex-treff": val_raw,
                            "Komponenttekst (TFM)": tfm_desc(val_can),
                        })
                    start = m.end()

    # 2) dokument-nivå joined
    full_joined = join_word_whitespace(" ".join(lines))
    for rx in rxs:
        for m in rx.finditer(full_joined):
            val_raw = m.group(0)
            val_can = canonical(val_raw)
            if val_can not in seen_vals:
                seen_vals.add(val_can)
                rows.append({
                    "Filnavn": filename,
                    "Komplett navn": val_raw,
                    "Regex-treff": val_raw,
                    "Komponenttekst (TFM)": tfm_desc(val_can),
                })

    # 3) alltid minst én rad pr fil (viser filnavn ved 0 treff)
    if not rows:
        rows.append({
            "Filnavn": filename,
            "Komplett navn": "",
            "Regex-treff": "",
            "Komponenttekst (TFM)": "",
        })

    return rows

# ==============================
# Dedupe: behold lengste "Komplett navn" pr (Filnavn, Regex-treff)
# ==============================
def dedupe_keep_longest(rows: List[Dict]) -> List[Dict]:
    """
    Dedupliser per (Filnavn, Regex-treff), behold raden med lengst 'Komplett navn'.
    'Regex-treff' dedupliseres case-insensitivt og uten mellomrom/-_/ (canonical).
    'Tom-rad' (0 treff i fil) beholder vi én pr. fil.
    """
    best: dict[tuple[str, str], Dict] = {}
    for r in rows:
        fil = r.get("Filnavn", "")
        match_val = (r.get("Regex-treff") or "").strip()
        if match_val:
            key = (fil, canonical(match_val))
        else:
            key = (fil, "__NOHIT__")

        prev = best.get(key)
        if prev is None or len(r.get("Komplett navn") or "") > len(prev.get("Komplett navn") or ""):
            best[key] = r
    return list(best.values())

# ==============================
# Excel for Søk
# ==============================
def build_search_excel(rows: List[Dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Syntaktisk søk"

    headers = ["Filnavn", "Komplett navn", "Regex-treff", "Komponenttekst (TFM)"]
    ws.append(headers)
    bold = Font(bold=True)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).font = bold

    for r in rows:
        ws.append([
            r.get("Filnavn",""),
            r.get("Komplett navn",""),
            r.get("Regex-treff",""),
            r.get("Komponenttekst (TFM)",""),
        ])

    widths = [40, 34, 24, 42]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Filter på header og frys første rad
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ==============================
# Diff-hjelpere for /diff_sjekk
# ==============================
def normalize_text_for_flags(s: str, ignore_case: bool, ignore_ws: bool, ignore_digits: bool) -> str:
    if ignore_case:
        s = s.lower()
    if ignore_digits:
        s = re.sub(r"\d", "", s)
    if ignore_ws:
        s = compact_spaces(s)
    return s

def extract_text_for_diff(path: str, use_ocr: bool) -> List[str]:
    """Les dokument og returnér linjer (robust for PDF). OCR hvis ønsket og ingen tekst."""
    ext = Path(path).suffix.lower()
    lines = extract_lines_backup(path)
    if ext == ".pdf" and not lines and use_ocr and OCR_AVAILABLE:
        try:
            doc = fitz_open(path)
            for page in doc:
                pix = page.get_pixmap()
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                txt = pytesseract.image_to_string(img) or ""
                if txt:
                    lines.extend(txt.splitlines())
        except Exception:
            logging.exception("OCR feilet")
    return [ln for ln in lines if (ln or "").strip()]

def to_blocks(a_lines: List[str], b_lines: List[str],
              ignore_case: bool, ignore_ws: bool, ignore_digits: bool) -> List[Dict]:
    """Lag blokker (equal/delete/insert/replace) for UI + wordDiff for replace."""
    # Normaliser for sammenligning men behold original for visning
    a_norm = [normalize_text_for_flags(ln, ignore_case, ignore_ws, ignore_digits) for ln in a_lines]
    b_norm = [normalize_text_for_flags(ln, ignore_case, ignore_ws, ignore_digits) for ln in b_lines]

    sm = difflib.SequenceMatcher(a=a_norm, b=b_norm)
    blocks: List[Dict] = []

    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        textA = "\n".join(a_lines[i1:i2]) if i2 > i1 else ""
        textB = "\n".join(b_lines[j1:j2]) if j2 > j1 else ""

        locA = {"page": None, "para": i1 + 1} if i2 > i1 else None
        locB = {"page": None, "para": j1 + 1} if j2 > j1 else None

        if tag == "equal":
            blocks.append({"type": "equal", "textA": textA, "textB": textB,
                           "locationA": locA, "locationB": locB})
        elif tag == "delete":
            blocks.append({"type": "delete", "textA": textA, "textB": "",
                           "locationA": locA, "locationB": None})
        elif tag == "insert":
            blocks.append({"type": "insert", "textA": "", "textB": textB,
                           "locationA": None, "locationB": locB})
        elif tag == "replace":
            # Word-level diff for inline-visning
            wd = word_diff(textA, textB)
            blocks.append({"type": "replace", "textA": textA, "textB": textB,
                           "locationA": locA, "locationB": locB, "wordDiff": wd})
    return blocks

def word_diff(a_text: str, b_text: str) -> List[Dict]:
    """Lag enkel ord-diff (liste av {op:'=', '+', '-', text})."""
    a_words = a_text.split()
    b_words = b_text.split()
    sm = difflib.SequenceMatcher(a=a_words, b=b_words)
    parts: List[Dict] = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            if i2 > i1:
                parts.append({"op": "=", "text": " ".join(a_words[i1:i2])})
        elif tag == "delete":
            if i2 > i1:
                parts.append({"op": "-", "text": " ".join(a_words[i1:i2])})
        elif tag == "insert":
            if j2 > j1:
                parts.append({"op": "+", "text": " ".join(b_words[j1:j2])})
        elif tag == "replace":
            if i2 > i1:
                parts.append({"op": "-", "text": " ".join(a_words[i1:i2])})
            if j2 > j1:
                parts.append({"op": "+", "text": " ".join(b_words[j1:j2])})
    return parts


# =========================
# Ruter – hovedside
# =========================
@dokumentsammenligning_bp.route('/dokumentsammenligning', methods=['GET', 'POST'])
def dokumentsammenligning():
    if request.method == 'GET':
        return render_template('dokumentsammenligning.html')

    # POST (Syntaktisk)
    syntaktisk_action = (request.form.get('syntaktisk_action') or '').strip().lower() or 'diff'
    keyword_pattern = (request.form.get('keyword_pattern') or '').strip()
    use_hoveddokument_keywords_only = request.form.get('use_hoveddokument_keywords_only') == 'on'

    # ---------- Syntaktisk – DIFF (Excel-rapport, uendret) ----------
    if syntaktisk_action == 'diff':
        hoveddokument_file = request.files.get('hoveddokument_file')
        sammenligningsfiler = request.files.getlist('sammenligningsfiler')

        if not keyword_pattern:
            return "Vennligst angi et mønster for nøkkelord."
        if not hoveddokument_file or hoveddokument_file.filename == '':
            return "Vennligst last opp et hoveddokument."
        if not sammenligningsfiler or all(f.filename == '' for f in sammenligningsfiler):
            return "Vennligst last opp minst én sammenligningsfil."

        hoveddokument_path = TEMP_ROOT / secure_filename(hoveddokument_file.filename)
        compare_paths: List[Path] = []

        try:
            hoveddokument_file.save(hoveddokument_path)
            for fil in sammenligningsfiler:
                if fil.filename:
                    p = TEMP_ROOT / secure_filename(fil.filename)
                    fil.save(p)
                    compare_paths.append(p)

            output = generate_report(
                str(hoveddokument_path),
                [str(p) for p in compare_paths],
                keyword_pattern,
                use_hoveddokument_keywords_only,
                use_ocr=False
            )

            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="Dokumentsammenligning.xlsx"
            )

        except Exception as e:
            logging.exception("Feil i dokumentsammenligning (diff):")
            return f"En feil oppsto: {e}"

        finally:
            try:
                if hoveddokument_path.exists():
                    hoveddokument_path.unlink()
            except Exception:
                pass
            for p in compare_paths:
                try:
                    if p.exists():
                        p.unlink()
                except Exception:
                    pass

    # ---------- Syntaktisk – SØK ----------
    if syntaktisk_action == 'search':
        if not keyword_pattern:
            return "Vennligst angi et mønster for nøkkelord."
        files = request.files.getlist('sammenligningsfiler')
        if not files or all(f.filename == '' for f in files):
            return "Vennligst last opp minst én fil å søke i."

        temp_paths: List[Path] = []
        try:
            for f in files:
                if f and f.filename:
                    p = TEMP_ROOT / secure_filename(f.filename)
                    f.save(p)
                    temp_paths.append(p)

            all_rows: List[Dict] = []
            for p in temp_paths:
                lines = merged_lines_for_search(str(p))
                rows = find_matches_build_rows(p.name, lines, keyword_pattern)
                all_rows.extend(rows)

            # dedupe per (Filnavn, Regex-treff), behold lengste "Komplett navn"
            all_rows = dedupe_keep_longest(all_rows)

            bio = build_search_excel(all_rows)
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="SyntaktiskSok.xlsx"
            )

        except Exception as e:
            logging.exception("Feil i syntaktisk søk:")
            return f"En feil oppsto i søk: {e}"

        finally:
            for p in temp_paths:
                try:
                    if p.exists():
                        p.unlink()
                except Exception:
                    pass

    # Ukjent action
    return abort(400, "Ukjent handling for Syntaktisk.")


# =========================
# Diff-sjekk (for fanen “Diff-sjekk” i UI-et)
# =========================
@dokumentsammenligning_bp.route('/diff_sjekk', methods=['POST'])
def diff_sjekk():
    """
    Mottar filA/filB og opsjoner, returnerer JSON:
    { blocks: [{type, textA, textB, locationA, locationB, wordDiff?}, ...] }
    """
    filA = request.files.get('filA')
    filB = request.files.get('filB')
    if not filA or not filB or not filA.filename or not filB.filename:
        abort(400, "Mangler dokument A og/eller B.")

    ignore_case = bool(request.form.get('ignore_case'))
    ignore_ws = bool(request.form.get('ignore_ws', 'on'))  # default true i UI
    ignore_digits = bool(request.form.get('ignore_digits'))
    use_ocr = bool(request.form.get('use_ocr'))

    pA = TEMP_ROOT / secure_filename(filA.filename)
    pB = TEMP_ROOT / secure_filename(filB.filename)

    try:
        filA.save(pA)
        filB.save(pB)

        a_lines = extract_text_for_diff(str(pA), use_ocr=use_ocr)
        b_lines = extract_text_for_diff(str(pB), use_ocr=use_ocr)

        blocks = to_blocks(a_lines, b_lines, ignore_case, ignore_ws, ignore_digits)

        payload = {"blocks": blocks}
        bio = BytesIO(json.dumps(payload, ensure_ascii=False).encode('utf-8'))
        bio.seek(0)
        # frontend forventer JSON direkte
        return (bio.read(), 200, {'Content-Type': 'application/json; charset=utf-8'})

    except Exception as e:
        logging.exception("Feil i /diff_sjekk:")
        abort(500, f"Feil i diff: {e}")

    finally:
        for p in (pA, pB):
            try:
                if p.exists():
                    p.unlink()
            except Exception:
                pass


@dokumentsammenligning_bp.route('/diff_download_excel', methods=['POST'])
def diff_download_excel():
    """
    Mottar JSON-payloadet fra UI (lastDiffPayload) og genererer Excel:
    Kolonner: Endringstype, Posisjon A, Tekst A, Posisjon B, Tekst B
    """
    try:
        payload = request.get_json(force=True, silent=False)
        blocks = payload.get("blocks", []) if isinstance(payload, dict) else []

        wb = Workbook()
        ws = wb.active
        ws.title = "Diff"

        headers = ["Endringstype", "Posisjon A", "Tekst A", "Posisjon B", "Tekst B"]
        ws.append(headers)
        bold = Font(bold=True)
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i).font = bold

        def loc_str(loc: Optional[Dict]) -> str:
            if not loc:
                return ""
            bits = []
            if loc.get("page") is not None:
                bits.append(f"Side {loc['page']}")
            if loc.get("para") is not None:
                bits.append(f"Avsnitt {loc['para']}")
            return ", ".join(bits)

        for b in blocks:
            t = b.get("type", "")
            a_loc = loc_str(b.get("locationA"))
            b_loc = loc_str(b.get("locationB"))
            a_txt = b.get("textA") or ""
            b_txt = b.get("textB") or ""
            ws.append([t.capitalize(), a_loc, a_txt, b_loc, b_txt])

        # litt formatering
        widths = [16, 18, 80, 18, 80]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name="Diff.xlsx")

    except Exception as e:
        logging.exception("Feil i /diff_download_excel:")
        abort(500, f"Feil ved generering av Excel: {e}")


# ---------- Kombinert søk (uendret fra backup) ----------
@dokumentsammenligning_bp.route('/kombinert_sok', methods=['POST'])
def kombinert_sok():
    data = request.form.get('kombinert_data')
    files = request.files.getlist('kombinert_filer')
    if not data or not files:
        abort(400, "Manglende tabell eller filer for kombinert søk.")
    table = json.loads(data)

    # Lagre midlertidige filer
    paths = []
    for f in files:
        if f and f.filename:
            p = TEMP_ROOT/secure_filename(f.filename)
            f.save(p)
            paths.append(str(p))

    # Ekstraher tekst for hvert dokument (samme som i backup)
    texts = {p: extract_lines_backup(p) for p in paths}

    wb = Workbook()
    ws = wb.active
    ws.title = 'KombinertSok'

    maxc = max(len(r) for r in table)
    headers = [f'Kolonne {i+1}' for i in range(maxc)] + ['Dokument','Status','Avvik']
    ws.append(headers)

    # Generer fargepalette
    colors = []
    for i in range(len(paths)):
        hue = i / len(paths)
        import colorsys as cs
        r,g,b = cs.hls_to_rgb(hue, 0.8, 0.4)
        colors.append((int(r*255),int(g*255),int(b*255)))

    THRESH = 80

    for row in table:
        raw_vals = [c.strip().lstrip('=') if isinstance(c, str) else '' for c in row] + [''] * (maxc - len(row))
        display_vals = [(' ' + c) if isinstance(c, str) and c.startswith('=') else c for c in row] + [''] * (maxc - len(row))
        alt2 = raw_vals[1].replace(',', '.') if len(raw_vals) > 1 and raw_vals[1] else ''

        for idx, path in enumerate(paths):
            lines = texts[path]
            hit_lines = []
            for i, text in enumerate(lines, start=1):
                if raw_vals[0]:
                    if raw_vals[0] in text:
                        hit_lines.append(i)
                    elif fuzz.partial_ratio(raw_vals[0], text) > THRESH:
                        hit_lines.append(i)
                    elif fuzz.token_set_ratio(raw_vals[0], text) > THRESH:
                        hit_lines.append(i)

            status = 'Ingen treff'
            avvik = ''

            if hit_lines:
                all_ok = True
                for col_idx, val in enumerate(raw_vals[:maxc], start=1):
                    if not val:
                        continue
                    matched = False
                    for ln in hit_lines:
                        line_text = lines[ln-1]
                        if val in line_text or fuzz.partial_ratio(val, line_text) > THRESH:
                            matched = True
                            break
                    if not matched:
                        all_ok = False
                status = 'Komplett' if all_ok else 'Delvis'
                if not all_ok and raw_vals[1]:
                    found_elsewhere = False
                    for i, text in enumerate(lines, start=1):
                        if (raw_vals[1] in text or
                            (alt2 and alt2 in text) or
                            fuzz.partial_ratio(raw_vals[1], text) > THRESH or
                            (alt2 and fuzz.partial_ratio(alt2, text) > THRESH) or
                            fuzz.token_set_ratio(raw_vals[1], text) > THRESH):
                            found_elsewhere = True
                            avvik = f"{raw_vals[1]} funnet i linje {i}"
                            break
                    if not found_elsewhere:
                        avvik = f"{raw_vals[1]}, ikke funnet i samme dokument"

            # Skriv rad
            row_index = ws.max_row + 1
            out_vals = display_vals + [Path(path).name, status, avvik]
            for col_num, v in enumerate(out_vals, start=1):
                cell = ws.cell(row=row_index, column=col_num, value=v)
                cell.number_format = '@'
                if col_num <= maxc:
                    if status.startswith('Komplett') or (status == 'Delvis' and col_num == 1):
                        r, g, b = colors[idx]
                        cell.fill = PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                                                end_color=f"{r:02X}{g:02X}{b:02X}",
                                                fill_type='solid')

    # Autofilter og autojuster kolonner
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column+1):
        letter = get_column_letter(col)
        max_len = max(len(str(c.value or '')) for c in ws[letter])
        ws.column_dimensions[letter].width = min(max_len + 2, 80)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    username = getattr(current_user, 'username', 'bruker')
    filename = f"{username}_KombinertSammenligning.xlsx"

    for p in paths:
        try:
            Path(p).unlink()
        except Exception:
            pass

    return send_file(bio, as_attachment=True, download_name=filename)
