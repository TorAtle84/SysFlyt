# Fil: app/routes/protokoller_core.py

import logging, joblib
from flask import Blueprint, request, jsonify, Response, stream_with_context, render_template, current_app
import io, re, json, os
import fitz, docx
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.models.user import User
from flask_login import current_user, login_required
from typing import Dict, Iterable, Iterator, Optional

# ------------------- Blueprint og grunnkonfig -------------------
bp = Blueprint("protokoller", __name__, url_prefix="/protokoller")
basedir = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.abspath(os.path.join("app", "data"))
STATIC_DATA_DIR = os.path.abspath(os.path.join("app", "static", "data"))
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(STATIC_DATA_DIR, exist_ok=True)

# ------------------- TFM-ordbok (uendret) -------------------
TFM_DICT = {"AB": "Bjelke", "AD": "Dekker", "AE": "Elementer", "AF": "Fagverk", "AG": "Glassfelt", "AH": "Fundamenter",
"AK": "Komplette konstruksjoner", "AL": "List - Beslistning", "AO": "Oppbygende - Utforende", "AP": "Plate", "AR": "Ramme - Oppheng",
"AS": "Søyle", "AU": "Åpninger", "AV": "Vegg", "BA": "Armering Forsterkning", "BB": "Beskyttende - Stoppende", "BC": "Begrensende",
"BF": "Fuging", "BG": "Pakning Tetning", "BI": "Isolasjon", "BP": "Avretningsmasse", "BS": "Spikerslag", "CB": "Balkong",
"CC": "Baldakin", "CD": "Karnapp", "CG": "Rampe - Repos", "CK": "Komplett konstruksjon", "CM": "Kjemiske stoffer", "CO": "Kobling Overgang",
"CP": "Pipe - Skostein", "CQ": "Festemateriell", "CR": "Rammeverk - Oppheng", "CT": "Trapp - Leider", "CX": "Tunnel - Bru (inne-ute)",
"DB": "Dør med brannklasse", "DF": "Foldedør - Foldevegg", "DI": "Dør - innvendig", "DK": "Kjøretøyregulering", "DL": "Luke", "DP": "Port",
"DT": "Dør - tilfluktsrom", "DU": "Dør - utvendig", "DV": "Vindu", "EB": "Overflatebekledning", "EC": "Overflatebehandling", "EH": "Himling",
"FA": "Ventilert arbeidsplass", "FB": "Benk - Bord - Plate - Tavle", "FC": "Beslag", "FD": "Disk - Skranke", "FF": "Fryserom", "FH": "Hylle - Reol",
"FI": "Kabinett", "FK": "Kjølerom - Svalrom", "FO": "Sittebenk - Sofa - Stol", "FR": "Rom", "FS": "Skap - Skuff", "FT": "Speil",
"FV": "Vaskemaskin", "FX": "Krok - Knagg - Håndtak", "GA": "Automat - Maskin", "GB": "Benk - Bord - Plate - Tavle", "GD": "Dekontamator",
"GE": "Autoklaver", "GF": "Fryseskap", "GG": "Gardiner - Forheng", "GH": "Hylle - Reol", "GK": "Kjøleskap - Kjøledisk", "GL": "Lås - Beslag",
"GM": "Mattilbereding", "GN": "Nøkler", "GO": "Sofa - Sittebenk", "GP": "Stol", "GQ": "Seng - Liggebenk", "GS": "Skap - Skuffer",
"GT": "Tørkeskap - Varmeskap", "GV": "Vaskemaskin", "GW": "Vekt", "GX": "Holder", "GY": "Avfallsbeholder", "HA": "Automobil - Bil",
"HB": "Rullebord - Tralle", "HC": "Container - Vogn", "HM": "Maskin", "HS": "Rullestol", "HT": "Truck - kran", "HV": "Verktøy",
"IB": "Brenner", "IC": "Solceller-solfangere", "ID": "Kjel for destruksjon", "IE": "Elektrokjel", "IF": "Kjel for fast-bio brensel",
"IG": "Generator", "IK": "Kuldeaggregat", "IL": "Energibrønn", "IM": "Motor", "IO": "Oljekjel", "IP": "Gasskjel", "IT": "Trykkluftaggregat (enhet)",
"IU": "Turbin", "IV": "Aggregatenhet", "JF": "Forsterker", "JK": "Kompressor", "JP": "Pumpe", "JQ": "Pumpe i VA-installasjoner", "JV": "Vifte",
"JW": "Spesialvifte", "KA": "Aktuator", "KD": "Drivhjul - Drev", "KE": "Induktiv energioverføring", "KG": "Gjennomføring",
"KH": "Transportenhet (hevende forflyttende)", "KJ": "Jordingskomponenter", "KK": "Kanal", "KM": "Mast - Antenne", "KN": "Nedløp",
"KO": "Kraftoverføring", "KQ": "Rør - spesielt", "KR": "Rør - generelt", "KS": "Skinne - Bane - Spor", "KU": "Kombinert kabel",
"KV": "Høyspenningskabel > 1000V", "KW": "Lavspenningskabel 50 til 1000V", "KX": "Lavspenningskabel < 50V", "KY": "Optisk kabel",
"KZ": "Slange", "LB": "Varmeomformende med vifte", "LC": "Kjøleomformende med vifte", "LD": "Kjøleflater", "LE": "Kondensator",
"LF": "Fordamper", "LG": "Gear clutch", "LH": "Varmeflate", "LI": "Varmeelement", "LK": "Kjøleomformende", "LL": "Lyskilde",
"LN": "Likeretter", "LO": "Omformer", "LP": "Pens - Veksel - Sjalter", "LQ": "Vekselretter", "LR": "Frekvensomformer",
"LS": "Strålevarme", "LU": "Luftfukter", "LV": "Varmeomformende", "LX": "Varmegjenvinner", "LZ": "Varmerkabel - Varmerør",
"MA": "Absoluttfilter", "MB": "ABC-filter", "MC": "UV-filter", "ME": "Elektrostatiske filter", "MF": "Luftfilter", "MG": "Fettfilter",
"MK": "Kondenspotte", "ML": "Luftutskiller", "MM": "Membran", "MO": "Utskiller", "MR": "Rist - Sil", "MS": "Syklon", "MT": "Tørke",
"MU": "Filter for Lyd - Bilde - Frekvensutjevner", "MV": "Vannfilter", "MX": "Støyfilter", "NB": "Batteri - UPS", "NC": "Kondensator",
"NI": "Informasjonslagring", "NK": "Kum", "NM": "Badekar - Basseng", "NO": "Åpen tank", "NT": "Tank med trykk", "NU": "Tank uten trykk",
"NV": "Vekt Lodd", "NW": "Varmtvannsbereder", "NX": "Toalett", "NY": "Servant", "NZ": "Brannslukkingsapparat", "OA": "AV-maskiner",
"OB": "Shuntgruppe", "OD": "Datamaskin", "OE": "Energimåler", "OF": "Systembærer", "OM": "Mottaker - Sender", "OP": "PBX",
"OQ": "Dataprogramprogramvare", "OR": "Router Fordeler", "OS": "Sentralenhet Mikser i Lydsystem", "OT": "Telefonapparat",
"OU": "Undersentral", "QB": "Belastningsvakt", "QD": "Differansetrykkvakt", "QE": "Elektrisk vern", "QF": "Strømningsvakt",
"QH": "Fuktvakt", "QL": "Lyddemper", "QM": "Mekanisk beskyttelse", "QN": "Nivåvakt", "QO": "Overtrykksventil", "QP": "Trykkvakt",
"QQ": "Vibrasjonsvakt", "QR": "Rotasjonsvakt", "QS": "Strømvakt", "QT": "Temperaturvakt", "QV": "Sikkerhetsventil", "QX": "Solavskjerming",
"QY": "Lynavleder", "QZ": "Brannvern", "RA": "AV-opptaker", "RB": "Bevegelse", "RC": "Seismometer", "RD": "Differansetrykkgiver",
"RE": "Elektriske variabler", "RF": "Strømningsmåler", "RG": "Posisjon - Lengde", "RH": "Fuktighetsgiver", "RI": "Termometer",
"RJ": "Fotocelle", "RK": "Kortleser", "RM": "Multifunksjonell Kombinert føler", "RN": "Nivågiver", "RP": "Trykkgiver",
"RQ": "Manometer Trykkmåler", "RR": "Giver generelt", "RS": "Hastighetsmåler", "RT": "Temperaturgiver", "RU": "Ur", "RV": "Veieceller",
"RW": "Virkningsgradsmåler", "RX": "Målepunkt", "RY": "Gassdetektor-Røykdetektor", "RZ": "Branndeteksjon",
"SA": "Reguleringsventil manuell", "SB": "Reguleringsventil motorstyrt", "SC": "Stengeventil motorstyrt",
"SD": "Alarmventil sprinkler", "SE": "Ekspansjonsventil", "SF": "Fraluftsventil", "SG": "Tilbakeslagsventil - Overtrykkspjeld",
"SH": "Hurtigkobling", "SI": "Effektregulator", "SJ": "Jevntrykksventil", "SM": "Stengeventil manuell", "SP": "Trykkutjevningsventil",
"SQ": "Strømningsregulator - VAV", "SR": "Reguleringsspjeld", "SS": "Stengespjeld", "ST": "Tilluftsventil", "SU": "Sugetrykksventil",
"SV": "Strupeventil", "SW": "Plenumskammer", "SX": "Regulator", "SZ": "Brannspjeld - Røyks-p-j-e-l-d",
"UA": "Uttak alarm", "UD": "Uttak data", "UE": "Uttak el", "UF": "Fellesuttak", "UG": "Uttak gass", "UK": "Kontrollpanel - Tablå",
"UL": "Uttak trykkluft", "UM": "Monitor - Display", "UN": "Nødbelysning", "UO": "Trommel", "UP": "Belysningsarmatur", "UR": "Uttak radio",
"US": "Stasjon", "UT": "Uttak telefon", "UV": "Uttak vann", "UX": "Koblingsboks", "UY": "Uttak antenne", "UZ": "Dyse - Spreder",
"VB": "Bærelag", "VD": "Dekke", "VG": "Gress", "VK": "Kantstein heller", "VL": "Masse", "VM": "Mekanisk beskyttelse", "VP": "Planterbuskertrær",
"VS": "Skilt", "XC": "Kondensator", "XD": "Komp. for binærlogikk", "XF": "Komponenter for vern", "XG": "Komponenter for krafttilførsel",
"XH": "Komponenter for signalering", "XI": "Potensiometer", "XK": "Releer - Kontaktorer", "XL": "Induktiv komponenter", "XM": "Motor",
"XN": "Integrerte kretser", "XO": "Urbryter - Timer", "XP": "Komponenter for måling og prøving", "XQ": "Effektbryter", "XR": "Motstand",
"XS": "Bryter / Vender", "XT": "Transformator", "XV": "Halvlederkomponenter og elektronrør", "XX": "Rekkeklemmer - Samlesignal",
"XZ": "Terminering og tilpasning"}

# ------------------- COMPAT: RegexCounts + trygg modelllasting -------------------
class RegexCounts:
    """Compatibility stub for older pickled models that reference RegexCounts."""
    pass

DEFAULT_FAGMODELL_PATH = os.environ.get(
    "PROSJEKTBASEN_FAGMODELL",
    os.path.join(DATA_DIR, "models", "fag_model.joblib")
)

_fag_model_cache = {"loaded": False, "model": None, "error": None}

def get_fag_model(path: Optional[str] = None):
    """
    Laster fag-modellen trygt (én gang). Returnerer None hvis lasting feiler.
    Sørger for at __main__.RegexCounts finnes for gamle pickle-filer.
    """
    global _fag_model_cache
    if _fag_model_cache["loaded"]:
        return _fag_model_cache["model"]

    model_path = path or DEFAULT_FAGMODELL_PATH
    try:
        # Kritisk: sørg for at __main__ har RegexCounts før load
        import importlib
        main_mod = importlib.import_module("__main__")
        if not hasattr(main_mod, "RegexCounts"):
            setattr(main_mod, "RegexCounts", RegexCounts)

        _fag_model_cache["model"] = joblib.load(model_path)
        _fag_model_cache["loaded"] = True
        return _fag_model_cache["model"]
    except Exception as e:
        _fag_model_cache["error"] = e
        _fag_model_cache["loaded"] = True
        logging.error("Fag-modell feilet: %s", e)
        return None


# ------------------- Streng/generisk regex + parser-API -------------------
# GENERELL (Excel/Doc/Txt) – tolerant
# NYTT: komponent-mønster støtter 2–4 bokstaver, fleksibel tallblokk og valgfri /NNN
MASTER_REGEX_GENERIC = re.compile(
    r"(?P<full_tag>"
    r"(?:\+(?P<byggnr>[A-Z0-9]+))?"                                   # +BYGG
    r"(?:=(?P<system>[\d.:/]+))?"                                     # =SYSTEM (tolerant; støtter kolon)
    r"(?:-?(?P<komponent>([A-Za-z]{2,4})[A-Za-z0-9]{0,6}\d{2,5}[A-Za-z0-9/]*))?"  # -LK009T/001
    r"(?:%(?P<typekode>[A-Z0-9./:_-]+))?"                             # %TYPE (tolerant)
    r")",
    re.IGNORECASE,
)

# STRENG (PDF) – “kynisk”
# NYTT: utvidet komponent-mønster identisk med over. System beholdes strengt (0000.0000), vi henter ev. kolon-del fra full_tag ved behov.
MASTER_REGEX_PDF = re.compile(
    r"(?P<full_tag>"
    r"(?:\+(?P<pdf_byggnr>[^=\s+]+)=(?=\d))?"                         # +BYGGNR=
    r"(?:=(?P<pdf_system>\b\d{3,4}\.\d{3,4}\b))?"                     # =000.000 / 0000.0000 / ...
    r"(?:-?(?P<komponent>([A-Za-z]{2,4})[A-Za-z0-9]{0,6}\d{2,5}[A-Za-z0-9/]*))?"  # -LK009T/001
    r"(?:%(?P<pdf_typekode>[^\s+]+))?"                                # %TYPE til whitespace eller '+'
    r")",
    re.IGNORECASE,
)

def select_regex_for_filename(filename: Optional[str]):
    """Velger riktig regex basert på filtype (PDF => streng)."""
    fn = (filename or "").lower()
    return MASTER_REGEX_PDF if fn.endswith(".pdf") else MASTER_REGEX_GENERIC

def get_unique_system_id(system_string: Optional[str]) -> str:
    """
    Returnerer kanonisk system-ID: 3–4 siffer, punktum, 3–4 siffer.
    Fallback: første tall-blokk i starten. 'Uspesifisert' hvis ingenting.
    (Hvis inn-kommer '360.0002:001', normaliseres det til '360.0002'.)
    """
    if not system_string:
        return "Uspesifisert"
    s = system_string.strip()
    m = re.search(r'\b(\d{3,4}\.\d{3,4})\b', s)
    if m:
        return m.group(1)
    m2 = re.match(r'^\d+(?:\.\d+)?', s)
    return m2.group(0) if m2 else "Uspesifisert"

def iter_tags(text: str, filename: Optional[str]) -> Iterator[Dict[str, str]]:
    """
    Generator: finner tagger i tekst, normaliserer til keys:
    byggnr, system, komponent (uppercase), typekode, full_tag.
    """
    rx = select_regex_for_filename(filename)
    for m in rx.finditer(text or ""):
        gd = m.groupdict()
        byggnr    = gd.get("byggnr")    or gd.get("pdf_byggnr")
        system    = gd.get("system")    or gd.get("pdf_system")
        komponent = gd.get("komponent") or gd.get("pdf_komponent")
        typekode  = gd.get("typekode")  or gd.get("pdf_typekode")
        full_tag  = (gd.get("full_tag") or "").strip()

        # NYTT: robust fallback – “siste bindestrek + komponent”-detektor
        if not komponent and full_tag:
            m_fallback = re.search(
                r"-([A-Za-z]{2,4}[A-Za-z0-9]{0,6}\d{2,5}[A-Za-z0-9]*?(?:/[0-9]{1,4})?)(?=$|[\s,.;:\)\]\}])",
                full_tag,
                re.IGNORECASE
            )
            if m_fallback:
                komponent = m_fallback.group(1)

        if not komponent:
            continue

        yield {
            "byggnr": (byggnr or "").strip(),
            "system": (system or "").strip(),
            "komponent": (komponent or "").strip().upper(),
            "typekode": (typekode or "").strip(),
            "full_tag": full_tag,
        }

def parse_rows_from_text(text: str, filename: Optional[str]) -> Iterable[Dict[str, str]]:
    """
    Konverterer funn til rader for UI/Excel:
    - unique/system_number/system_full_name settes til kanonisk ID
    - full_id = full_tag
    """
    for tag in iter_tags(text, filename):
        uid = get_unique_system_id(tag["system"])
        yield {
            "source": filename or "",
            "unique_system": uid,
            "system_number": uid,
            "system_full_name": uid,
            "full_id": tag["full_tag"],
            "komponent": tag["komponent"],
            "typekode": tag["typekode"],
            "byggnr": tag["byggnr"],
        }

# Bakoverkompatibilitet (hvis noe fortsatt importerer MASTER_REGEX)
MASTER_REGEX = MASTER_REGEX_GENERIC

def extract_tfm_prefix(komponent: str):
    if not komponent: return None
    m = re.match(r'^[A-Z]{2,3}', komponent.strip().upper())
    return m.group(0) if m else None

# ------------------- Les tekst fra filer -------------------
def extract_text_from_file(file_storage):
    filename = (file_storage.filename or "").lower()
    try:
        file_storage.seek(0)
    except Exception:
        pass

    # Råtekst
    if filename.endswith((".csv", ".txt")):
        try:
            return file_storage.read().decode("utf-8-sig")
        except Exception:
            try:
                file_storage.seek(0)
            except Exception:
                pass
            return file_storage.read().decode("latin-1", errors="ignore")

    # Les bytes én gang
    file_bytes = file_storage.read()
    bio = io.BytesIO(file_bytes)

    if filename.endswith(".pdf"):
        try:
            doc = fitz.open(stream=bio, filetype="pdf")
            return "".join([p.get_text() for p in doc])
        except Exception:
            return ""

    if filename.endswith(".docx"):
        try:
            d = docx.Document(bio)
            return "\n".join(p.text for p in d.paragraphs)
        except Exception:
            return ""

    if filename.endswith(".xlsx"):
        try:
            wb = load_workbook(bio, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    parts = [str(c) for c in row if c is not None]
                    if parts:
                        lines.append(" ".join(parts))
            return "\n".join(lines)
        except Exception:
            return ""

    return ""

# ------------------- Routes -------------------
@bp.route("/generate_underlag", methods=["POST"])
@login_required
def generate_underlag():
    """
    MC: Streamer NDJSON-linjer. Bufferer filer i RAM (unngår 'I/O on closed file').
    Støtter systemkriterier (prefiks) og TFM-beskrivelse fra komponentens to første bokstaver.
    """
    files = request.files.getlist("files")

    # --- Systemkriterier (eks. "36" eller "36, 37")
    raw_criteria = (request.form.get("system_kriterier") or "").strip()
    allowed_prefixes = [s for s in re.split(r"[,\s]+", raw_criteria) if s]

    def _allowed(uid: str) -> bool:
        if not allowed_prefixes:
            return True
        s = str(uid or "")
        return any(s.startswith(pref) for pref in allowed_prefixes)

    # 1) Kopiér alle filer til RAM først
    files_mem = []
    for f in files:
        if not f or not f.filename:
            continue
        try:
            f.seek(0)
        except Exception:
            pass
        files_mem.append((f.filename, f.read()))

    # --- HJELPERE -------------------------------------------------------------
    _COMP_RE = re.compile(r"-?([A-Za-z]{2,4}\d{2,5})(?:%[^\s+]+)?")  # plukk ut KOMPID fra full_id
    _KOMP_FROM_ID = re.compile(r"^[A-Za-z]{2,4}\d+")

    def parse_component_from_full_id(full_id: str) -> str | None:
        """Returner 'UEA0122' fra '+524=4320.001-UEA0122%...' osv."""
        s = str(full_id or "")
        m = _COMP_RE.search(s)
        return m.group(1) if m else None

    def tfm_prefix_from_component(komp: str) -> str:
        """
        TFM-lookup bruker normalt 2 bokstaver. Vi tar de to første bokstavene
        av komponent-ID'en uavhengig av om den har 2 eller 3 bokstaver (UEA -> UE).
        """
        if not komp:
            return ""
        letters = re.match(r"^[A-Za-z]{2,}", komp)
        return (letters.group(0)[:2].upper()) if letters else ""

    # -------------------------------------------------------------------------

    def generate():
        all_rows = []
        print("\n--- MC: starter ---")
        if allowed_prefixes:
            print(f"[MC] Systemkriterier aktiv: {allowed_prefixes}")
        else:
            print("[MC] Ingen systemkriterier valgt (viser alle)")

        for filename, file_bytes in files_mem:
            yield f'data: {json.dumps({"currentFile": filename})}\n\n'
            try:
                class _FS:
                    def __init__(self, name, data):
                        self.filename = name
                        self._bio = io.BytesIO(data)
                    def read(self): return self._bio.read()
                    def seek(self, pos): self._bio.seek(pos)

                text = extract_text_from_file(_FS(filename, file_bytes))
                hits_total = 0
                hits_after_filter = 0

                for row in parse_rows_from_text(text, filename):
                    hits_total += 1
                    uid = row.get("unique_system", "")
                    if not _allowed(uid):
                        continue
                    hits_after_filter += 1

                    # --- Robust komponent finnes ofte allerede via iter_tags; behold fallback:
                    komponent = (row.get("komponent") or
                                 parse_component_from_full_id(row.get("full_id", "")) or
                                 "")

                    # --- Slå opp TFM på de to første bokstavene av komponenten
                    tfm_key = tfm_prefix_from_component(komponent)
                    desc = TFM_DICT.get(tfm_key, "Ukjent beskrivelse")

                    all_rows.append({
                        "source": row.get("source", filename),
                        "unique_system": uid,
                        "full_id": row.get("full_id", ""),
                        "komponent": komponent,
                        "desc": desc
                    })

                print(f"[MC] {filename}: {hits_after_filter}/{hits_total} funn etter filter")
            except Exception as e:
                msg = f"Feil under {filename}: {type(e).__name__}: {e}"
                print("[MC][FEIL]", msg)
                yield f'data: {json.dumps({"error": msg})}\n\n'

        print(f"--- MC: ferdig; total {len(all_rows)} rader ---")
        yield f'data: {json.dumps({"rows": all_rows})}\n\n'

    return Response(stream_with_context(generate()), mimetype="application/x-ndjson")

@bp.route("/api/tfm-liste")
@login_required
def hent_tfm_liste():
    funksjon = request.args.get("funksjon", "").upper()
    mapping = {
        "MC": "tfm-settings-mc.json",
        "FUNKSJONSTEST": "tfm-settings-funksjonstest.json",
        "INNREGULERING": "tfm-settings-innregulering.json"
    }
    if funksjon not in mapping:
        return jsonify({"error": "Ugyldig funksjonstype"}), 400
    settings_path = os.path.abspath(os.path.join(basedir, "..", "static", "data", mapping[funksjon]))
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bp.route("/api/tfm-liste/save", methods=["POST"])
@login_required
def lagre_tfm_liste():
    data = request.get_json()
    funksjon = request.args.get("funksjon", "mc").lower()
    if not data:
        return jsonify({"error": "Manglende data"}), 400
    filnavn = f"tfm-settings-{funksjon}.json"
    path = os.path.join(current_app.static_folder, "data", filnavn)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@bp.route("/api/users")
@login_required
def api_users():
    users = User.query.filter_by(approved=True).all()
    return jsonify([f"{u.first_name} {u.last_name}" for u in users])

@bp.route("/api/technicians")
@login_required
def api_technicians():
    techs = User.query.filter_by(role="tekniker", approved=True).all()
    return jsonify([f"{t.first_name} {t.last_name}" for t in techs])

@bp.route("/api/me")
@login_required
def api_me():
    return f"{current_user.first_name} {current_user.last_name}"

@bp.route("/")
@login_required
def index():
    return render_template("protokoller.html")

# ------------------- Funksjonsbank (uendret) -------------------
_FUNKBANK_CACHE = {"mtime": None, "data": {}}

def _funbank_excel_path():
    return os.path.join(DATA_DIR, "Funksjonstest_db.xlsx")

def _load_bank_from_excel(selected=None):
    xlsx_path = _funbank_excel_path()
    if not os.path.exists(xlsx_path):
        return {}
    mtime = os.path.getmtime(xlsx_path)
    if _FUNKBANK_CACHE["mtime"] != mtime:
        wb = load_workbook(xlsx_path, data_only=True)
        data = {}
        for ws in wb.worksheets:
            code = (ws.title or "").strip().upper()
            if not code: 
                continue
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header0 = str(rows[0][0]).lower() if rows and rows[0] and rows[0][0] else ""
            start_idx = 1 if any(k in header0 for k in ["beskrivelse", "navn"]) else 0
            bank = []
            for row in rows[start_idx:]:
                if not row or not any(row):
                    continue
                navn = (str(row[0]).strip() if len(row) > 0 and row[0] else "")
                test = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
                aksept = (str(row[2]).strip() if len(row) > 2 and row[2] else "")
                if not (navn or test or aksept):
                    continue
                bank.append({"navn": navn, "test": test, "aksept": aksept})
            if bank:
                data[code] = bank
        _FUNKBANK_CACHE["data"] = data
        _FUNKBANK_CACHE["mtime"] = mtime
    if selected:
        wanted = {t.strip().upper() for t in selected if t and t.strip()}
        return {k: v for k, v in _FUNKBANK_CACHE["data"].items() if k in wanted}
    return _FUNKBANK_CACHE["data"]

@bp.route("/api/funksjonsbank")
@login_required
def api_funksjonsbank():
    tfm_query = request.args.get("tfm", "")
    wanted = set(t.strip().upper() for t in tfm_query.split(",") if t.strip()) or None
    return jsonify(_load_bank_from_excel(wanted))

# ------------------- Excel-hjelpere (uendret) -------------------
PASTELL_LYSEBLAA = "DCEBFF"
SECTION_ORDER = ["Start og Stopp funksjoner", "Reguleringsfunksjoner", "Sikkerhetsfunksjoner", "Øvrig"]

def insert_section_header(ws, row_index: int, title: str):
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row_index <= mr.max_row:
            to_unmerge.append(mr)
    for mr in to_unmerge:
        ws.unmerge_cells(str(mr))
    max_col = max(14, ws.max_column)
    for col in range(1, max_col + 1):
        ws.cell(row=row_index, column=col, value=None)
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=14)
    c = ws.cell(row=row_index, column=1, value=title)
    c.fill = PatternFill(fill_type="solid", start_color=PASTELL_LYSEBLAA, end_color=PASTELL_LYSEBLAA)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.font = Font(bold=True)

def add_betingelser(worksheet):
    conditions_text = [
        "",
        "Kontroll av protokoll",
        "Dokumentasjon må overleveres før igangkjøring av anlegg.",
        "Protokollen skal fylles ut og signeres av utførende montør.",
        "Protokollen skal fylles ut og signeres av kontrollør som kontrollerer arbeidet.",
        "Dokumentasjonen skal fylles ut fortløpende avhengig av fremdrift i prosjektet."
    ]
    start_row = 202
    for i, line in enumerate(conditions_text):
        worksheet.cell(row=start_row + i, column=1, value=line)
        if line:
            worksheet.merge_cells(start_row=start_row + i, start_column=1, end_row=start_row + i, end_column=9)
            worksheet.cell(row=start_row + i, column=1).font = Font(bold=True)
            worksheet.cell(row=start_row + i, column=1).alignment = Alignment(wrapText=True)