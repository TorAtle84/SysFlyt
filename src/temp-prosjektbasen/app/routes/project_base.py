import pandas as pd
import os
import sqlite3
import openpyxl
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, jsonify
from flask_login import login_required, current_user
from app.models import Project, Merkskilt, db, User
from app.models.systembygging import Systembygging
from app.models.komponentopptelling import Komponentopptelling
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell

project_bp = Blueprint('project_bp', __name__)

ALL_LOKASJONER = ["Bergen", "Oslo", "Trondheim", "Region Vest", "Region Øst", "Region Nord", "Region Sør"]

@project_bp.route('/prosjektoversikt')
@login_required
def vis_prosjektoversikt():
    return redirect(url_for('project_bp.prosjektbase'))


def get_db_connection():
    db_path = current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


@project_bp.route('/prosjektbase', methods=['GET'])
@login_required
def prosjektbase():
    selected_location = request.args.get("location")
    locations = ALL_LOKASJONER

    if selected_location:
        projects = Project.query.filter_by(location=selected_location).order_by(Project.project_name).all()
    else:
        projects = Project.query.order_by(Project.project_name).all()

    users = User.query.order_by(User.first_name, User.last_name).all()

    return render_template(
        'prosjektbase.html',
        locations=locations,
        selected_location=selected_location,
        projects=projects,
        users=users
    )


@project_bp.route('/opprett-prosjekt', methods=['POST'])
@login_required
def opprett_prosjekt():
    project_name = request.form['project_name']
    order_number = request.form['order_number']
    location = request.form['location']
    description = request.form['description']

    # === Slå opp prosjektleder etter navn ===
    leader_name = request.form['project_leader_id'].strip()
    leader = User.query.filter(
        (User.first_name + " " + User.last_name) == leader_name
    ).first()
    if not leader:
        flash(f"Fant ikke prosjektleder med navn: {leader_name}", "danger")
        return redirect(url_for("project_bp.prosjektbase"))

    # === Slå opp prosjekteier etter navn (valgfritt) ===
    manager_id = None
    manager_name = request.form.get("project_manager_id", "").strip()
    if manager_name:
        manager = User.query.filter(
            (User.first_name + " " + User.last_name) == manager_name
        ).first()
        if not manager:
            flash(f"Fant ikke prosjekteier med navn: {manager_name}", "danger")
            return redirect(url_for("project_bp.prosjektbase"))
        manager_id = manager.id

    nytt_prosjekt = Project(
        project_name=project_name,
        order_number=order_number,
        location=location,
        project_leader_id=leader.id,
        project_manager_id=manager_id,
        description=description
    )

    db.session.add(nytt_prosjekt)
    db.session.commit()

    flash('Prosjektet ble opprettet.', 'success')
    return redirect(url_for("project_bp.prosjektbase"))

@project_bp.route("/prosjekt")
@login_required
def vis_prosjekt():
    project_id = request.args.get("id", type=int)
    active_tab = request.args.get("tab", "systembygging")  # 👈 standard nå systembygging

    if not project_id:
        flash("Prosjekt-ID mangler.")
        return redirect(url_for("project_bp.prosjektbase"))

    project = Project.query.get(project_id)
    if not project:
        flash("Prosjektet ble ikke funnet.")
        return redirect(url_for("project_bp.prosjektbase"))

    merkeskilt_posts = Merkskilt.query.filter_by(project_id=project.id).all() if active_tab == "merkeskilt" else []
    system_rader = Systembygging.query.filter_by(project_id=project.id).all() if active_tab == "systembygging" else []
    komponent_rader = Komponentopptelling.query.filter_by(project_id=project.id).all() if active_tab == "komponentopptelling" else []

    from collections import defaultdict

    system_grupper = defaultdict(list)

    for rad in system_rader:
        if rad.full_id:
            base = rad.full_id.split(":")[0]
            system_grupper[base].append(rad)

    return render_template(
        "project_detail.html",
        project=project,
        active_tab=active_tab,
        merkeskilt_posts=merkeskilt_posts,
        system_rader=system_rader,
        komponent_rader=komponent_rader,
        system_grupper=system_grupper
    )



# === API: Hent lokasjoner ===
@project_bp.route("/api/lokasjoner")
@login_required
def api_hent_lokasjoner():
    lokasjoner = User.query.with_entities(User.location).distinct().all()
    return jsonify(sorted([loc[0] for loc in lokasjoner if loc[0]]))

# === API: Hent brukere med filtrering ===
@project_bp.route("/api/brukere")
@login_required
def api_hent_brukere():
    søk = request.args.get("q", "").lower()
    lokasjon = request.args.get("lokasjon", "")
    query = User.query
    if lokasjon:
        query = query.filter(User.location == lokasjon)
    if søk:
        query = query.filter(
            (User.first_name.ilike(f"%{søk}%")) |
            (User.last_name.ilike(f"%{søk}%"))
        )
    brukere = query.order_by(User.first_name, User.last_name).limit(20).all()
    return jsonify([
        {
            "id": u.id,
            "navn": f"{u.first_name} {u.last_name}",
            "rolle": u.role,
            "lokasjon": u.location
        } for u in brukere
    ])

@project_bp.route("/prosjekt/merkeskilt/lastned")
@login_required
def last_ned_merkeskilt():
    project_id = request.args.get("id", type=int)
    if not project_id:
        return "Mangler prosjekt-ID", 400

    project = Project.query.get(project_id)
    if not project:
        return "Prosjekt ikke funnet", 404

    merkeskilt = Merkskilt.query.filter_by(project_id=project_id).all()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Merkeskilt")

    headers = ["Komponent-ID", "Beskrivelse", "Himlingsskilt", "Stripsskilt"]
    header_row = []
    for h in headers:
        cell = WriteOnlyCell(ws, value=h)
        cell.font = Font(bold=True)
        header_row.append(cell)
    ws.append(header_row)

    # For å justere kolonnebredder:
    col_widths = [len(h) for h in headers]
    data_rows = []

    for rad in merkeskilt:
        row = []

        komponent_id = str(rad.komponent_id)
        if komponent_id.startswith("="):
            komponent_id = " " + komponent_id  # escape '='
        c1 = WriteOnlyCell(ws, value=komponent_id)
        c1.number_format = '@'
        row.append(c1)
        col_widths[0] = max(col_widths[0], len(komponent_id))

        beskrivelse = rad.beskrivelse or ""
        c2 = WriteOnlyCell(ws, value=beskrivelse)
        c2.number_format = '@'
        row.append(c2)
        col_widths[1] = max(col_widths[1], len(beskrivelse))

        himling = "Ja" if rad.aktiv else "Nei"
        c3 = WriteOnlyCell(ws, value=himling)
        c3.number_format = '@'
        row.append(c3)
        col_widths[2] = max(col_widths[2], len(himling))

        strips = "Ja" if rad.inaktiv else "Nei"
        c4 = WriteOnlyCell(ws, value=strips)
        c4.number_format = '@'
        row.append(c4)
        col_widths[3] = max(col_widths[3], len(strips))

        data_rows.append(row)

    # Skriv data
    for r in data_rows:
        ws.append(r)

    # Sett kolonnebredder
    for i, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width + 2  # liten buffer

    # Lagre og send fil
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filnavn = f"{project.project_name}_Merkeskilt.xlsx".replace(" ", "_")

    return send_file(
        output,
        as_attachment=True,
        download_name=filnavn,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




@project_bp.route("/prosjekt/merkeskilt/oppdater", methods=["POST"])
@login_required
def oppdater_merkeskilt_rad():
    data = request.get_json()
    rad = Merkskilt.query.get(data.get("id"))
    if not rad:
        return jsonify(success=False, error="Fant ikke rad.")

    rad.komponent_id = data.get("komponent_id", rad.komponent_id)
    rad.beskrivelse = data.get("beskrivelse", rad.beskrivelse)
    rad.aktiv = data.get("aktiv", rad.aktiv)
    rad.inaktiv = data.get("inaktiv", rad.inaktiv)
    db.session.commit()

    return jsonify(success=True)
    
@project_bp.route("/prosjekt/merkeskilt/slett/<int:rad_id>", methods=["POST"])
@login_required
def slett_merkeskilt_rad(rad_id):
    rad = Merkskilt.query.get(rad_id)
    if not rad:
        return jsonify(success=False, error="Fant ikke rad.")

    db.session.delete(rad)
    db.session.commit()
    return jsonify(success=True)