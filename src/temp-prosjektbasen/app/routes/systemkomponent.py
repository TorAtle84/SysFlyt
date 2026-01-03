import io
import json
import re
import fitz  # PyMuPDF
import docx
from flask import Blueprint, request, jsonify, send_file, Response, stream_with_context, render_template
from flask_login import login_required
from openpyxl.styles import Alignment
from openpyxl import Workbook, load_workbook
from pathlib import Path
from app.models.project import Project
from app.models.systembygging import Systembygging
from app.models.komponentopptelling import Komponentopptelling
from app.models.db import db

systemkomponent_bp = Blueprint("systemkomponent", __name__, url_prefix="/systemkomponent")

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_ROOT = BASE_DIR / "temp"
TEMP_ROOT.mkdir(exist_ok=True)

# ────────────────────────────────────────────────────────────────────────────────
# TFM_DICT (forkortet ikke – beholdt slik du la inn)
# ────────────────────────────────────────────────────────────────────────────────
TFM_DICT = {
    "AB": "Bjelke", "AD": "Dekker", "AE": "Elementer", "AF": "Fagverk", "AG": "Glassfelt", "AH": "Fundamenter",
    "AK": "Komplette konstruksjoner", "AL": "List - Beslistning", "AO": "Oppbygende - Utforende", "AP": "Plate",
    "AR": "Ramme - Oppheng", "AS": "Søyle", "AU": "Åpninger", "AV": "Vegg", "BA": "Armering Forsterkning",
    "BB": "Beskyttende - Stoppende", "BC": "Begrensende", "BF": "Fuging", "BG": "Pakning Tetning", "BI": "Isolasjon",
    "BP": "Avretningsmasse", "BS": "Spikerslag", "CB": "Balkong", "CC": "Baldakin", "CD": "Karnapp",
    "CG": "Rampe - Repos", "CK": "Komplett konstruksjon", "CM": "Kjemiske stoffer", "CO": "Kobling Overgang",
    "CP": "Pipe - Skostein", "CQ": "Festemateriell", "CR": "Rammeverk - Oppheng", "CT": "Trapp - Leider",
    "CX": "Tunnel - Bru (inne-ute)", "DB": "Dør med brannklasse", "DF": "Foldedør - Foldevegg", "DI": "Dør - innvendig",
    "DK": "Kjøretøyregulering", "DL": "Luke", "DP": "Port", "DT": "Dør - tilfluktsrom", "DU": "Dør - utvendig",
    "DV": "Vindu", "EB": "Overflatebekledning", "EC": "Overflatebehandling", "EH": "Himling",
    "FA": "Ventilert arbeidsplass", "FB": "Benk - Bord - Plate - Tavle", "FC": "Beslag", "FD": "Disk - Skranke",
    "FF": "Fryserom", "FH": "Hylle - Reol", "FI": "Kabinett", "FK": "Kjølerom - Svalrom",
    "FO": "Sittebenk - Sofa - Stol", "FR": "Rom", "FS": "Skap - Skuff", "FT": "Speil", "FV": "Vaskemaskin",
    "FX": "Krok - Knagg - Håndtak", "GA": "Automat - Maskin", "GB": "Benk - Bord - Plate - Tavle", "GD": "Dekontamator",
    "GE": "Autoklaver", "GF": "Fryseskap", "GG": "Gardiner - Forheng", "GH": "Hylle - Reol",
    "GK": "Kjøleskap - Kjøledisk", "GL": "Lås - Beslag", "GM": "Mattilbereding", "GN": "Nøkler",
    "GO": "Sofa - Sittebenk", "GP": "Stol", "GQ": "Seng - Liggebenk", "GS": "Skap - Skuffer",
    "GT": "Tørkeskap - Varmeskap", "GV": "Vaskemaskin", "GW": "Vekt", "GX": "Holder", "GY": "Avfallsbeholder",
    "HA": "Automobil - Bil", "HB": "Rullebord - Tralle", "HC": "Container - Vogn", "HM": "Maskin",
    "HS": "Rullestol", "HT": "Truck - kran", "HV": "Verktøy", "IB": "Brenner", "IC": "Solceller-solfangere",
    "ID": "Kjel for destruksjon", "IE": "Elektrokjel", "IF": "Kjel for fast-bio brensel", "IG": "Generator",
    "IK": "Kuldeaggregat", "IL": "Energibrønn", "IM": "Motor", "IO": "Oljekjel", "IP": "Gasskjel",
    "IT": "Trykkluftaggregat (enhet)", "IU": "Turbin", "IV": "Aggregatenhet", "JF": "Forsterker", "JK": "Kompressor",
    "JP": "Pumpe", "JQ": "Pumpe i VA-installasjoner", "JV": "Vifte", "JW": "Spesialvifte",
    "KA": "Aktuator", "KD": "Drivhjul - Drev", "KE": "Induktiv energioverføring", "KG": "Gjennomføring",
    "KH": "Transportenhet (hevende forflyttende)", "KJ": "Jordingskomponenter", "KK": "Kanal", "KM": "Mast - Antenne",
    "KN": "Nedløp", "KO": "Kraftoverføring", "KQ": "Rør - spesielt", "KR": "Rør - generelt", "KS": "Skinne - Bane - Spor",
    "KU": "Kombinert kabel", "KV": "Høyspenningskabel > 1000V", "KW": "Lavspenningskabel 50 til 1000V",
    "KX": "Lavspenningskabel < 50V", "KY": "Optisk kabel", "KZ": "Slange", "LB": "Varmeomformende med vifte",
    "LC": "Kjøleomformende med vifte", "LD": "Kjøleflater", "LE": "Kondensator", "LF": "Fordamper",
    "LG": "Gear clutch", "LH": "Varmeflate", "LI": "Varmeelement", "LK": "Kjøleomformende", "LL": "Lyskilde",
    "LN": "Likeretter", "LO": "Omformer", "LP": "Pens - Veksel - Sjalter", "LQ": "Vekselretter",
    "LR": "Frekvensomformer", "LS": "Strålevarme", "LU": "Luftfukter", "LV": "Varmeomformende",
    "LX": "Varmegjenvinner", "LZ": "Varmerkabel - Varmerør", "MA": "Absoluttfilter", "MB": "ABC-filter",
    "MC": "UV-filter", "ME": "Elektrostatiske filter", "MF": "Luftfilter", "MG": "Fettfilter", "MK": "Kondenspotte",
    "ML": "Luftutskiller", "MM": "Membran", "MO": "Utskiller", "MR": "Rist - Sil", "MS": "Syklon", "MT": "Tørke",
    "MU": "Filter for Lyd - Bilde - Frekvensutjevner", "MV": "Vannfilter", "MX": "Støyfilter",
    "NB": "Batteri - UPS", "NC": "Kondensator", "NI": "Informasjonslagring", "NK": "Kum", "NM": "Badekar - Basseng",
    "NO": "Åpen tank", "NT": "Tank med trykk", "NU": "Tank uten trykk", "NV": "Vekt Lodd", "NW": "Varmtvannsbereder",
    "NX": "Toalett", "NY": "Servant", "NZ": "Brannslukkingsapparat", "OA": "AV-maskiner", "OB": "Shuntgruppe",
    "OD": "Datamaskin", "OE": "Energimåler", "OF": "Systembærer", "OM": "Mottaker - Sender", "OP": "PBX",
    "OQ": "Dataprogramprogramvare", "OR": "Router Fordeler", "OS": "Sentralenhet Mikser i Lydsystem",
    "OT": "Telefonapparat", "OU": "Undersentral", "QB": "Belastningsvakt", "QD": "Differansetrykkvakt",
    "QE": "Elektrisk vern", "QF": "Strømningsvakt", "QG": "", "QH": "Fuktvakt", "QI": "", "QJ": "", "QK": "",
    "QL": "Lyddemper", "QM": "Mekanisk beskyttelse", "QN": "Nivåvakt", "QO": "Overtrykksventil", "QP": "Trykkvakt",
    "QQ": "Vibrasjonsvakt", "QR": "Rotasjonsvakt", "QS": "Strømvakt", "QT": "Temperaturvakt", "QU": "",
    "QV": "Sikkerhetsventil", "QW": "", "QX": "Solavskjerming", "QY": "Lynavleder", "QZ": "Brannvern",
    "RA": "AV-opptaker", "RB": "Bevegelse", "RC": "Seismometer", "RD": "Differansetrykkgiver",
    "RE": "Elektriske variabler", "RF": "Strømningsmåler", "RG": "Posisjon - Lengde", "RH": "Fuktighetsgiver",
    "RI": "Termometer", "RJ": "Fotocelle", "RK": "Kortleser", "RL": "", "RM": "Multifunksjonell Kombinert føler",
    "RN": "Nivågiver", "RO": "", "RP": "Trykkgiver", "RQ": "Manometer Trykkmåler", "RR": "Giver generelt",
    "RS": "Hastighetsmåler", "RT": "Temperaturgiver", "RU": "Ur", "RV": "Veieceller", "RW": "Virkningsgradsmåler",
    "RX": "Målepunkt", "RY": "Gassdetektor-Røykdetektor", "RZ": "Branndeteksjon",
    "SA": "Reguleringsventil manuell", "SB": "Reguleringsventil motorstyrt", "SC": "Stengeventil motorstyrt",
    "SD": "Alarmventil sprinkler", "SE": "Ekspansjonsventil", "SF": "Fraluftsventil",
    "SG": "Tilbakeslagsventil - Overtrykkspjeld", "SH": "Hurtigkobling", "SI": "Effektregulator",
    "SJ": "Jevntrykksventil", "SK": "Strømningsregulator - CAV", "SL": "", "SM": "Stengeventil manuell", "SN": "",
    "SO": "", "SP": "Trykkutjevningsventil", "SQ": "Strømningsregulator - VAV", "SR": "Reguleringsspjeld",
    "SS": "Stengespjeld", "ST": "Tilluftsventil", "SU": "Sugetrykksventil", "SV": "Strupeventil",
    "SW": "Plenumskammer", "SX": "Regulator", "SY": "", "SZ": "Brannspjeld - Røykspjeld",
    "UA": "Uttak alarm", "UB": "Blandebatteri", "UC": "", "UD": "Uttak data", "UE": "Uttak el", "UF": "Fellesuttak",
    "UG": "Uttak gass", "UH": "Høyttaler", "UI": "", "UJ": "Skriver", "UK": "Kontrollpanel - Tablå",
    "UL": "Uttak trykkluft", "UM": "Monitor - Display", "UN": "Nødbelysning", "UO": "Trommel",
    "UP": "Belysningsarmatur", "UQ": "", "UR": "Uttak radio", "US": "Stasjon", "UT": "Uttak telefon", "UU": "",
    "UV": "Uttak vann", "UW": "", "UX": "Koblingsboks", "UY": "Uttak antenne", "UZ": "Dyse - Spreder",
    "VA": "", "VB": "Bærelag", "VC": "", "VD": "Dekke", "VE": "", "VF": "", "VG": "Gress", "VH": "", "VI": "",
    "VJ": "", "VK": "Kantstein heller", "VL": "Masse", "VM": "Mekanisk beskyttelse", "VN": "", "VO": "",
    "VP": "Planterbuskertrær", "VQ": "", "VR": "", "VS": "Skilt", "VT": "", "VU": "", "VV": "", "VW": "", "VX": "",
    "VY": "", "VZ": "", "XA": "", "XB": "", "XC": "Kondensator", "XD": "Komp. for binærlogikk", "XE": "",
    "XF": "Komponenter for vern", "XG": "Komponenter for krafttilførsel", "XH": "Komponenter for signalering",
    "XI": "Potensiometer", "XJ": "", "XK": "Releer - Kontaktorer", "XL": "Induktiv komponenter", "XM": "Motor",
    "XN": "Integrerte kretser", "XO": "Urbryter - Timer", "XP": "Komponenter for måling og prøving",
    "XQ": "Effektbryter", "XR": "Motstand", "XS": "Bryter / Vender", "XT": "Transformator", "XU": "",
    "XV": "Halvlederkomponenter og elektronrør", "XW": "", "XX": "Rekkeklemmer - Samlesignal", "XY": "",
    "XZ": "Terminering og tilpasning", "YA": "", "YB": "", "YC": "", "YD": "", "YE": "", "YF": "", "YG": "",
    "YH": "", "YI": "", "YJ": "", "YK": "", "YL": "", "YM": "", "YN": "", "YO": "", "YP": "", "YQ": "", "YR": "",
    "YS": "", "YT": "", "YU": "", "YV": "", "YW": "", "YX": "", "YY": "", "YZ": "", "ZA": "", "ZB": "", "ZC": "",
    "ZD": "", "ZE": "", "ZF": "", "ZG": "", "ZH": "", "ZI": "", "ZJ": "", "ZK": "", "ZL": "", "ZM": "", "ZN": "",
    "ZO": "", "ZP": "", "ZQ": "", "ZR": "", "ZS": "", "ZT": "", "ZU": "", "ZV": "", "ZW": "", "ZX": "", "ZY": "",
    "ZZ": ""
}

# ────────────────────────────────────────────────────────────────────────────────
# UI
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/")
@login_required
def index():
    prosjekter = Project.query.order_by(Project.project_name.asc()).all()
    return render_template("systemkomponent.html", prosjekter=prosjekter)

# ────────────────────────────────────────────────────────────────────────────────
# Tekstekstraksjon
# ────────────────────────────────────────────────────────────────────────────────
def extract_text_from_bytes(filename: str, data: bytes) -> str:
    """
    Ekstraherer tekst fra bytes for PDF, DOCX, XLSX, TXT.
    Trygg å bruke under streaming (ingen avhengighet til åpne filhåndtak).
    """
    name = filename.lower()
    bio = io.BytesIO(data)

    if name.endswith(".pdf"):
        try:
            doc = fitz.open(stream=data, filetype="pdf")
            text = ""
            for page in doc:
                text += page.get_text()
            return text
        except Exception as e:
            print(f"Feil ved lesing av PDF {filename}: {e}")
            return ""

    elif name.endswith(".docx"):
        try:
            docx_file = docx.Document(bio)
            return "\n".join(p.text for p in docx_file.paragraphs)
        except Exception as e:
            print(f"Feil ved lesing av DOCX {filename}: {e}")
            return ""

    elif name.endswith(".xlsx"):
        try:
            wb = load_workbook(filename=bio, data_only=True)
            # Deaktiver aktive filtre
            for ws in wb.worksheets:
                if ws.auto_filter and ws.auto_filter.ref:
                    ws.auto_filter.ref = None
            text = ""
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = " ".join("" if c is None else str(c) for c in row)
                    text += line + "\n"
            return text
        except Exception as e:
            print(f"Feil ved lesing av XLSX {filename}: {e}")
            return ""

    elif name.endswith(".txt"):
        try:
            return data.decode(errors="ignore")
        except Exception as e:
            print(f"Feil ved lesing av TXT {filename}: {e}")
            return ""

    else:
        print(f"Ukjent filtype: {filename}. Hopper over.")
        return ""

# ────────────────────────────────────────────────────────────────────────────────
# Regex-bygging (mal → regex som fanger prefiks + innhold pr segment)
# ────────────────────────────────────────────────────────────────────────────────
def build_dynamic_regex(format_string: str) -> str:
    """
    Bygg regex fra en formatstreng ved å fange PREFIKS + INNHOLD for hvert segment.
    - {byggnr}   -> (?P<byggnr_prefix>\++)(?P<byggnr>[A-Za-z0-9]+)
    - {system}   -> (?:(?P<system_prefix>=)(?P<system>[^-\s%\n]+))?
    - {komponent}-> (?P<komponent_prefix>-)(?P<komponent>[^\s%\n]+)
    - {typekode} -> (?:(?P<typekode_prefix>%)(?P<typekode>\S+))?
    """
    parts = []
    i = 0
    while i < len(format_string):
        if format_string[i:].startswith("{byggnr}"):
            parts.append(r"(?P<byggnr_prefix>\++)(?P<byggnr>[A-Za-z0-9]+)")
            i += len("{byggnr}")
        elif format_string[i:].startswith("{system}"):
            parts.append(r"(?:(?P<system_prefix>=)(?P<system>[^-\s%\n]+))?")
            i += len("{system}")
        elif format_string[i:].startswith("{komponent}"):
            parts.append(r"(?P<komponent_prefix>-)(?P<komponent>[^\s%\n]+)")
            i += len("{komponent}")
        elif format_string[i:].startswith("{typekode}"):
            parts.append(r"(?:(?P<typekode_prefix>%)(?P<typekode>\S+))?")
            i += len("{typekode}")
        else:
            parts.append(re.escape(format_string[i]))
            i += 1
    return "".join(parts)

# ────────────────────────────────────────────────────────────────────────────────
# Hjelpere: system-ID, Excel-format
# ────────────────────────────────────────────────────────────────────────────────
def get_unique_system_id(system_string: str) -> str:
    """
    Returner "kanonisk" del av system-ID før ':' eller '*' eller '-'.
    """
    if not system_string:
        return ""
    m = re.search(r'[:*\-]', system_string)
    return system_string[:m.start()] if m else system_string

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_value_str = "" if cell.value is None else str(cell.value)
                if cell_value_str:
                    max_length = max(max_length, len(cell_value_str))
            except Exception:
                pass
        ws.column_dimensions[column].width = max(8, (max_length + 2))

from openpyxl.styles import Alignment

def apply_cell_format(cell):
    # Behold wrap – vi bruker Tekst-format separat på kolonne A via write_text_cell
    cell.alignment = Alignment(wrap_text=True)

def sanitize_for_excel_text(value: str) -> str:
    """
    Returner som tekst uten å legge inn apostrof.
    Viktig: Vi gjør IKKE lenger noe med ledende '=' – det håndteres av write_text_cell().
    """
    return "" if value is None else str(value)

def write_text_cell(ws, row_idx: int, col_idx: int, value):
    """
    Skriv en tekstcelle (Tekst-format) uten at Excel tolker '=' som formel.
    - Setter number_format='@' (Tekst)
    - Tvinger data_type='s'
    - Setter quotePrefix=True hvis verdien starter med '=' (Excel viser '=' uten å beregne)
    """
    from openpyxl.styles import Alignment

    text = "" if value is None else str(value)
    cell = ws.cell(row=row_idx, column=col_idx)

    # 1) Format som Tekst + wrap
    cell.number_format = "@"
    cell.alignment = Alignment(wrap_text=True)

    # 2) Marker som "quoted text" dersom den starter med '='
    if text.startswith("="):
        try:
            # Støttes i openpyxl og gjør at Excel håndterer dette som tekst uten å vise apostrof
            cell.quotePrefix = True
        except Exception:
            pass

    # 3) Sett som streng (etter quotePrefix, før verdi er ok også – men vi tvinger type etterpå)
    cell.value = text
    try:
        cell.data_type = "s"
    except Exception:
        pass

    return cell

def force_first_column_text(ws):
    """
    (Oppdatert) Sikrer at eksisterende celler i kolonne A står som Tekst,
    uten å legge inn apostrof. Brukes bare som "belt and braces" etterpå om ønskelig.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"
            try:
                cell.data_type = "s"
            except Exception:
                pass
            # Dersom noen verdier i A starter med '=', merk som quoted text
            v = "" if cell.value is None else str(cell.value)
            if v.startswith("="):
                try:
                    cell.quotePrefix = True
                except Exception:
                    pass

# ────────────────────────────────────────────────────────────────────────────────
# SYSTEMBYGGING (streaming)
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/systembygging", methods=["POST"])
def systembygging():
    """
    Lister komponent-treff (én rad per treff) basert på formatmal og kriterier.
    """
    files = request.files.getlist("files")
    mem_files = [{"filename": f.filename, "data": f.read()} for f in files]
    kriterier_str = request.form.get("kriterier", "")
    system_valg = [s.strip() for s in kriterier_str.split(",") if s.strip()]
    formatval = request.form.get("format", "")  # f.eks. "{byggnr}{system}{komponent}{typekode}"

    if not formatval:
        return Response(json.dumps({"error": "Mangler format"}), mimetype="application/json", status=400)

    regex_str = build_dynamic_regex(formatval)
    pattern = re.compile(regex_str)

    def generate():
        rows = []
        for mf in mem_files:
            yield json.dumps({"currentFile": mf["filename"]}) + "\n"
            text = extract_text_from_bytes(mf["filename"], mf["data"])
            if not text:
                continue

            for m in pattern.finditer(text):
                gr = m.groupdict()

                byggnr    = (gr.get("byggnr") or "").strip()
                system    = (gr.get("system") or "").strip()
                komponent = (gr.get("komponent") or "").strip()
                typekode  = (gr.get("typekode") or "").strip()

                # Komponent er obligatorisk
                if not komponent:
                    continue

                # Hvis malen inneholder {system}, krever vi at '=' faktisk var til stede i teksten
                if "{system}" in formatval and not gr.get("system_prefix"):
                    continue

                # Kriteriefilter: sjekk de to første tegnene i system-ID (når system finnes)
                if system_valg and system and system[:2] not in system_valg:
                    continue

                # Rekonstruer full_id med eksakt fanget prefiks
                full_id_parts = []
                i = 0
                while i < len(formatval):
                    if formatval[i:].startswith("{byggnr}"):
                        full_id_parts.append((gr.get("byggnr_prefix") or "") + byggnr)
                        i += len("{byggnr}")
                    elif formatval[i:].startswith("{system}"):
                        if gr.get("system_prefix"):
                            full_id_parts.append(gr["system_prefix"] + system)
                        i += len("{system}")
                    elif formatval[i:].startswith("{komponent}"):
                        full_id_parts.append((gr.get("komponent_prefix") or "-") + komponent)
                        i += len("{komponent}")
                    elif formatval[i:].startswith("{typekode}"):
                        if gr.get("typekode_prefix"):
                            full_id_parts.append(gr["typekode_prefix"] + typekode)
                        i += len("{typekode}")
                    else:
                        full_id_parts.append(formatval[i])
                        i += 1

                current_full_id = "".join(full_id_parts)

                actual_id = komponent.lstrip("-")
                tfm2 = actual_id[:2] if len(actual_id) >= 2 else ""
                desc = TFM_DICT.get(tfm2, "Ukjent")

                rows.append({
                    "full_id": current_full_id,
                    "desc": desc,
                    "system": system,
                    "component": actual_id,
                    "files": [mf["filename"]],
                    "unique_system_key": get_unique_system_id(system),
                })

        # Sorter stabilt for frontend
        rows_sorted = sorted(rows, key=lambda x: (x["system"], x["component"]))
        yield json.dumps({"rows": rows_sorted})

    return Response(stream_with_context(generate()), mimetype="text/plain")

# ────────────────────────────────────────────────────────────────────────────────
# KOMPONENTOPPTELLING (streaming)
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/komponentopptelling", methods=["POST"])
def komponentopptelling():
    """
    Teller forekomster av komponenter basert på TFM-prefix og samler dem.
    Skiller mellom med/uten system-ID i kilden.
    """
    files = request.files.getlist("files")
    mem_files = [{"filename": f.filename, "data": f.read()} for f in files]
    kriterier_str = request.form.get("kriterier", "")
    system_valg = [s.strip() for s in kriterier_str.split(",") if s.strip()]
    formatval = request.form.get("format", "")

    if not formatval:
        return Response(json.dumps({"error": "Mangler format"}), mimetype="application/json", status=400)

    regex_str = build_dynamic_regex(formatval)
    pattern = re.compile(regex_str)

    # Aggregater
    unique_component_with_system = {}
    unique_component_without_system = {}

    def generate():
        for mf in mem_files:
            yield json.dumps({"currentFile": mf["filename"]}) + "\n"
            text = extract_text_from_bytes(mf["filename"], mf["data"])
            if not text:
                continue

            for m in pattern.finditer(text):
                gr = m.groupdict()
                system = (gr.get("system") or "").strip()
                komponent = (gr.get("komponent") or "").strip()

                if "{komponent}" in formatval and not gr.get("komponent_prefix"):
                    continue
                if not komponent:
                    continue

                has_system = bool(gr.get("system_prefix") and system)

                actual_id = komponent.lstrip("-")
                tfm2 = actual_id[:2] if len(actual_id) >= 2 else ""

                # Kun gyldige TFM-prefiks (bokstaver og finnes i ordboka)
                if not tfm2.isalpha() or tfm2 not in TFM_DICT:
                    continue

                # Aggregeringsnøkkel: TFM + påfølgende bokstaver/streker før tall/annet
                mkey = re.match(r'([A-Za-z]{2}[A-Za-z\-]*)', actual_id)
                comp_key = mkey.group(1) if mkey else tfm2

                # Kriterier på system (hvis til stede)
                if system_valg and has_system and system[:2] not in system_valg:
                    continue

                target = unique_component_with_system if has_system else unique_component_without_system
                if comp_key not in target:
                    target[comp_key] = {"per_file": {}, "files": set()}

                target_entry = target[comp_key]
                fname = mf["filename"]

                target_entry["per_file"][fname] = target_entry["per_file"].get(fname, 0) + 1
                target_entry["files"].add(fname)

        # Flater ut for frontend
        rows_to_send = []
        for target_dict, has_system_flag in [
            (unique_component_with_system, "Ja"),
            (unique_component_without_system, "Nei"),
        ]:
            for agg_id in sorted(target_dict.keys()):
                data = target_dict[agg_id]
                desc = TFM_DICT.get(agg_id[:2] if len(agg_id) >= 2 else "", "Ukjent")
                rows_to_send.append({
                    "id": agg_id,
                    "desc": desc,
                    "per_file": data["per_file"],
                    "files": list(data["files"]),
                    "has_system": has_system_flag
                })

        # Stabil sortering
        rows_to_send.sort(key=lambda x: (x["has_system"], x["id"]))
        yield json.dumps({"rows": rows_to_send})

    return Response(stream_with_context(generate()), mimetype="text/plain")

# ────────────────────────────────────────────────────────────────────────────────
# GENERER EXCEL
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/excel", methods=["POST"])
def generate_excel():
    """
    Genererer en Excel-fil basert på aggregerte resultater.
    - Systembygging: ett ark per system + ett for manglende
    - Komponentopptelling: ett ark med summerte 'count'
    VIKTIG: Kolonne A i Systembygging-ark skrives som TEKST (number_format='@').
            Verdier som starter med '=' beholdes som ren tekst (ingen apostrof).
    """
    data = request.json or {}
    rows = data.get("rows", [])

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    is_system_building = bool(rows) and ("full_id" in rows[0])
    download_filename = "Systembygging.xlsx" if is_system_building else "Komponentopptelling.xlsx"

    if is_system_building:
        systems = {}
        missing_system_rows = []

        for row in rows:
            sys_val = row.get("system", "")
            uniq = get_unique_system_id(sys_val)
            if not sys_val:
                missing_system_rows.append(row)
            else:
                systems.setdefault(uniq, []).append(row)

        # Ett ark per system
        for sys_name, comp_rows in sorted(systems.items()):
            sheet_name = (sys_name or "Ukjent System")[:31]
            invalid = [":", "/", "\\", "?", "*", "[", "]"]
            for ch in invalid:
                sheet_name = sheet_name.replace(ch, "_")
            base = sheet_name
            n = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base[:28]}_{n}"
                n += 1

            ws = wb.create_sheet(title=sheet_name)

            # Header
            headers = ["Full ID", "Beskrivelse", "Filer"]
            ws.append(headers)
            ws.auto_filter.ref = ws.dimensions
            for c in ws[1]:
                apply_cell_format(c)

            # Rader
            for r in comp_rows:
                # Radindeks vi skal skrive til
                row_idx = ws.max_row + 1

                # Kol A: TEKST, tillat '=' først uten apostrof
                full_id = sanitize_for_excel_text(r["full_id"])
                write_text_cell(ws, row_idx, 1, full_id)

                # Kol B og C som vanlig tekst med wrap
                ws.cell(row=row_idx, column=2, value=r["desc"])
                ws.cell(row=row_idx, column=3, value=", ".join(r.get("files", [])))
                apply_cell_format(ws.cell(row=row_idx, column=2))
                apply_cell_format(ws.cell(row=row_idx, column=3))

            # Juster bredder
            adjust_column_widths(ws)

        # Ark for manglende system
        if missing_system_rows:
            sheet_name = "Systemnummer mangler"
            base = sheet_name
            n = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base[:28]}_{n}"
                n += 1

            ws = wb.create_sheet(title=sheet_name)

            headers = ["Full ID", "Beskrivelse", "Filer"]
            ws.append(headers)
            ws.auto_filter.ref = ws.dimensions
            for c in ws[1]:
                apply_cell_format(c)

            for r in missing_system_rows:
                row_idx = ws.max_row + 1
                full_id = sanitize_for_excel_text(r["full_id"])
                write_text_cell(ws, row_idx, 1, full_id)
                ws.cell(row=row_idx, column=2, value=r["desc"])
                ws.cell(row=row_idx, column=3, value=", ".join(r.get("files", [])))
                apply_cell_format(ws.cell(row=row_idx, column=2))
                apply_cell_format(ws.cell(row=row_idx, column=3))

            adjust_column_widths(ws)

    else:
        # Komponentopptelling: summer count fra per_file om nødvendig
        ws = wb.active or wb.create_sheet(title="Komponentopptelling")
        ws.title = "Komponentopptelling"

        headers = ["Med system?", "Komponent", "Beskrivelse", "Antall", "Filer"]
        ws.append(headers)
        ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            apply_cell_format(c)

        for r in rows:
            per_file = r.get("per_file", {})
            count = sum(per_file.values()) if isinstance(per_file, dict) else (r.get("count") or 0)
            row_idx = ws.max_row + 1

            # Kolonne B (Komponent) kan i prinsippet også starte med '=' i noen tilfeller.
            # Vi skriver den som tekst via helper for å være konsekvent og trygg.
            ws.cell(row=row_idx, column=1, value=r.get("has_system", ""))
            write_text_cell(ws, row_idx, 2, r.get("id", ""))

            ws.cell(row=row_idx, column=3, value=r.get("desc", ""))
            ws.cell(row=row_idx, column=4, value=count)
            ws.cell(row=row_idx, column=5, value=", ".join(r.get("files", [])))

            apply_cell_format(ws.cell(row=row_idx, column=1))
            apply_cell_format(ws.cell(row=row_idx, column=3))
            apply_cell_format(ws.cell(row=row_idx, column=4))
            apply_cell_format(ws.cell(row=row_idx, column=5))

        adjust_column_widths(ws)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=download_filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
# ────────────────────────────────────────────────────────────────────────────────
# SEND TIL PROSJEKT (DB)
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/send/systembygging", methods=["POST"])
@login_required
def send_systembygging():
    data = request.get_json()
    if not data:
        return "Ingen JSON-data", 400

    project_id = data.get("project_id")
    rows = data.get("rows", [])
    if not project_id or not rows:
        return "Mangler prosjekt eller data", 400

    Systembygging.query.filter_by(project_id=project_id).delete()

    for r in rows:
        db.session.add(Systembygging(
            project_id=project_id,
            full_id=r.get("full_id", ""),
            desc=r.get("desc", ""),
            files=", ".join(r.get("files", []))
        ))
    db.session.commit()
    return "OK", 200

@systemkomponent_bp.route("/send/komponentopptelling", methods=["POST"])
@login_required
def send_komponentopptelling():
    data = request.get_json()
    if not data:
        return "Ingen JSON-data", 400

    project_id = data.get("project_id")
    rows = data.get("rows", [])
    if not project_id or not rows:
        return "Mangler prosjekt eller data", 400

    Komponentopptelling.query.filter_by(project_id=project_id).delete()

    for r in rows:
        per_file = r.get("per_file", {})
        count = sum(per_file.values()) if isinstance(per_file, dict) else (r.get("count") or 0)

        db.session.add(Komponentopptelling(
            project_id=project_id,
            komponent=r.get("id", ""),
            desc=r.get("desc", ""),
            count=count,
            has_system=r.get("has_system") == "Ja",
            files=", ".join(r.get("files", []))
        ))

    db.session.commit()
    return "OK", 200

# ────────────────────────────────────────────────────────────────────────────────
# VISNING TABS
# ────────────────────────────────────────────────────────────────────────────────
@systemkomponent_bp.route("/tab/system/<int:project_id>")
@login_required
def vis_systembygging(project_id):
    prosjekt = Project.query.get_or_404(project_id)
    rader = Systembygging.query.filter_by(project_id=project_id).all()
    return render_template("partials/tab_system.html", prosjekt=prosjekt, rader=rader)

@systemkomponent_bp.route("/tab/komponent/<int:project_id>")
@login_required
def vis_komponentopptelling(project_id):
    prosjekt = Project.query.get_or_404(project_id)
    rader = Komponentopptelling.query.filter_by(project_id=project_id).all()
    return render_template("tab_komponent.html", prosjekt=prosjekt, rader=rader)